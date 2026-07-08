# Copyright (c) 2025, TFSS and contributors
# For license information, please see license.txt

import base64
import io

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import get_fullname, now_datetime, today, flt

# ---------------------------------------------------------------------------
# Single source of truth for the registration state machine
# ---------------------------------------------------------------------------
VALID_TRANSITIONS = {
    "Draft":                    ["Selected"],
    "Selected":                 ["Pending REGO"],
    "Pending REGO":             ["Pending FINO"],
    "Pending FINO":             ["Pending Registration"],
    "Pending Registration":     ["Pending Print & Scan"],
    "Pending Print & Scan":     ["Pending Residences"],
    "Pending Residences":       ["Pending IT"],
    "Pending IT":               ["Final Verification REGO"],
    "Final Verification REGO":  ["Completed"],
    "Completed":                ["Re-Open"],
    "Re-Open":                  ["Pending REGO"],
}

# Roles that are allowed to move INTO each target status
TRANSITION_ROLES = {
    "Pending REGO":             ["slcm_Student", "slcm_Registration User", "System Manager"],
    "Pending FINO":             ["slcm_REGO Officer", "System Manager"],
    "Pending Registration":     ["slcm_FINO Officer", "System Manager"],
    "Pending Print & Scan":     ["slcm_Registration Officer", "System Manager"],
    "Pending Residences":       ["slcm_Documentation Officer", "System Manager"],
    "Pending IT":               ["slcm_Hostel Admin", "System Manager"],
    "Final Verification REGO":  ["slcm_IT Admin", "System Manager"],
    "Completed":                ["slcm_Registration Officer", "System Manager"],
    "Re-Open":                  ["System Manager"],
}


# ---------------------------------------------------------------------------
# Fee structure helpers (module-level, reusable across methods and jobs)
# ---------------------------------------------------------------------------

def _get_valid_fee_structure_for_program(program):
    """Return the currently date-valid active Fee Structure for the given program, or None."""
    current_date = today()
    result = frappe.db.sql(
        """
        SELECT name, total_amount_for_indian AS total_amount, valid_from, valid_until
        FROM `tabFee Structure`
        WHERE program = %s
          AND status = 'Active'
          AND applicable = 'Student'
          AND valid_from <= %s
          AND (valid_until IS NULL OR valid_until >= %s)
        ORDER BY valid_from DESC, creation DESC
        LIMIT 1
        """,
        (program, current_date, current_date),
        as_dict=True,
    )
    return result[0] if result else None


def _resolve_program(programme):
    """Resolve a Cohort name (or bare Program name) to a Program name."""
    program = frappe.db.get_value("Batch", programme, "program")
    if not program and frappe.db.exists("Programme", programme):
        program = programme
    return program


def _calculate_discount(total_fee, applying_scholarship, scholarship_percentage, scholarship_amount):
    """Return scholarship discount amount based on student's scholarship fields."""
    scholarship_pct = flt(scholarship_percentage or 0)
    scholarship_amt = flt(scholarship_amount or 0)

    if applying_scholarship == "Yes" and scholarship_pct:
        return round((total_fee * scholarship_pct) / 100, 2)
    elif applying_scholarship == "Yes" and scholarship_amt:
        return min(scholarship_amt, total_fee)
    return 0.0


# ---------------------------------------------------------------------------
# Main DocType class
# ---------------------------------------------------------------------------

class StudentMaster(Document):
    def validate(self):
        self.validate_status_transition()

    def before_insert(self):
        """Auto-populate fee details from the currently valid Fee Structure on new record."""
        self._auto_fetch_fee_structure()
        if not self.current_year:
            self.current_year = "1"

    def before_save(self):
        """Track status and fee structure changes and append to audit history."""
        if self.is_new():
            return

        self._track_fee_structure_change()

        previous_status = frappe.db.get_value("Student Master", self.name, "registration_status")

        if previous_status == self.registration_status:
            return

        self.status_updated_by = frappe.session.user
        self.status_updated_on = now_datetime()

        confirmed_previous_state = None
        if previous_status and frappe.db.exists("Workflow State", previous_status):
            confirmed_previous_state = previous_status

        self.append("workflow_history", {
            "workflow_state":  self.registration_status,
            "previous_state":  confirmed_previous_state,
            "updated_by":      frappe.session.user,
            "updated_on":      now_datetime(),
            "remarks":         self.status_remarks,
        })

        frappe.get_doc({
            "doctype":          "Comment",
            "comment_type":     "Workflow",
            "reference_doctype": self.doctype,
            "reference_name":   self.name,
            "content": _("Status changed from {0} to {1} by {2}").format(
                previous_status or "Selected",
                self.registration_status,
                get_fullname(frappe.session.user),
            ),
        }).insert(ignore_permissions=True)

    def _track_fee_structure_change(self):
        """If fee_structure was changed manually, update fee fields and log to history."""
        prev_fs = frappe.db.get_value("Student Master", self.name, "fee_structure")
        if prev_fs == self.fee_structure or not self.fee_structure:
            return

        try:
            fs_doc = frappe.get_doc("Fee Structure", self.fee_structure)
        except Exception:
            return

        total_fee = flt(fs_doc.total_amount_for_indian or 0)
        if total_fee:
            discount = _calculate_discount(
                total_fee,
                self.applying_scholarship,
                self.scholarship_percentage,
                self.scholarship_amount,
            )
            net_fee = total_fee - discount
            self.total_program_fee  = total_fee
            self.discount_amount    = discount
            self.net_program_fee    = net_fee
            self.outstanding_balance = max(net_fee - flt(self.total_paid_amount or 0), 0)

        if fs_doc.instalment_enabled and fs_doc.max_instalments:
            self.number_of_instalments = fs_doc.max_instalments

        fs_label = fs_doc.fee_structure_name or self.fee_structure
        self.append("fee_structure_history", {
            "fee_structure":       self.fee_structure,
            "fee_structure_label": fs_label,
            "total_program_fee":   flt(fs_doc.total_amount_for_indian or 0),
            "valid_from":          fs_doc.valid_from,
            "valid_until":         fs_doc.valid_until,
            "applied_on":          now_datetime(),
            "applied_by":          frappe.session.user,
            "reason":              "Manual change by admin",
        })

    def on_update(self):
        """Sync derived fields and send completion email."""
        self._sync_active_statuses()
        self._handle_registration_email()
        _rebuild_fee_invoices(self)

    def _rebuild_fee_invoices_table(self):
        """Internal: rebuild fee_invoices child rows. Called from on_update and the API."""
        _rebuild_fee_invoices(self)

    def _sync_active_statuses(self):
        """Keep student_status and academic_status consistent with registration_status."""
        if self.registration_status == "Completed":
            if self.student_status != "Active":
                frappe.db.set_value("Student Master", self.name, "student_status", "Active")
            if self.academic_status != "Active":
                frappe.db.set_value("Student Master", self.name, "academic_status", "Active")

    def _handle_registration_email(self):
        doc_before_save = self.get_doc_before_save()
        previous_status = doc_before_save.registration_status if doc_before_save else None

        if self.registration_status == "Completed" and previous_status != "Completed":
            from slcm.slcm.utils.student_email import handle_registration_completion
            handle_registration_completion(self.name, frappe.session.user)

    def _auto_fetch_fee_structure(self):
        """Populate fee fields from the currently date-valid Fee Structure.

        Runs on before_insert. Skips if fee data is already provided (e.g. from
        the admission pipeline). Records the assignment in the history table.
        """
        if not self.programme:
            return
        if self.fee_structure or flt(self.total_program_fee):
            return

        program = _resolve_program(self.programme)
        if not program:
            return

        fs = _get_valid_fee_structure_for_program(program)
        if not fs:
            return

        total_fee = flt(fs.total_amount or 0)
        if not total_fee:
            return

        discount = _calculate_discount(
            total_fee,
            self.applying_scholarship,
            self.scholarship_percentage,
            self.scholarship_amount,
        )

        self.fee_structure = fs.name
        self.total_program_fee = total_fee
        self.discount_amount = discount
        self.net_program_fee = total_fee - discount
        self.outstanding_balance = max(self.net_program_fee - flt(self.total_paid_amount or 0), 0)

        fs_label = frappe.db.get_value("Fee Structure", fs.name, "fee_structure_name") or fs.name
        self.append("fee_structure_history", {
            "fee_structure":       fs.name,
            "fee_structure_label": fs_label,
            "total_program_fee":   total_fee,
            "valid_from":          fs.valid_from,
            "valid_until":         fs.valid_until,
            "applied_on":          now_datetime(),
            "applied_by":          frappe.session.user,
            "reason":              "Auto-assigned on student creation",
        })

    def validate_status_transition(self):
        """Enforce the registration workflow sequence."""
        if self.is_new():
            if not self.registration_status:
                self.registration_status = "Selected"
            return

        previous_status = frappe.db.get_value("Student Master", self.name, "registration_status")

        if previous_status == self.registration_status:
            return

        if "System Manager" not in frappe.get_roles():
            if previous_status in VALID_TRANSITIONS:
                if self.registration_status not in VALID_TRANSITIONS[previous_status]:
                    frappe.throw(_(
                        "Invalid status transition from {0} to {1}. Please follow the workflow sequence."
                    ).format(previous_status, self.registration_status))


# ---------------------------------------------------------------------------
# Hook called from hooks.py doc_events before_save
# ---------------------------------------------------------------------------
def before_save_hook(doc, method=None):
    doc.before_save()


# ---------------------------------------------------------------------------
# Fee Structure change trigger (called from Fee Structure on_update hook)
# ---------------------------------------------------------------------------
def on_fee_structure_update(doc, method=None):
    """Enqueue a background sync when a Student fee structure is saved with relevant changes."""
    if doc.applicable != "Student":
        return
    if doc.status != "Active":
        return

    changed = (
        doc.has_value_changed("status")
        or doc.has_value_changed("valid_from")
        or doc.has_value_changed("valid_until")
        or doc.has_value_changed("total_amount_for_indian")
    )
    if not changed:
        return

    frappe.enqueue(
        "slcm.slcm.doctype.student_master.student_master.sync_fee_structures_for_program",
        program=doc.program,
        queue="default",
        timeout=600,
        job_id=f"fee_sync_{doc.program}_{today()}",
    )
    frappe.msgprint(
        _("Fee structure change detected. Student fee data will be updated in the background."),
        indicator="blue",
        alert=True,
    )


# ---------------------------------------------------------------------------
# Sync helpers – used by both the background job and the daily scheduler
# ---------------------------------------------------------------------------

def _sync_single_student_fee(student_name):
    """Check and update the fee structure for one student. Returns True if updated."""
    student = frappe.get_doc("Student Master", student_name)

    if not student.programme:
        return False

    program = _resolve_program(student.programme)
    if not program:
        return False

    fs = _get_valid_fee_structure_for_program(program)
    if not fs:
        return False

    if student.fee_structure == fs.name:
        return False

    total_fee = flt(fs.total_amount or 0)
    if not total_fee:
        return False

    discount = _calculate_discount(
        total_fee,
        student.applying_scholarship,
        student.scholarship_percentage,
        student.scholarship_amount,
    )
    net_fee    = total_fee - discount
    paid       = flt(student.total_paid_amount or 0)
    outstanding = max(net_fee - paid, 0)

    fs_label = frappe.db.get_value("Fee Structure", fs.name, "fee_structure_name") or fs.name

    student.fee_structure      = fs.name
    student.total_program_fee  = total_fee
    student.discount_amount    = discount
    student.net_program_fee    = net_fee
    student.outstanding_balance = outstanding

    student.append("fee_structure_history", {
        "fee_structure":       fs.name,
        "fee_structure_label": fs_label,
        "total_program_fee":   total_fee,
        "valid_from":          fs.valid_from,
        "valid_until":         fs.valid_until,
        "applied_on":          now_datetime(),
        "applied_by":          "System",
        "reason":              "Auto-synced: new fee structure validity period is active",
    })

    student.save(ignore_permissions=True)
    return True


def sync_fee_structures_for_program(program):
    """Update fee structures for all active students in the given program.

    Called as a background job when a Fee Structure is saved.
    """
    students = frappe.get_all(
        "Student Master",
        filters={"student_status": "Active"},
        fields=["name", "programme"],
        limit=0,
    )

    updated = 0
    errors  = 0

    for s in students:
        if not s.programme:
            continue
        resolved = _resolve_program(s.programme)
        if resolved != program:
            continue

        try:
            if _sync_single_student_fee(s.name):
                updated += 1
        except Exception:
            frappe.log_error(
                frappe.get_traceback(),
                f"Fee Structure Sync Error: {s.name}",
            )
            errors += 1

    frappe.logger().info(
        f"Fee Structure Sync for program '{program}': {updated} updated, {errors} errors"
    )


def auto_sync_all_student_fee_structures():
    """Daily scheduler: update fee structures for every active student."""
    students = frappe.get_all(
        "Student Master",
        filters={"student_status": "Active"},
        fields=["name"],
        limit=0,
    )

    updated = 0
    errors  = 0

    for s in students:
        try:
            if _sync_single_student_fee(s.name):
                updated += 1
        except Exception:
            frappe.log_error(
                frappe.get_traceback(),
                f"Daily Fee Sync Error: {s.name}",
            )
            errors += 1

    if updated or errors:
        frappe.logger().info(
            f"Daily Fee Sync: {updated} updated, {errors} errors / {len(students)} total active students"
        )


# ---------------------------------------------------------------------------
# Whitelisted API
# ---------------------------------------------------------------------------
@frappe.whitelist()
def update_registration_status(student_id, new_status, remarks=None):
    """Update registration status with role-based and sequence validation."""
    user_roles = frappe.get_roles()
    is_admin = "System Manager" in user_roles or frappe.session.user == "Administrator"

    if not remarks or not remarks.strip():
        frappe.throw(_("Remarks is mandatory for status update"))

    student = frappe.get_doc("Student Master", student_id)
    current_status = student.registration_status or "Selected"

    if not is_admin:
        required_roles = TRANSITION_ROLES.get(new_status, [])
        if not any(role in user_roles for role in required_roles):
            frappe.throw(_(
                "You do not have permission to set status to {0}. Required roles: {1}"
            ).format(new_status, ", ".join(required_roles)))

        if current_status in VALID_TRANSITIONS:
            if new_status not in VALID_TRANSITIONS[current_status]:
                frappe.throw(_(
                    "Invalid status transition from {0} to {1}. Please follow the workflow sequence."
                ).format(current_status, new_status))

    _validate_transition_requirements(student, new_status)

    student.registration_status = new_status
    student.status_updated_by = frappe.session.user
    student.status_updated_on = now_datetime()
    student.status_remarks = remarks

    try:
        student.save(ignore_permissions=True)
        frappe.db.commit()
    except Exception as e:
        frappe.db.rollback()
        frappe.throw(_("Error updating status: {0}").format(str(e)))

    return {"status": "success", "message": _("Status updated to {0}").format(new_status)}


@frappe.whitelist()
def get_available_status_actions(student_id):
    """Return valid next actions for the current user and student status."""
    student = frappe.get_doc("Student Master", student_id)
    current_status = student.registration_status or "Selected"
    user_roles = frappe.get_roles()
    is_admin = "System Manager" in user_roles or frappe.session.user == "Administrator"

    ACTION_LABELS = {
        "Pending REGO":             "Submit for REGO",
        "Pending FINO":             "Approve Documents",
        "Pending Registration":     "Approve Finances",
        "Pending Print & Scan":     "Submit for Print & Scan",
        "Pending Residences":       "Upload Documents",
        "Pending IT":               "Allocate Room",
        "Final Verification REGO":  "Allocate Assets",
        "Completed":                "Complete Registration",
        "Re-Open":                  "Re-Open",
    }

    available_actions = []

    if is_admin:
        for state in VALID_TRANSITIONS:
            if state != current_status:
                available_actions.append({
                    "action":     f"Set to {state}",
                    "next_state": state,
                    "label":      f"Set to {state}",
                })
    else:
        next_states = VALID_TRANSITIONS.get(current_status, [])
        for next_state in next_states:
            required_roles = TRANSITION_ROLES.get(next_state, [])
            if any(role in user_roles for role in required_roles):
                available_actions.append({
                    "action":     ACTION_LABELS.get(next_state, next_state),
                    "next_state": next_state,
                    "label":      ACTION_LABELS.get(next_state, next_state),
                })

    return {"current_status": current_status, "available_actions": available_actions}


@frappe.whitelist()
def validate_new_enrollment(student_id):
    """Check whether a student is eligible to be enrolled."""
    try:
        student = frappe.get_doc("Student Master", student_id)
    except frappe.DoesNotExistError:
        return {"allowed": False, "message": "Student record not found."}

    if student.academic_status == "Inactive":
        return {"allowed": False, "message": "Student Academic Status is Inactive."}

    if student.student_status == "Inactive":
        return {"allowed": False, "message": "Student Status is Inactive."}

    if student.registration_status != "Completed":
        return {
            "allowed": False,
            "message": (
                f"Student Registration Status is '{student.registration_status}'. "
                "Must be 'Completed' to enroll."
            ),
        }

    if not student.programme:
        return {"allowed": False, "message": "Programme (Cohort) is not set in Student Master."}

    existing_enrollment = frappe.db.exists(
        "Student Enrollment",
        {"student": student.name, "cohort": student.programme, "docstatus": ["<", 2]},
    )

    if existing_enrollment:
        return {
            "allowed": False,
            "message": (
                f"Student is already enrolled in this Cohort ({student.programme}). "
                f"Enrollment ID: {existing_enrollment}"
            ),
        }

    return {"allowed": True}


@frappe.whitelist()
def bulk_student_enrollment(students):
    """Bulk-enroll a list of students."""
    import json

    if isinstance(students, str):
        students = json.loads(students)

    success = []
    failed  = []

    for student_id in students:
        validation = validate_new_enrollment(student_id)

        if not validation.get("allowed"):
            failed.append({"student": student_id, "reason": validation.get("message")})
            continue

        try:
            student = frappe.get_doc("Student Master", student_id)

            new_enrollment = frappe.get_doc({
                "doctype":      "Student Enrollment",
                "student":      student.name,
                "student_name": " ".join(filter(None, [
                    student.first_name, student.middle_name, student.last_name
                ])),
                "cohort":        student.programme,
                "batch_year_ref": frappe.db.get_value("Batch", student.programme, "section") or "",
                "academic_year": student.academic_year,
                "status":        "Enrolled",
                "enrollment_date": frappe.utils.today(),
            })

            new_enrollment.insert()
            success.append(student_id)

        except Exception as e:
            frappe.log_error(message=str(e), title="Bulk Enrollment Error")
            failed.append({"student": student_id, "reason": str(e)})

    return {"success": success, "failed": failed}


# ---------------------------------------------------------------------------
# Internal transition requirement validator
# ---------------------------------------------------------------------------
def _validate_transition_requirements(student, new_status):
    """Raise frappe.throw if mandatory fields/docs are missing for the transition."""

    if new_status == "Pending FINO":
        required_docs = ["aadhaar_card", "pan_card", "std_x_marksheet", "passport_size_photo"]
        missing = [d for d in required_docs if not student.get(d)]
        if missing:
            labels = [frappe.get_meta("Student Master").get_label(d) for d in missing]
            frappe.throw(_(
                "Cannot move to Pending FINO. Missing documents: {0}"
            ).format(", ".join(labels)))

    elif new_status == "Pending Registration":
        if student.fee_payment_status not in ["Paid", "Partially Paid"]:
            frappe.throw(_(
                "Cannot move to Pending Registration. "
                "Fee Payment Status must be 'Paid' or 'Partially Paid'. Current: {0}"
            ).format(student.fee_payment_status))

    elif new_status == "Pending Print & Scan":
        required_fields = ["first_name", "last_name", "dob", "gender", "email", "phone", "programme", "programme_of_study"]
        missing = [f for f in required_fields if not student.get(f)]
        if missing:
            labels = [frappe.get_meta("Student Master").get_label(f) for f in missing]
            frappe.throw(_(
                "Cannot move to Pending Print & Scan. Missing mandatory details: {0}"
            ).format(", ".join(labels)))

    elif new_status == "Pending Residences":
        if not student.id_card_issued:
            frappe.throw(_("Cannot move to Pending Residences. ID Card must be issued."))
        if not student.aadhaar_verified:
            frappe.throw(_("Cannot move to Pending Residences. Aadhaar must be verified."))

    elif new_status == "Pending IT":
        if student.is_hosteller:
            if not student.hostel_room:
                frappe.throw(_("Cannot move to Pending IT. Hostel Room must be allocated for hostellers."))
            if not student.keys_handed_over:
                frappe.throw(_("Cannot move to Pending IT. Keys must be handed over."))

    elif new_status == "Final Verification REGO":
        if not student.official_email_id:
            frappe.throw(_("Cannot move to Final Verification. Official Email ID must be set."))

    elif new_status == "Completed":
        doc_checks = {
            "aadhaar_card":        "aadhaar_verified",
            "pan_card":            "pan_verified",
            "offer_letter":        "offer_letter_verified",
            "student_declaration": "student_declaration_verified",
        }
        for doc_field, verified_field in doc_checks.items():
            if not student.get(doc_field):
                label = frappe.get_meta("Student Master").get_label(doc_field)
                frappe.throw(_("Cannot complete registration. {0} is not uploaded.").format(label))
            if not student.get(verified_field):
                label = frappe.get_meta("Student Master").get_label(doc_field)
                frappe.throw(_("Cannot complete registration. {0} is not verified.").format(label))

        if not student.id_card_issued:
            frappe.throw(_("Cannot complete registration. ID Card must be issued."))
        if not student.official_email_id:
            frappe.throw(_("Cannot complete registration. Official Email ID must be set."))


@frappe.whitelist()
def fetch_program_fee_details(programme):
    """Return fee details from the currently date-valid Student Fee Structure for the given cohort."""
    if not programme:
        return None

    program = _resolve_program(programme)
    if not program:
        return None

    fs = _get_valid_fee_structure_for_program(program)
    if not fs:
        return None

    return {
        "fee_structure":      fs.name,
        "fee_structure_name": frappe.db.get_value("Fee Structure", fs.name, "fee_structure_name") or fs.name,
        "total_program_fee":  flt(fs.total_amount or 0),
        "valid_from":         str(fs.valid_from or ""),
        "valid_until":        str(fs.valid_until or ""),
    }


# ---------------------------------------------------------------------------
# Fee Invoice child-table helpers
# ---------------------------------------------------------------------------

def _rebuild_fee_invoices(sm_doc):
    """Rebuild the fee_invoices child table rows from live Fee Invoice records.

    Accepts a StudentMaster document instance. Uses db_update() so it never
    triggers validate/on_update loops.
    """
    try:
        invoices = frappe.get_all(
            "Fee Invoice",
            filters={"student": sm_doc.name},
            fields=[
                "name", "academic_term", "invoice_date", "due_date",
                "final_payable_amount", "paid_amount", "outstanding_amount", "status",
            ],
            order_by="invoice_date desc, creation desc",
            ignore_permissions=True,
        )

        sm_doc.set("fee_invoices", [])
        for inv in invoices:
            sm_doc.append("fee_invoices", {
                "invoice":            inv.name,
                "academic_term":      inv.academic_term or "",
                "invoice_date":       inv.invoice_date,
                "due_date":           inv.due_date,
                "net_payable":        flt(inv.final_payable_amount or 0),
                "paid_amount":        flt(inv.paid_amount or 0),
                "outstanding_amount": max(flt(inv.outstanding_amount or 0), 0),
                "status":             inv.status or "Unpaid",
            })

        sm_doc.db_update()
        frappe.db.commit()
        return len(invoices)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "_rebuild_fee_invoices failed")
        return 0


def _enrollments_by_recency(student_name):
    """Return the student's enrollments ordered by actual academic recency:
    the enrollment still marked "Enrolled" first, then by the linked
    Academic Year's start date (most recent first) — NOT by row creation
    time, since enrollments (e.g. for a past term) may be entered into the
    system after a later enrollment already exists."""
    enrollments = frappe.get_all(
        "Student Enrollment",
        filters={"student": student_name, "docstatus": ["<", 2]},
        fields=["name", "academic_year", "term_name", "status", "creation"],
    )
    if not enrollments:
        return []

    ay_names = {e.academic_year for e in enrollments if e.academic_year}
    ay_start = {}
    if ay_names:
        for ay in frappe.get_all(
            "Academic Year",
            filters={"name": ["in", list(ay_names)]},
            fields=["name", "year_start_date"],
        ):
            ay_start[ay.name] = ay.year_start_date

    enrollments.sort(
        key=lambda e: (
            1 if e.status == "Enrolled" else 0,
            ay_start.get(e.academic_year) or e.creation,
        ),
        reverse=True,
    )
    return enrollments


@frappe.whitelist()
def get_academic_progress_list(student_name):
    """Return all enrollments for a student (current first, then most
    recent past terms) for the Current/Previous Academic Progress selector."""
    if not frappe.db.exists("Student Master", student_name):
        frappe.throw(_("Student Master not found: {0}").format(student_name))

    enrollments = _enrollments_by_recency(student_name)

    ay_names = {e.academic_year for e in enrollments if e.academic_year}
    ay_map = {}
    if ay_names:
        for ay in frappe.get_all(
            "Academic Year",
            filters={"name": ["in", list(ay_names)]},
            fields=["name", "academic_year_name"],
        ):
            ay_map[ay.name] = ay.academic_year_name

    return [
        {
            "name": e.name,
            "academic_year": e.academic_year or "",
            "ay_name": ay_map.get(e.academic_year) or e.academic_year or "",
            "term_name": e.term_name or "",
            "status": e.status or "",
            "is_current": idx == 0,
        }
        for idx, e in enumerate(enrollments)
    ]


@frappe.whitelist()
def get_academic_progress(student_name, enrollment_name=None):
    """Return academic year, term, semester and enrolled courses for a student.

    By default returns the most recent (current) enrollment. Pass
    `enrollment_name` to fetch a specific past enrollment instead, for the
    "Previous Academic Progress" view.

    Also checks the active Promotion Policy to compute promotion eligibility
    (promotion eligibility always reflects the student's current standing,
    regardless of which enrollment is being viewed).
    """
    if not frappe.db.exists("Student Master", student_name):
        frappe.throw(_("Student Master not found: {0}").format(student_name))

    sm = frappe.db.get_value(
        "Student Master",
        student_name,
        ["programme", "current_year", "academic_term", "current_cgpa", "attendance_status",
         "first_name", "last_name"],
        as_dict=True,
    )
    batch_year = frappe.db.get_value("Batch", sm.programme, "section") if sm.programme else ""

    # ── Enrollment: a specific one if requested, else the current one ───────
    if enrollment_name:
        enrollment = frappe.db.get_value(
            "Student Enrollment",
            {"student": student_name, "docstatus": ["<", 2], "name": enrollment_name},
            [
                "name", "academic_year", "term_name", "status",
                "program", "cohort", "batch_year_ref", "enrollment_date",
            ],
            as_dict=True,
        )
    else:
        recent = _enrollments_by_recency(student_name)
        enrollment = None
        if recent:
            enrollment = frappe.db.get_value(
                "Student Enrollment",
                recent[0].name,
                [
                    "name", "academic_year", "term_name", "status",
                    "program", "cohort", "batch_year_ref", "enrollment_date",
                ],
                as_dict=True,
            )

    result = {
        "student_name": student_name,
        "current_year": sm.current_year or "",
        "current_term": sm.academic_term or "",
        "current_cgpa": flt(sm.current_cgpa or 0),
        "attendance_status": sm.attendance_status or "",
        "batch_year": batch_year or "",
        "enrollment": None,
        "courses": [],
        "promotion": None,
    }

    if not enrollment:
        return result

    # Academic year details
    ay_doc = None
    if enrollment.academic_year:
        ay_doc = frappe.db.get_value(
            "Academic Year",
            enrollment.academic_year,
            ["academic_year_name", "academic_system", "year_start_date", "year_end_date", "status"],
            as_dict=True,
        )

    # Academic term details
    at_doc = None
    if enrollment.term_name:
        at_doc = frappe.db.get_value(
            "Academic Term",
            enrollment.term_name,
            ["term_name", "sequence", "term_start_date", "term_end_date", "status"],
            as_dict=True,
        )

    # Cohort / Section details
    cohort_doc = None
    if enrollment.cohort:
        cohort_doc = frappe.db.get_value(
            "Batch",
            enrollment.cohort,
            ["cohort_name", "cohort_code", "term_year", "current_year", "status"],
            as_dict=True,
        )

    # Program details
    program_name = ""
    if enrollment.program:
        program_name = frappe.db.get_value("Programme", enrollment.program, "program_name") or enrollment.program

    # Faculty advisor name
    faculty_advisor_name = ""
    if enrollment.faculty_advisor:
        faculty_advisor_name = frappe.db.get_value(
            "Faculty", enrollment.faculty_advisor, "faculty_name"
        ) or enrollment.faculty_advisor

    result["enrollment"] = {
        "name":                 enrollment.name,
        "academic_year":        enrollment.academic_year or "",
        "ay_name":              ay_doc.academic_year_name if ay_doc else (enrollment.academic_year or ""),
        "ay_system":            ay_doc.academic_system if ay_doc else "",
        "ay_start":             str(ay_doc.year_start_date) if ay_doc and ay_doc.year_start_date else "",
        "ay_end":               str(ay_doc.year_end_date) if ay_doc and ay_doc.year_end_date else "",
        "ay_status":            ay_doc.status if ay_doc else "",
        "term_name":            enrollment.term_name or "",
        "term_sequence":        at_doc.sequence if at_doc else "",
        "term_start":           str(at_doc.term_start_date) if at_doc and at_doc.term_start_date else "",
        "term_end":             str(at_doc.term_end_date) if at_doc and at_doc.term_end_date else "",
        "term_status":          at_doc.status if at_doc else "",
        "status":               enrollment.status or "",
        "program":              enrollment.program or "",
        "program_name":         program_name,
        "cohort":               enrollment.cohort or "",
        "cohort_name":          cohort_doc.cohort_name if cohort_doc else (enrollment.cohort or ""),
        "cohort_code":          cohort_doc.cohort_code if cohort_doc else "",
        "cohort_term_year":     cohort_doc.term_year if cohort_doc else "",
        "cohort_status":        cohort_doc.status if cohort_doc else "",
        "batch_year":           enrollment.batch_year_ref or "",
        "faculty_advisor":      enrollment.faculty_advisor or "",
        "faculty_advisor_name": faculty_advisor_name,
        "enrollment_date":      str(enrollment.enrollment_date) if enrollment.enrollment_date else "",
    }

    # ── Enrolled courses ─────────────────────────────────────────────────────
    sec_rows = frappe.get_all(
        "Student Enrollment Course",
        filters={"parent": enrollment.name, "parenttype": "Student Enrollment"},
        fields=["course_offering", "course", "course_type", "credits", "status", "grade"],
        order_by="idx asc",
    )
    co_names = [r.course_offering for r in sec_rows if r.course_offering]
    co_map = {}
    if co_names:
        for co in frappe.get_all(
            "Course Offering",
            filters={"name": ["in", co_names]},
            fields=["name", "course_name", "faculty", "credit_value"],
            ignore_permissions=True,
        ):
            co_map[co.name] = co

    # ── Attendance percentage per course (from Attendance Summary) ──────────
    att_map = {}
    for a in frappe.get_all(
        "Attendance Summary",
        filters={"student": student_name},
        fields=["course_offering", "course", "attendance_percentage"],
        ignore_permissions=True,
    ):
        if a.course_offering:
            att_map[a.course_offering] = a.attendance_percentage
        if a.course:
            att_map.setdefault(a.course, a.attendance_percentage)

    courses = []
    for r in sec_rows:
        co = co_map.get(r.course_offering) or frappe._dict()
        att_pct = att_map.get(r.course_offering)
        if att_pct is None:
            att_pct = att_map.get(r.course)
        courses.append({
            "course_offering": r.course_offering or "",
            "course": r.course or "",
            "course_name": co.get("course_name") or r.course or "",
            "course_type": r.course_type or "",
            "course_status": r.status or "Enrolled",
            "credit_value": co.get("credit_value") or r.credits or 0,
            "faculty": co.get("faculty") or "",
            "grade": r.grade or "",
            "attendance_percentage": att_pct if att_pct is not None else None,
        })
    result["courses"] = courses

    # ── Promotion Policy check ────────────────────────────────────────────────
    if enrollment.program and enrollment.academic_year:
        policy = frappe.db.get_value(
            "Promotion Policy",
            {
                "program": enrollment.program,
                "academic_year": enrollment.academic_year,
                "status": "Active",
            },
            [
                "name", "from_year", "to_year",
                "enable_cgpa_check", "min_cgpa",
                "enable_backlog_check", "max_backlogs_allowed",
                "enable_attendance_check", "min_attendance_percent",
                "conditional_promotion_action",
            ],
            order_by="creation desc",
            as_dict=True,
        )

        if policy:
            cgpa_pass = True
            backlog_pass = True
            attendance_pass = True

            student_cgpa = flt(sm.current_cgpa or 0)

            # Count backlogs from promotion records
            backlog_count = frappe.db.count(
                "Student Promotion",
                {"student": student_name, "promotion_status": "Not Promoted"},
            )

            # Parse attendance percent
            attendance_str = sm.attendance_status or "0"
            try:
                attendance_pct = flt(attendance_str.replace("%", "").strip())
            except Exception:
                attendance_pct = 0.0

            if policy.enable_cgpa_check:
                cgpa_pass = student_cgpa >= flt(policy.min_cgpa or 0)

            if policy.enable_backlog_check:
                backlog_pass = backlog_count <= int(policy.max_backlogs_allowed or 0)

            if policy.enable_attendance_check:
                attendance_pass = attendance_pct >= flt(policy.min_attendance_percent or 0)

            all_pass = cgpa_pass and backlog_pass and attendance_pass

            result["promotion"] = {
                "policy_name":        policy.name,
                "from_year":          policy.from_year,
                "to_year":            policy.to_year,
                "cgpa_check":         bool(policy.enable_cgpa_check),
                "min_cgpa":           flt(policy.min_cgpa or 0),
                "student_cgpa":       student_cgpa,
                "cgpa_pass":          cgpa_pass,
                "backlog_check":      bool(policy.enable_backlog_check),
                "max_backlogs":       int(policy.max_backlogs_allowed or 0),
                "backlog_count":      backlog_count,
                "backlog_pass":       backlog_pass,
                "attendance_check":   bool(policy.enable_attendance_check),
                "min_attendance":     flt(policy.min_attendance_percent or 0),
                "attendance_pct":     attendance_pct,
                "attendance_pass":    attendance_pass,
                "eligible":           all_pass,
                "conditional_action": policy.conditional_promotion_action or "",
            }

    return result


@frappe.whitelist()
def send_parent_login_invite(student_name):
    if not frappe.db.exists("Student Master", student_name):
        frappe.throw(_("Student Master not found: {0}").format(student_name))

    sm = frappe.get_doc("Student Master", student_name, ignore_permissions=True)
    full_student_name = f"{sm.first_name} {sm.last_name or ''}".strip()

    parents = sm.get("parents") or []
    if not parents:
        frappe.throw(_("No parent records found on this Student Master."))

    results = []
    for p in parents:
        if not p.email:
            results.append({"name": f"{p.first_name} {p.last_name or ''}".strip(), "status": "no_email"})
            continue
        from slcm.slcm.doctype.parent_login_invite_tool.parent_login_invite_tool import _create_parent_user_and_invite
        parent_full = f"{p.first_name} {p.last_name or ''}".strip()
        already = frappe.db.exists("User", p.email)
        _create_parent_user_and_invite(p.email, parent_full, full_student_name)
        results.append({
            "name": parent_full,
            "email": p.email,
            "status": "existing" if already else "invited",
        })

    return results


@frappe.whitelist()
def get_fee_structure_details(fee_structure):
    """Return key fields from a Fee Structure for auto-population in the Student Master form."""
    if not fee_structure or not frappe.db.exists("Fee Structure", fee_structure):
        return None

    fs = frappe.get_doc("Fee Structure", fee_structure)
    components = []
    for c in (fs.fee_components_for_indian or []):
        components.append({
            "component_name": c.component_name,
            "amount":         flt(c.amount or 0),
            "total_amount":   flt(c.total_amount or c.amount or 0),
        })

    return {
        "fee_structure":       fs.name,
        "fee_structure_name":  fs.fee_structure_name or fs.name,
        "total_amount":        flt(fs.total_amount_for_indian or 0),
        "valid_from":          str(fs.valid_from or ""),
        "valid_until":         str(fs.valid_until or ""),
        "status":              fs.status,
        "instalment_enabled":  int(fs.instalment_enabled or 0),
        "max_instalments":     int(fs.max_instalments or 0),
        "components":          components,
    }


@frappe.whitelist()
def change_fee_structure_admin(student_name, new_fee_structure, reason):
    """Admin-only API: change a student's fee structure with audit trail.

    Updates all derived fee fields and appends a history entry with the given reason.
    """
    user_roles = frappe.get_roles()
    allowed_roles = ["System Manager", "slcm_FINO Officer", "Administrator"]
    if not any(r in user_roles for r in allowed_roles) and frappe.session.user != "Administrator":
        frappe.throw(_("You do not have permission to change the Fee Structure."))

    if not frappe.db.exists("Student Master", student_name):
        frappe.throw(_("Student Master not found: {0}").format(student_name))
    if not frappe.db.exists("Fee Structure", new_fee_structure):
        frappe.throw(_("Fee Structure not found: {0}").format(new_fee_structure))

    reason = (reason or "").strip()
    if not reason:
        frappe.throw(_("Reason is mandatory when changing the Fee Structure."))

    sm = frappe.get_doc("Student Master", student_name, ignore_permissions=True)
    fs = frappe.get_doc("Fee Structure", new_fee_structure)

    total_fee = flt(fs.total_amount_for_indian or 0)
    discount = _calculate_discount(
        total_fee,
        sm.applying_scholarship,
        sm.scholarship_percentage,
        sm.scholarship_amount,
    )
    net_fee     = total_fee - discount
    outstanding = max(net_fee - flt(sm.total_paid_amount or 0), 0)

    sm.fee_structure       = new_fee_structure
    sm.total_program_fee   = total_fee
    sm.discount_amount     = discount
    sm.net_program_fee     = net_fee
    sm.outstanding_balance = outstanding

    if fs.instalment_enabled and fs.max_instalments:
        sm.number_of_instalments = fs.max_instalments

    fs_label = fs.fee_structure_name or new_fee_structure
    sm.append("fee_structure_history", {
        "fee_structure":       new_fee_structure,
        "fee_structure_label": fs_label,
        "total_program_fee":   total_fee,
        "valid_from":          fs.valid_from,
        "valid_until":         fs.valid_until,
        "applied_on":          now_datetime(),
        "applied_by":          frappe.session.user,
        "reason":              reason,
    })

    sm.save(ignore_permissions=True)
    frappe.db.commit()

    _append_payment_log(
        student_name,
        "Fee Structure Changed",
        amount=total_fee,
        from_status=frappe.db.get_value("Student Master", student_name, "fee_payment_status") or "",
        to_status="",
        remarks=f"Fee structure changed to {fs_label}. Reason: {reason}",
    )

    return {
        "status":              "success",
        "fee_structure":       new_fee_structure,
        "fee_structure_name":  fs_label,
        "total_program_fee":   total_fee,
        "net_program_fee":     net_fee,
        "valid_from":          str(fs.valid_from or ""),
        "valid_until":         str(fs.valid_until or ""),
    }


@frappe.whitelist()
def sync_fee_invoices(student_name):
    """Module-level whitelisted function — called from the JS button.

    Rebuilds the fee_invoices child table for the given Student Master and
    returns the count of synced rows.
    """
    if not frappe.db.exists("Student Master", student_name):
        frappe.throw(_("Student Master not found: {0}").format(student_name))

    sm_doc = frappe.get_doc("Student Master", student_name, ignore_permissions=True)
    count  = _rebuild_fee_invoices(sm_doc)
    return {"synced": count}



@frappe.whitelist()
def get_fee_demand_receipt(fee_demand_name):
    """Return the Fee Receipt name linked to a paid Fee Demand.

    Queries the Fee Payment Demand Row child table (no public permissions) with
    ignore_permissions so admin users can always retrieve the receipt link.
    """
    row = frappe.db.get_value(
        "Fee Payment Demand Row",
        {"fee_demand": fee_demand_name},
        "parent",
        order_by="creation desc",
    )
    if not row:
        return None
    receipt = frappe.db.get_value("Fee Payment", row, "receipt")
    if not receipt:
        return None
    return {"receipt": receipt, "fee_demand": fee_demand_name}


# ── Payment Log helper ────────────────────────────────────────────────────────

def _append_payment_log(student_name, event_type, **kwargs):
    """Insert a Student Fee Payment Log row directly, bypassing SM validate.

    Accepts all audit fields:
        student_name, event_type, amount, currency, invoice,
        payment_mode, payment_method, razorpay_payment_id, razorpay_order_id,
        transaction_id, triggered_by, attempt_type, retry_count,
        webhook_status, from_status, to_status, ip_address,
        gateway_response, error_message, failure_reason, remarks, timestamp
    """
    import json as _json

    # Determine attempt type and retry count automatically when not supplied
    attempt_type = kwargs.get("attempt_type") or ""
    if not attempt_type:
        if event_type == "Payment Initiated":
            prev_count = frappe.db.count(
                "Student Fee Payment Log",
                filters={
                    "parent": student_name,
                    "parenttype": "Student Master",
                    "event_type": "Payment Initiated",
                },
            )
            attempt_type = "Retry" if prev_count > 0 else "Initial Attempt"
        elif event_type == "Webhook Received":
            attempt_type = "Webhook Update"
        elif event_type == "Refunded":
            attempt_type = "Refund"

    retry_count = kwargs.get("retry_count")
    if retry_count is None:
        retry_count = max(
            frappe.db.count(
                "Student Fee Payment Log",
                filters={
                    "parent": student_name,
                    "parenttype": "Student Master",
                    "event_type": "Payment Initiated",
                },
            ) - 1,
            0,
        )

    # Safely serialise gateway_response to string
    gw_resp = kwargs.get("gateway_response") or ""
    if gw_resp and not isinstance(gw_resp, str):
        try:
            gw_resp = _json.dumps(gw_resp, indent=2)
        except Exception:
            gw_resp = str(gw_resp)

    try:
        row = frappe.get_doc({
            "doctype":              "Student Fee Payment Log",
            "parent":               student_name,
            "parenttype":           "Student Master",
            "parentfield":          "fee_payment_log",
            "event_type":           event_type,
            "timestamp":            kwargs.get("timestamp") or now_datetime(),
            "amount":               flt(kwargs.get("amount") or 0),
            "currency":             kwargs.get("currency") or "INR",
            "invoice":              kwargs.get("invoice") or "",
            "fee_demand":           kwargs.get("fee_demand") or "",
            "payment_mode":         kwargs.get("payment_mode") or "",
            "payment_method":       kwargs.get("payment_method") or "",
            "razorpay_payment_id":  kwargs.get("razorpay_payment_id") or "",
            "razorpay_order_id":    kwargs.get("razorpay_order_id") or "",
            "transaction_id":       kwargs.get("transaction_id") or "",
            "triggered_by":         kwargs.get("triggered_by") or frappe.session.user,
            "paid_by_role":         kwargs.get("paid_by_role") or "",
            "paid_by_name":         kwargs.get("paid_by_name") or "",
            "attempt_type":         attempt_type,
            "retry_count":          retry_count,
            "webhook_status":       kwargs.get("webhook_status") or "Not Applicable",
            "from_status":          kwargs.get("from_status") or "",
            "to_status":            kwargs.get("to_status") or "",
            "ip_address":           kwargs.get("ip_address") or _get_request_ip(),
            "gateway_response":     gw_resp,
            "error_message":        (kwargs.get("error_message") or "")[:500],
            "failure_reason":       (kwargs.get("failure_reason") or "")[:500],
            "remarks":              kwargs.get("remarks") or "",
        })
        row.insert(ignore_permissions=True)
    except Exception:
        frappe.log_error(frappe.get_traceback(), f"Payment log insert failed — {student_name}")


def _get_request_ip():
    """Return the client IP from the current Frappe request, or empty string."""
    try:
        if hasattr(frappe, "request") and frappe.request:
            return (
                frappe.request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
                or frappe.request.remote_addr
                or ""
            )
    except Exception:
        pass
    return ""


@frappe.whitelist()
def get_fee_demand_payment_logs(fee_demand_name):
    """Return payment log rows and analytics for a specific Fee Demand.

    Used by the Payment Details button on the Fee Demand form.
    Requires read permission on the linked Student Master.
    """
    student_name = frappe.db.get_value("Fee Demand", fee_demand_name, "student")
    if not student_name:
        frappe.throw(frappe._("Fee Demand not found."), frappe.DoesNotExistError)

    if not frappe.has_permission("Student Master", "read", doc=student_name):
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    rows = frappe.db.get_all(
        "Student Fee Payment Log",
        filters={
            "parent":     student_name,
            "parenttype": "Student Master",
            "fee_demand": fee_demand_name,
        },
        fields=[
            "name", "event_type", "timestamp", "amount", "currency",
            "invoice", "fee_demand", "payment_mode", "payment_method",
            "razorpay_payment_id", "razorpay_order_id", "transaction_id",
            "triggered_by", "paid_by_role", "paid_by_name",
            "attempt_type", "retry_count", "webhook_status",
            "from_status", "to_status", "ip_address",
            "gateway_response", "error_message", "failure_reason", "remarks",
        ],
        order_by="timestamp desc",
    )

    total      = len(rows)
    successful = sum(1 for r in rows if r.event_type in ("Captured", "Payment Recorded"))
    failed     = sum(1 for r in rows if r.event_type == "Payment Failed")
    cancelled  = sum(1 for r in rows if r.event_type == "Payment Cancelled")
    initiated  = sum(1 for r in rows if r.event_type == "Payment Initiated")
    refunded   = sum(1 for r in rows if r.event_type == "Refunded")
    last_attempt = str(rows[-1].timestamp) if rows else None

    analytics = {
        "total_attempts": total,
        "successful":     successful,
        "failed":         failed,
        "cancelled":      cancelled,
        "initiated":      initiated,
        "refunded":       refunded,
        "webhook_events": 0,
        "last_attempt":   last_attempt,
        "success_rate":   round((successful / initiated * 100) if initiated > 0 else 0, 1),
    }

    return {"logs": rows, "analytics": analytics}


@frappe.whitelist()
def get_payment_logs(student_name):
    """Return all payment log rows for a Student Master as a list of dicts.

    Used by the 'View Payment Logs' dialog in the admin UI.
    Requires read permission on Student Master.
    """
    if not frappe.has_permission("Student Master", "read", doc=student_name):
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    rows = frappe.db.get_all(
        "Student Fee Payment Log",
        filters={"parent": student_name, "parenttype": "Student Master"},
        fields=[
            "name", "event_type", "timestamp", "amount", "currency",
            "invoice", "fee_demand", "payment_mode", "payment_method",
            "razorpay_payment_id", "razorpay_order_id", "transaction_id",
            "triggered_by", "paid_by_role", "paid_by_name",
            "attempt_type", "retry_count", "webhook_status",
            "from_status", "to_status", "ip_address",
            "gateway_response", "error_message", "failure_reason", "remarks",
        ],
        order_by="timestamp desc",
    )
    return rows


# ---------------------------------------------------------------------------
# Student ID generation (bulk upload + auto generate)
# ---------------------------------------------------------------------------

@frappe.whitelist()
def get_batch_filter_options():
    """Return one row per Batch with a human-friendly Programme label, for the
    cascading Programme / Academic Year / Term filters in the Student ID
    generator dialogs."""
    batches = frappe.get_all(
        "Batch",
        fields=["name", "program", "academic_year", "academic_term"],
        order_by="academic_year desc, academic_term asc",
    )
    if not batches:
        return []

    programme_names = list({b["program"] for b in batches if b.get("program")})
    programmes = frappe.get_all(
        "Programme",
        filters={"name": ["in", programme_names]},
        fields=["name", "program_name"],
    ) if programme_names else []
    programme_map = {p["name"]: p for p in programmes}

    master_names = list({
        p["program_name"] for p in programmes
        if p.get("program_name") and frappe.db.exists("Programme Master", p["program_name"])
    })
    masters = frappe.get_all(
        "Programme Master",
        filters={"name": ["in", master_names]},
        fields=["name", "programme_name"],
    ) if master_names else []
    master_label = {m["name"]: m["programme_name"] for m in masters}

    options = []
    for b in batches:
        if not b.get("program"):
            continue
        prog = programme_map.get(b["program"]) or {}
        display_name = master_label.get(prog.get("program_name")) or prog.get("program_name") or b["program"]
        label = f"{display_name} ({b.get('academic_year') or ''})".strip()
        options.append({
            "batch":            b["name"],
            "programme":        b.get("program"),
            "programme_label":  label,
            "academic_year":    b.get("academic_year") or "",
            "term_name":        b.get("academic_term") or "",
        })
    return options


def _programme_code_for_batch(batch_name):
    """Resolve the Programme code (Programme Master's programme_code, e.g.
    "TLLM") for a Batch, via Batch.program -> Programme.program_code."""
    programme = frappe.db.get_value("Batch", batch_name, "program")
    if not programme:
        return ""
    return frappe.db.get_value("Programme", programme, "program_code") or ""


@frappe.whitelist()
def preview_student_ids(batches):
    """Given a list of Batch names, recompute Student IDs for EVERY student in
    those batches (not just ones missing an ID), sorted alphabetically by
    name, in the form <programme_code><academic_year><seq>, e.g.
    B20262027001. Re-running this after adding a new student re-sorts the
    whole group, so a student who now sorts earlier correctly takes over a
    lower number and everyone after them shifts up - existing IDs are not
    preserved across runs."""
    batches = frappe.parse_json(batches) if isinstance(batches, str) else batches
    if not batches:
        frappe.throw(_("Please select Programme, Academic Year and Term."))

    students = frappe.get_all(
        "Student Master",
        filters={"programme": ["in", batches]},
        fields=["name", "first_name", "last_name", "academic_year", "registration_id"],
    )
    if not students:
        return []

    students.sort(key=lambda s: f"{s.get('first_name') or ''} {s.get('last_name') or ''}".strip().lower())

    programme_code = _programme_code_for_batch(batches[0])
    academic_year = (students[0].get("academic_year") or "").replace("-", "").replace(" ", "")
    prefix = f"{programme_code}{academic_year}"

    return [
        {
            "name":         s["name"],
            "student_name": f"{s.get('first_name') or ''} {s.get('last_name') or ''}".strip(),
            "student_id":   f"{prefix}{idx + 1:03d}",
            "current_id":   s.get("registration_id") or "",
        }
        for idx, s in enumerate(students)
    ]


@frappe.whitelist()
def apply_student_ids(assignments):
    """Persist the Student IDs generated by preview_student_ids."""
    assignments = frappe.parse_json(assignments) if isinstance(assignments, str) else assignments
    updated = 0
    for row in (assignments or []):
        if not row.get("name") or not row.get("student_id"):
            continue
        frappe.db.set_value("Student Master", row["name"], "registration_id", row["student_id"])
        updated += 1
    frappe.db.commit()
    return {"updated": updated}


@frappe.whitelist()
def download_student_id_bulk_template(batches=None):
    """Build an xlsx template pre-filled with Student Name / Programme /
    Academic Year / Term / Batch for the given Batches (or all students if
    none given), leaving Student ID blank (or showing the existing one) for
    the admin to fill in and upload back."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        frappe.throw(_("openpyxl is not installed. Run: bench pip install openpyxl"))

    batches = frappe.parse_json(batches) if isinstance(batches, str) else batches

    filters = {}
    if batches:
        filters["programme"] = ["in", batches]

    students = frappe.get_all(
        "Student Master",
        filters=filters,
        fields=[
            "name", "first_name", "last_name",
            "academic_year", "academic_term", "programme", "registration_id",
        ],
        order_by="first_name asc, last_name asc",
    )
    if not students:
        frappe.throw(_("No students found for the selected filters."))

    batch_names = {s["programme"] for s in students if s.get("programme")}
    batch_rows = frappe.get_all(
        "Batch",
        filters={"name": ["in", list(batch_names)]},
        fields=["name", "batch_name", "program"],
    ) if batch_names else []
    batch_info = {b["name"]: b for b in batch_rows}

    programme_names = {b["program"] for b in batch_rows if b.get("program")}
    programme_code = {
        p["name"]: p["program_code"] for p in frappe.get_all(
            "Programme",
            filters={"name": ["in", list(programme_names)]},
            fields=["name", "program_code"],
        )
    } if programme_names else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student IDs"

    headers = ["Row Key", "Student Name", "Programme", "Academic Year", "Term", "Batch", "Student ID"]
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3949AB")

    for s in students:
        student_name = f"{s.get('first_name') or ''} {s.get('last_name') or ''}".strip()
        batch = batch_info.get(s.get("programme")) or {}
        ws.append([
            s["name"],
            student_name,
            programme_code.get(batch.get("program")) or "",
            s.get("academic_year") or "",
            s.get("academic_term") or "",
            batch.get("batch_name") or s.get("programme") or "",
            s.get("registration_id") or "",
        ])

    ws.column_dimensions["A"].hidden = True
    for col, width in zip("BCDEFG", [26, 14, 14, 16, 20, 16]):
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return {
        "filename": "student_id_bulk_template.xlsx",
        "content":  base64.b64encode(output.read()).decode("utf-8"),
        "mime":     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


@frappe.whitelist()
def upload_student_ids_bulk(file_url):
    """Read the filled-in Student ID template and update registration_id for
    each matched Student Master row. Matches on the hidden Row Key column
    (column A, the Student Master document name) so edits to the other
    display columns don't break matching."""
    try:
        import openpyxl
    except ImportError:
        frappe.throw(_("openpyxl is not installed. Run: bench pip install openpyxl"))

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
    ws = wb.active

    updated = 0
    skipped = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        student = str(row[0]).strip()
        new_id = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        if not new_id:
            continue
        if not frappe.db.exists("Student Master", student):
            skipped.append(student)
            continue
        frappe.db.set_value("Student Master", student, "registration_id", new_id)
        updated += 1

    frappe.db.commit()
    return {"updated": updated, "skipped": skipped}


@frappe.whitelist()
def get_payment_analytics(student_name):
    """Return summary analytics for the payment log of a student.

    Returns counts by event type plus last-attempt metadata.
    """
    if not frappe.has_permission("Student Master", "read", doc=student_name):
        frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

    rows = frappe.db.get_all(
        "Student Fee Payment Log",
        filters={"parent": student_name, "parenttype": "Student Master"},
        fields=["event_type", "timestamp", "amount", "razorpay_payment_id"],
        order_by="timestamp asc",
    )

    total          = len(rows)
    successful     = sum(1 for r in rows if r.event_type in ("Captured", "Payment Recorded"))
    failed         = sum(1 for r in rows if r.event_type == "Payment Failed")
    cancelled      = sum(1 for r in rows if r.event_type == "Payment Cancelled")
    initiated      = sum(1 for r in rows if r.event_type == "Payment Initiated")
    refunded       = sum(1 for r in rows if r.event_type == "Refunded")
    webhook_events = sum(1 for r in rows if r.event_type in ("Webhook Received", "Pending Verification"))

    last_attempt   = None
    if rows:
        last_attempt = str(rows[-1].timestamp)

    success_rate = round((successful / initiated * 100) if initiated > 0 else 0, 1)

    return {
        "total_attempts":    total,
        "successful":        successful,
        "failed":            failed,
        "cancelled":         cancelled,
        "initiated":         initiated,
        "refunded":          refunded,
        "webhook_events":    webhook_events,
        "last_attempt":      last_attempt,
        "success_rate":      success_rate,
    }
