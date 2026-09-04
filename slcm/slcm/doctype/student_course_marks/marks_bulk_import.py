import frappe
from frappe.utils import now
import json
import io
import openpyxl
import traceback

@frappe.whitelist()
def parse_import_file(file_url, exam_plan, evaluation_schema, exam_component=None,
                      assessment_type=None, re_exam_component=None, re_exam_assessment_type=None):
    """
    Create a Marks Import Log and queue a background job to parse the uploaded file.
    Returns immediately with {import_log, status:"parsing"} so large files never timeout.
    Emits marks_parsing_progress and marks_parsing_complete realtime events.
    """
    try:
        if not exam_plan or not evaluation_schema:
            frappe.throw("Exam Plan and Evaluation Schema are mandatory.")

        # Verify the file record exists before queuing (fast check)
        if not frappe.db.exists("File", {"file_url": file_url}):
            frappe.throw(f"Uploaded file record not found. Please re-upload the file.")

        # Create the import log immediately so we can hand its name back to the browser
        log = frappe.get_doc({
            "doctype": "Marks Import Log",
            "import_file": file_url,
            "exam_plan": exam_plan,
            "evaluation_schema": evaluation_schema,
            "exam_component": exam_component,
            "assessment_type": assessment_type,
            "re_exam_component": re_exam_component,
            "re_exam_assessment_type": re_exam_assessment_type,
            "status": "Staging",
            "total_rows": 0,
            "success_count": 0,
            "failed_count": 0,
            "skipped_count": 0,
            "missing_offerings_count": 0
        })
        log.insert(ignore_permissions=True)

        # Queue the actual heavy parsing in the background
        frappe.enqueue(
            method="slcm.slcm.doctype.student_course_marks.marks_bulk_import._run_parse_job",
            queue="long",
            timeout=1800,
            import_log_name=log.name,
            file_url=file_url,
            triggering_user=frappe.session.user
        )
        return {"import_log": log.name, "status": "parsing"}

    except frappe.exceptions.ValidationError:
        raise
    except frappe.exceptions.DoesNotExistError:
        frappe.throw("File record not found. Please re-upload the file.")
    except Exception as e:
        frappe.log_error(message=frappe.get_traceback(), title=f"parse_import_file failed: {file_url}")
        frappe.throw(f"Failed to start file parsing: {e}")


def _run_parse_job(import_log_name, file_url, triggering_user=None):
    """Background worker: wraps _run_parse_job_impl with error handling."""
    try:
        _run_parse_job_impl(import_log_name, file_url, triggering_user)
    except Exception as e:
        frappe.db.set_value("Marks Import Log", import_log_name, "status", "Failed")
        frappe.db.commit()
        frappe.log_error(message=frappe.get_traceback(), title=f"Parse Job Failed: {import_log_name}")
        if triggering_user:
            frappe.publish_realtime("marks_parsing_complete", {
                "import_log": import_log_name,
                "error": str(e),
                "error_title": "File Parsing Failed"
            }, user=triggering_user)


def _run_parse_job_impl(import_log_name, file_url, triggering_user=None):
    """Actual parsing: reads Excel rows, creates Marks Import Log Detail records."""
    triggering_user = triggering_user or frappe.session.user
    log = frappe.get_doc("Marks Import Log", import_log_name)
    frappe.db.set_value("Marks Import Log", import_log_name, "status", "In Progress")
    frappe.db.commit()

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = file_doc.get_content()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        frappe.throw("Could not open the uploaded file. Please ensure it is a valid .xlsx Excel file.")

    sheet = wb.active
    headers = []
    total_sheet_rows = max(sheet.max_row - 1, 1)
    row_count = 0

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else "" for c in row]
            required = ["Registration ID", "Course Code", "Term Name"]
            missing_hdrs = [h for h in required if h not in headers]
            if missing_hdrs:
                frappe.throw(
                    f"File is missing required columns: {', '.join(missing_hdrs)}. "
                    f"Please download the sample template and use the correct column headers."
                )
            continue

        row_dict = {}
        empty = True
        for j, cell in enumerate(row):
            if j < len(headers):
                val = cell if cell is not None else ""
                row_dict[headers[j]] = val
                if str(val).strip():
                    empty = False

        if not empty:
            reg_id = row_dict.get("Registration ID")
            course_code = row_dict.get("Course Code")
            term_name = row_dict.get("Term Name")

            detail = frappe.new_doc("Marks Import Log Detail")
            detail.import_log = log.name
            detail.row_number = i + 1
            detail.registration_id = str(reg_id)[:255] if reg_id else ""
            detail.course_code = str(course_code)[:255] if course_code else ""
            detail.term_name = str(term_name)[:255] if term_name else ""
            detail.status = "Pending"
            detail.raw_row_json = json.dumps(row_dict)
            detail.flags.ignore_permissions = True
            detail.insert()
            row_count += 1

        data_row = i  # 0-indexed after header skip
        if data_row % 50 == 0:
            frappe.db.set_value("Marks Import Log", import_log_name, "processed_count", row_count, update_modified=False)
            frappe.db.commit()
            pct = min(int((data_row / total_sheet_rows) * 100), 99)
            frappe.publish_realtime("marks_parsing_progress", {
                "progress": row_count,
                "total": total_sheet_rows,
                "percent": pct,
                "import_log": import_log_name
            }, user=triggering_user)

    # Finalise the log
    log.db_set({"total_rows": row_count, "status": "Staging"})

    frappe.publish_realtime("marks_parsing_progress", {
        "progress": row_count,
        "total": row_count,
        "percent": 100,
        "import_log": import_log_name
    }, user=triggering_user)

    frappe.publish_realtime("marks_parsing_complete", {
        "import_log": import_log_name,
        "total_rows": row_count,
        "status": "Staging"
    }, user=triggering_user)

@frappe.whitelist()
def delete_import_log(import_log):
    if not import_log: return
    frappe.db.delete("Marks Import Log Detail", {"import_log": import_log})
    frappe.delete_doc("Marks Import Log", import_log, ignore_permissions=True)

@frappe.whitelist()
def discard_import_draft(import_log):
    if not import_log:
        return
    log = frappe.db.get_value("Marks Import Log", import_log,
        ["status", "import_file"], as_dict=True)
    if not log:
        return

    active_job_statuses = ["Queued", "In Progress"]
    if log.status in active_job_statuses:
        frappe.throw(
            "Cannot discard while a background job is still running "
            "for this import. Please wait for it to finish, then try "
            "again."
        )
    if log.status not in ["Staging", "Validated", "Failed", "Completed with Errors"]:
        frappe.throw("Cannot discard an import that is actively running or already completed successfully.")
    
    delete_import_log(import_log)

@frappe.whitelist()
def get_preview_page(import_log, page=1, page_size=50):
    page = int(page)
    page_size = int(page_size)
    details = frappe.get_all("Marks Import Log Detail", 
        filters={"import_log": import_log},
        fields=["name", "row_number", "status", "registration_id", "course_code", "term_name", "raw_row_json"],
        order_by="row_number asc",
        limit_start=(page - 1) * page_size,
        limit_page_length=page_size
    )
    
    for d in details:
        if d.raw_row_json:
            d.raw_data = json.loads(d.raw_row_json)
        else:
            d.raw_data = {}
            
    total = frappe.db.count("Marks Import Log Detail", {"import_log": import_log})
    return {
        "rows": details,
        "total": total,
        "page": page,
        "page_size": page_size
    }

@frappe.whitelist()
def get_errors_page(import_log, page=1, page_size=20, status_filter=None):
    page = int(page)
    page_size = int(page_size)
    
    filters = {"import_log": import_log}
    if status_filter:
        filters["status"] = status_filter
    else:
        filters["status"] = ["!=", "Valid"]
        
    details = frappe.get_all("Marks Import Log Detail", 
        filters=filters,
        fields=["name", "row_number", "status", "error_reason"],
        order_by="row_number asc",
        limit_start=(page - 1) * page_size,
        limit_page_length=page_size
    )
    total = frappe.db.count("Marks Import Log Detail", filters=filters)
    return {
        "rows": details,
        "total": total,
        "page": page,
        "page_size": page_size
    }

@frappe.whitelist()
def start_validation(import_log, retry_mode=None, triggering_user=None):
    # Normalize retry_mode — JS boolean `false` arrives as the string "false" (truthy!)
    # Accepted values: None / "" / "false" / False  → first validate (Pending rows only)
    #                  "revalidate"                  → re-validate all rows
    #                  any other truthy string        → retry failed rows
    if retry_mode in (None, False, "false", ""):
        retry_mode = None  # first-time validate

    if retry_mode == "revalidate":
        log_status = frappe.db.get_value("Marks Import Log", import_log, "status")
        if log_status not in ("Staging", "Validated"):
            frappe.throw("Cannot re-validate an import that has already started.")

    triggering_user = triggering_user or frappe.session.user
    # D1: Mark as Queued before queuing so UI state is consistent
    frappe.db.set_value("Marks Import Log", import_log, "status", "Queued")
    frappe.db.commit()
    frappe.enqueue(
        method="slcm.slcm.doctype.student_course_marks.marks_bulk_import.run_validation_job",
        queue="long", timeout=3600,
        import_log=import_log, retry_mode=retry_mode, triggering_user=triggering_user
    )
    return {"status": "queued"}

def run_validation_job(import_log, retry_mode=None, triggering_user=None):
    try:
        _run_validation_job_impl(import_log, retry_mode, triggering_user)
    except Exception as e:
        # D1b: Reset to Staging so the user can try again (not permanently Failed)
        frappe.db.set_value("Marks Import Log", import_log, "status", "Staging")
        frappe.db.commit()
        frappe.log_error(message=frappe.get_traceback(), title=f"Validation Job Failed for {import_log}")
        if triggering_user:
            frappe.publish_realtime("marks_validation_complete", {
                "import_log": import_log,
                "error": str(e),
                "error_title": "Validation Failed"
            }, user=triggering_user)

def _run_validation_job_impl(import_log, retry_mode=None, triggering_user=None):
    """
    P2: Validate rows using bulk pre-loaded lookup maps instead of per-row DB queries.
    Reduces DB round-trips from ~4×N to ~5 bulk queries regardless of row count.
    """
    triggering_user = triggering_user or frappe.session.user
    # Normalize retry_mode (may come as string "false" from enqueue serialization)
    if retry_mode in (None, False, "false", ""):
        retry_mode = None

    filters = {"import_log": import_log}
    if retry_mode == "revalidate":
        pass  # No status filter — re-check ALL rows
    elif retry_mode:
        filters["status"] = ["in", ["Failed", "Missing Student", "Missing Course",
                                    "Missing Course Offering", "Duplicate (Skip)"]]
    else:
        filters["status"] = "Pending"

    frappe.db.set_value("Marks Import Log", import_log, "status", "In Progress")
    frappe.db.commit()

    details = frappe.get_all("Marks Import Log Detail", filters=filters,
        fields=["name", "registration_id", "course_code", "term_name", "raw_row_json"])
    total = len(details)

    if total == 0:
        # Nothing to validate — emit complete with current counts
        _publish_validation_complete(import_log, triggering_user)
        return

    # ── P2: Parse all rows first to extract lookup keys ──────────────────────
    all_items = []
    for d in details:
        row = json.loads(d.raw_row_json) if d.raw_row_json else {}
        all_items.append({
            "detail": d,
            "row": row,
            "reg_id": str(row.get("Registration ID", "") or ""),
            "course_code": str(row.get("Course Code", "") or ""),
            "term_name": str(row.get("Term Name", "") or ""),
            "batch": str(row.get("Batch Year", "") or ""),
        })

    # ── P2: Bulk DB query 1 — Students ───────────────────────────────────────
    all_reg_ids = list({it["reg_id"] for it in all_items if it["reg_id"]})
    students_map = {}  # registration_id → student.name
    if all_reg_ids:
        for sm in frappe.get_all("Student Master",
                filters={"registration_id": ["in", all_reg_ids]},
                fields=["name", "registration_id"]):
            students_map[sm.registration_id] = sm.name

    # ── P2: Bulk DB query 2 — Courses (by course_code + by name fallback) ────
    all_course_codes = list({it["course_code"] for it in all_items if it["course_code"]})
    courses_map = {}  # code_or_name → course.name
    if all_course_codes:
        for c in frappe.get_all("Course",
                filters={"course_code": ["in", all_course_codes]},
                fields=["name", "course_code"]):
            courses_map[c.course_code] = c.name
        # Fallback: some rows use the course name directly as "Course Code"
        missing_codes = [cc for cc in all_course_codes if cc not in courses_map]
        if missing_codes:
            for c in frappe.get_all("Course",
                    filters={"name": ["in", missing_codes]},
                    fields=["name"]):
                courses_map[c.name] = c.name

    # ── P2: Bulk DB query 3 — All Course Offerings ───────────────────────────
    # Load ALL offerings once; build a (course_title, cohort, term_name) → name map
    offerings_map = {}  # (course_title, cohort, term_name) → offering.name
    for off in frappe.get_all("Course Offering",
            fields=["name", "course_title", "cohort", "term_name"]):
        key = (off.course_title, off.cohort, off.term_name)
        offerings_map[key] = off.name

    # ── P2: Bulk DB query 4 — Existing marks (duplicate check) ───────────────
    all_student_names = list(set(students_map.values()))
    existing_marks = set()  # set of (student.name, course_offering.name)
    if all_student_names:
        for mk in frappe.get_all("Student Course Marks",
                filters={"student": ["in", all_student_names]},
                fields=["student", "course_offering"]):
            existing_marks.add((mk.student, mk.course_offering))

    # ── Validation loop — now O(1) dict lookups per row ──────────────────────
    missing_offering_groups = {}
    updates = {}

    for i, item in enumerate(all_items):
        d         = item["detail"]
        reg_id    = item["reg_id"]
        course_code = item["course_code"]
        term_name = item["term_name"]
        batch     = item["batch"]

        status = "Valid"
        error_reason = ""

        # 1. Student check
        student = students_map.get(reg_id)
        if not student:
            status = "Missing Student"
            error_reason = f"Student not found for Registration ID: {reg_id}"

        # 2. Course check
        if not error_reason:
            course_name = courses_map.get(course_code)
            if not course_name:
                status = "Missing Course"
                error_reason = f"Course Code '{course_code}' does not exist in Course master"

        # 3. Course Offering check
        if not error_reason:
            offering_key = (course_name, batch, term_name)
            course_offering = offerings_map.get(offering_key)
            if not course_offering:
                group_key = (course_code, batch, term_name)
                missing_offering_groups[group_key] = missing_offering_groups.get(group_key, 0) + 1
                status = "Missing Course Offering"
                error_reason = (f"No Course Offering for Course '{course_code}', "
                                f"Batch '{batch}', Term '{term_name}'")

        # 4. Duplicate check
        if not error_reason and student and course_offering:
            if (student, course_offering) in existing_marks:
                status = "Duplicate (Skip)"
                error_reason = "Student Course Marks already exists for this combination"

        updates[d.name] = {"status": status, "error_reason": error_reason}

        if len(updates) >= 200 or i == total - 1:
            frappe.db.bulk_update("Marks Import Log Detail", updates)
            updates = {}
            frappe.db.commit()

        if i % 50 == 0 or i == total - 1:
            frappe.db.set_value("Marks Import Log", import_log, "processed_count", i + 1, update_modified=False)
            frappe.db.commit()
            pct = round(((i + 1) / total) * 100, 1)
            frappe.publish_realtime("marks_validation_progress", {
                "import_log": import_log,
                "progress": i + 1,
                "total": total,
                "percent": pct
            }, user=triggering_user)

    # Build missing offering groups list from accumulated dict
    missing_groups_list = [
        {"course_code": k[0], "batch": k[1], "term_name": k[2],
         "affected_row_count": v}
        for k, v in missing_offering_groups.items()
    ]
    # Merge with any pre-existing groups from DB (if this is a partial revalidate)
    _publish_validation_complete(import_log, triggering_user,
                                 extra_missing_groups=missing_groups_list)


def _publish_validation_complete(import_log, triggering_user, extra_missing_groups=None):
    """Aggregate final counts, update the log, and emit marks_validation_complete."""
    valid_count          = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Valid"})
    error_count          = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Failed"})
    skip_count           = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Duplicate (Skip)"})
    missing_student_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Missing Student"})
    missing_course_count  = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Missing Course"})
    missing_off_count     = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Missing Course Offering"})

    missing_groups_list = extra_missing_groups or []
    if not missing_groups_list and missing_off_count > 0:
        # Rebuild from DB for display
        missing_offs = frappe.get_all("Marks Import Log Detail",
            {"import_log": import_log, "status": "Missing Course Offering"},
            ["course_code", "term_name", "raw_row_json"])
        groups = {}
        for m in missing_offs:
            r = json.loads(m.raw_row_json) if m.raw_row_json else {}
            b = r.get("Batch Year", "")
            k = (m.course_code, b, m.term_name)
            groups[k] = groups.get(k, 0) + 1
        missing_groups_list = [
            {"course_code": k[0], "batch": k[1], "term_name": k[2], "affected_row_count": v}
            for k, v in groups.items()
        ]

    frappe.db.set_value("Marks Import Log", import_log, {
        "success_count": valid_count,
        "failed_count": error_count + missing_student_count + missing_course_count,
        "skipped_count": skip_count,
        "missing_offerings_count": missing_off_count,
        "status": "Validated"
    })
    frappe.db.commit()

    summary = {
        "valid_count": valid_count,
        "error_count": error_count,
        "skip_count": skip_count,
        "missing_student_count": missing_student_count,
        "missing_course_count": missing_course_count,
        "missing_offering_count": missing_off_count,
        "missing_offering_groups": missing_groups_list
    }
    frappe.publish_realtime("marks_validation_complete",
        {"import_log": import_log, **summary}, user=triggering_user)
        



@frappe.whitelist()
def get_validation_summary(import_log):
    valid_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Valid"})
    error_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Failed"})
    skip_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Duplicate (Skip)"})
    missing_student_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Missing Student"})
    missing_course_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Missing Course"})
    missing_off_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log, "status": "Missing Course Offering"})
    
    missing_offering_groups = {}
    missing_offs = frappe.get_all("Marks Import Log Detail", {"import_log": import_log, "status": "Missing Course Offering"}, ["course_code", "term_name", "raw_row_json"])
    for m in missing_offs:
        r = json.loads(m.raw_row_json) if m.raw_row_json else {}
        b = r.get("Batch Year", "")
        k = (m.course_code, b, m.term_name)
        if k not in missing_offering_groups:
            missing_offering_groups[k] = 0
        missing_offering_groups[k] += 1

    missing_groups_list = [{"course_code": k[0], "batch": k[1], "term_name": k[2], "affected_row_count": v} for k, v in missing_offering_groups.items()]

    return {
        "import_log": import_log,
        "valid_count": valid_count,
        "error_count": error_count,
        "skip_count": skip_count,
        "missing_student_count": missing_student_count,
        "missing_course_count": missing_course_count,
        "missing_offering_count": missing_off_count,
        "missing_offering_groups": missing_groups_list
    }

@frappe.whitelist()
def start_bulk_import(import_log, skip_blocked=True):
    log = frappe.get_doc("Marks Import Log", import_log)
    log.db_set("status", "Queued")
    
    frappe.enqueue(
        method="slcm.slcm.doctype.student_course_marks.marks_bulk_import.process_import",
        queue="long",
        timeout=3600,
        import_log_name=import_log,
        triggering_user=frappe.session.user
    )
    
    return {"import_log": log.name}

@frappe.whitelist()
def retry_failed_rows(import_log):
    """Re-validate then re-import only failed/missing rows."""
    try:
        retry_statuses = ["Failed", "Missing Student", "Missing Course", "Missing Course Offering"]
        detail_names = frappe.get_all("Marks Import Log Detail",
            filters={"import_log": import_log, "status": ["in", retry_statuses]},
            pluck="name")
        if not detail_names:
            frappe.throw("No failed rows found to retry. Please re-validate the import first.")
        # Queue re-validation for the failed rows
        frappe.enqueue(
            method="slcm.slcm.doctype.student_course_marks.marks_bulk_import.run_validation_job",
            queue="long",
            timeout=3600,
            import_log=import_log,
            retry_mode="retry",
            triggering_user=frappe.session.user
        )
        frappe.db.set_value("Marks Import Log", import_log, "status", "Queued")
        return {"status": "Queued", "row_count": len(detail_names)}
    except frappe.exceptions.ValidationError:
        raise
    except Exception as e:
        frappe.log_error(message=frappe.get_traceback(), title=f"retry_failed_rows failed: {import_log}")
        frappe.throw(f"Failed to queue retry: {e}")

def process_import(import_log_name, retry_scope=None, triggering_user=None):
    try:
        _process_import_impl(import_log_name, retry_scope, triggering_user)
    except Exception as e:
        frappe.db.set_value("Marks Import Log", import_log_name, "status", "Failed")
        frappe.db.commit()
        frappe.log_error(message=frappe.get_traceback(), title=f"Import Job Failed for {import_log_name}")
        if triggering_user:
            frappe.publish_realtime("marks_import_progress", {"import_log": import_log_name, "progress": 100, "total": 100}, user=triggering_user)

def _process_import_impl(import_log_name, retry_scope=None, triggering_user=None):
    log = frappe.get_doc("Marks Import Log", import_log_name)
    log.db_set("status", "In Progress")
    if not triggering_user:
        triggering_user = log.imported_by
        
    filters = {"import_log": import_log_name, "status": "Valid"}
    if retry_scope:
        filters["name"] = ["in", retry_scope]
        
    valid_details = frappe.get_all("Marks Import Log Detail", 
        filters=filters,
        pluck="name"
    )
    
    total = len(valid_details)
    success_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log_name, "status": "Success"})
    
    if total == 0:
        frappe.publish_realtime("marks_import_progress", {
            "import_log": import_log_name,
            "progress": 0,
            "total": 0
        }, user=triggering_user)
        
    for i, docname in enumerate(valid_details):
        if i % 25 == 0:
            pct = round((i / total) * 100, 1) if total > 0 else 0
            frappe.publish_realtime("marks_import_progress", {
                "import_log": import_log_name,
                "progress": i,
                "total": total,
                "percent": pct
            }, user=triggering_user)
            
        detail = frappe.get_doc("Marks Import Log Detail", docname)
        row = json.loads(detail.raw_row_json)
        
        reg_id = row.get("Registration ID")
        course_code = row.get("Course Code")
        term_name = row.get("Term Name")
        batch = row.get("Batch Year")
        
        try:
            student = resolve_student(reg_id)
            course_offering = resolve_course_offering(course_code, batch, term_name)
            
            course_name = frappe.db.get_value("Course", {"course_code": course_code}, "name")
            if not course_name:
                course_name = frappe.db.get_value("Course", {"name": course_code}, "name")
                
            doc = frappe.new_doc("Student Course Marks")
            doc.student = student
            doc.student_id = reg_id
            doc.course_offering = course_offering
            doc.course = course_name or course_code
            doc.exam_plan = log.exam_plan
            doc.evaluation_schema = log.evaluation_schema
            doc.status = "Submitted"
            
            grade_point = row.get("Grade Point")
            try:
                gp_val = float(grade_point)
            except:
                gp_val = 0.0
            doc.consider_for_sgpa = 0 if gp_val == 0.0 else 1
            
            doc.total_marks = row.get("Total Marks")
            doc.grade = row.get("Grade")
            doc.re_exam_grade = row.get("Re-Exam Grade")
            
            computed = compute_updated_final(row)
            doc.updated_final_marks = computed["updated_final_marks"]
            doc.updated_grade = computed["updated_grade"]
            doc.improvement_marks = computed["improvement_marks"]
            doc.improvement_grade = computed["improvement_grade"]
            doc.improvement_applied = computed["improvement_applied"]
            
            if log.exam_component and log.assessment_type:
                doc.append("marks_entries", {
                    "component": log.exam_component,
                    "assessment_type": log.assessment_type,
                    "marks": doc.total_marks
                })
            
            if row.get("Re-Exam Total Marks") and log.re_exam_component and log.re_exam_assessment_type:
                doc.append("marks_entries", {
                    "component": log.re_exam_component,
                    "assessment_type": log.re_exam_assessment_type,
                    "marks": row.get("Re-Exam Total Marks"),
                    "is_reexam": 1
                })
            
            doc.flags.ignore_permissions = True
            doc.insert()
            
            detail.db_set("status", "Success")
            detail.db_set("student_course_marks", doc.name)
            detail.db_set("error_reason", "")
            success_count += 1
            
        except Exception as e:
            err_msg = traceback.format_exc()
            detail.db_set("status", "Failed")
            detail.db_set("error_reason", str(e) + "\n" + err_msg)
            
    if total > 0:
        frappe.publish_realtime("marks_import_progress", {
            "import_log": import_log_name,
            "progress": total,
            "total": total,
            "percent": 100
        }, user=triggering_user)
    
    # Final counts for Log
    failed_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log_name, "status": "Failed"})
    missing_student = frappe.db.count("Marks Import Log Detail", {"import_log": import_log_name, "status": "Missing Student"})
    missing_course = frappe.db.count("Marks Import Log Detail", {"import_log": import_log_name, "status": "Missing Course"})
    missing_off_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log_name, "status": "Missing Course Offering"})
    skipped_count = frappe.db.count("Marks Import Log Detail", {"import_log": import_log_name, "status": "Duplicate (Skip)"})
    
    total_failed = failed_count + missing_student + missing_course + missing_off_count
    
    final_status = "Completed with Errors" if total_failed > 0 else "Completed"
    log.db_set({
        "success_count": success_count,
        "failed_count": total_failed,
        "skipped_count": skipped_count,
        "missing_offerings_count": missing_off_count,
        "status": final_status
    })

    # E5: Log a summary to Frappe Error Log when there are failures, for admin visibility
    if total_failed > 0:
        err_lines = [
            f"Import {import_log_name} completed with errors.",
            f"  Successful         : {success_count}",
            f"  Total Failed       : {total_failed}",
            f"    Row errors       : {failed_count}",
            f"    Missing students : {missing_student}",
            f"    Missing courses  : {missing_course}",
            f"    Missing offerings: {missing_off_count}",
            f"  Skipped (dupes)    : {skipped_count}",
            "Open the Marks Import Log to view per-row error details.",
        ]
        frappe.log_error(
            message="\n".join(err_lines),
            title=f"Import Errors — {import_log_name}"
        )

    # Emit final import complete event
    if triggering_user:
        frappe.publish_realtime("marks_import_complete", {
            "import_log": import_log_name,
            "success_count": success_count,
            "failed_count": total_failed,
            "status": final_status
        }, user=triggering_user)
    
def compute_updated_final(row):
    original_grade = row.get("Grade") or ""
    try:
        total_marks = float(row.get("Total Marks") or 0)
    except:
        total_marks = 0.0
        
    reexam_grade = row.get("Re-Exam Grade")
    try:
        reexam_total_marks = float(row.get("Re-Exam Total Marks") or 0) if row.get("Re-Exam Total Marks") else None
    except:
        reexam_total_marks = None
        
    if reexam_total_marks is None:
        return {
            "updated_final_marks": total_marks,
            "updated_grade": original_grade,
            "improvement_marks": 0,
            "improvement_grade": "",
            "improvement_applied": 0
        }
        
    if original_grade == 'F' and reexam_grade != 'F':
        return {
            "updated_final_marks": reexam_total_marks,
            "updated_grade": reexam_grade,
            "improvement_marks": reexam_total_marks,
            "improvement_grade": reexam_grade,
            "improvement_applied": 0
        }
        
    if original_grade != 'F' and reexam_total_marks > total_marks:
        return {
            "updated_final_marks": total_marks,
            "updated_grade": original_grade,
            "improvement_marks": reexam_total_marks,
            "improvement_grade": reexam_grade,
            "improvement_applied": 1
        }
        
    return {
        "updated_final_marks": total_marks,
        "updated_grade": original_grade,
        "improvement_marks": reexam_total_marks,
        "improvement_grade": reexam_grade,
        "improvement_applied": 0
    }

def resolve_student(registration_id):
    if not registration_id: return None
    return frappe.db.get_value("Student Master", {"registration_id": registration_id}, "name")

def resolve_course_offering(course_code, batch, term_name):
    if not (course_code and batch and term_name): return None
    course_name = frappe.db.get_value("Course", {"course_code": course_code}, "name")
    if not course_name:
        course_name = frappe.db.get_value("Course", {"name": course_code}, "name")
    
    if not course_name: return None
    
    return frappe.db.get_value("Course Offering", {
        "course_title": course_name,
        "cohort": batch,
        "term_name": term_name
    }, "name")

@frappe.whitelist()
def download_sample_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Template"
    
    headers = [
        "Term Name", "Registration ID", "Student Name", "Programme Name", 
        "Batch Year", "Department Name", "Course Code", "Course Name", 
        "Total Marks", "Grade", "Grade Point", "Re-Exam Total Marks", 
        "Re-Exam Grade", "Re-Exam Grade Point", "SGPA", "CGPA"
    ]
    ws.append(headers)
    
    ws.append([
        "T1-2025-2026", "REG001", "John Doe", "B.A., LL.B.", 
        "B-AY 2025-2026", "Law", "LAW101", "Legal Methods", 
        85, "A", 4.0, "", "", "", 4.0, 4.0
    ])
    
    ws.append([
        "T1-2025-2026", "REG002", "Jane Smith", "B.A., LL.B.", 
        "B-AY 2025-2026", "Law", "LAW102", "Contracts", 
        45, "F", 0.0, 75, "B", 3.0, 2.5, 3.0
    ])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    frappe.response['filename'] = "Student_Marks_Import_Template.xlsx"
    frappe.response['filecontent'] = stream.read()
    frappe.response['type'] = 'download'

def _download_csv(headers, data, filename):
    import csv
    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(headers)
    writer.writerows(data)
    frappe.response['result'] = stream.getvalue()
    frappe.response['type'] = 'csv'
    frappe.response['doctype'] = filename

@frappe.whitelist()
def download_missing_course_offerings(import_log):
    rows = frappe.db.sql("""
        SELECT course_code, term_name, count(*) as affected_row_count
        FROM `tabMarks Import Log Detail`
        WHERE import_log = %s AND status = 'Missing Course Offering'
        GROUP BY course_code, term_name
    """, (import_log,))
    _download_csv(["Course Code", "Term Name", "Affected Row Count"], rows, "Missing_Course_Offerings.csv")

@frappe.whitelist()
def download_missing_courses(import_log):
    rows = frappe.db.sql("""
        SELECT course_code, count(*) as affected_row_count
        FROM `tabMarks Import Log Detail`
        WHERE import_log = %s AND status = 'Missing Course'
        GROUP BY course_code
    """, (import_log,))
    _download_csv(["Course Code", "Affected Row Count"], rows, "Missing_Courses.csv")

@frappe.whitelist()
def download_missing_students(import_log):
    rows = frappe.db.sql("""
        SELECT registration_id, count(*) as affected_row_count
        FROM `tabMarks Import Log Detail`
        WHERE import_log = %s AND status = 'Missing Student'
        GROUP BY registration_id
    """, (import_log,))
    _download_csv(["Registration ID", "Affected Row Count"], rows, "Missing_Students.csv")
@frappe.whitelist()
def download_missing_courses_template(import_log):
    rows = frappe.db.sql("""
        SELECT DISTINCT course_code, raw_row_json
        FROM `tabMarks Import Log Detail`
        WHERE import_log = %s AND status = 'Missing Course'
    """, import_log, as_dict=True)
    
    unique_courses = {}
    for r in rows:
        code = r.course_code
        if code not in unique_courses:
            row_data = json.loads(r.raw_row_json) if r.raw_row_json else {}
            unique_courses[code] = row_data.get("Course Name", "")
            
    out = [["ID", "Course Code", "Course Name", "Status"]]
    for code, name in unique_courses.items():
        out.append(["", code, name, "Active"])
        
    _download_csv(out[0], out[1:], f"Missing_Courses_Template_{import_log}.csv")

@frappe.whitelist()
def download_missing_course_offerings_template(import_log):
    rows = frappe.db.sql("""
        SELECT DISTINCT course_code, term_name, raw_row_json
        FROM `tabMarks Import Log Detail`
        WHERE import_log = %s AND status = 'Missing Course Offering'
    """, import_log, as_dict=True)
    
    unique_offerings = {}
    for r in rows:
        row_data = json.loads(r.raw_row_json) if r.raw_row_json else {}
        code = r.course_code
        name = row_data.get("Course Name", "")
        batch = row_data.get("Batch Year", "")
        
        key = (code, batch, r.term_name)
        if key not in unique_offerings:
            course = f"{code}-{name}" if name else code
            
            # Batch has program and academic year linked, but user said:
            # "Programme = Student Master.master_programme (plain value, NO compound string construction)"
            # "Academic Year = Student Master.academic_year"
            # Wait, the validation logic already fetched `resolve_student(reg_id)`. We don't have the student details in raw_row_json except Reg ID.
            # So let's fetch one missing row for each key to resolve the student details.
            student_id = row_data.get("Registration ID", "")
            # FIX: look up by registration_id field, not by docname
            sm = None
            if student_id:
                sm = frappe.db.get_value(
                    "Student Master",
                    {"registration_id": str(student_id)},
                    ["master_programme", "academic_year"],
                    as_dict=True
                )
            programme = (sm.master_programme or "") if sm else ""
            academic_year = (sm.academic_year or "") if sm else ""
            
            unique_offerings[key] = {
                "course": course,
                "batch": batch,
                "programme": programme,
                "academic_year": academic_year,
                "term": r.term_name
            }
            
    out = [["ID", "Course", "Batch", "Programme", "Academic Year", "Term", "Status"]]
    for key, data in unique_offerings.items():
        out.append([
            "",
            data["course"],
            data["batch"],
            data["programme"],
            data["academic_year"],
            data["term"],
            "Active"
        ])
        
    _download_csv(out[0], out[1:], f"Missing_Course_Offerings_Template_{import_log}.csv")
