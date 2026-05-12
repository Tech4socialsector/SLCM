import frappe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def execute():
	cols = [
		{"component": "External", "component_name": "External", "label": "Class Work Assessment", "assessment_type": "CWA", "maximum_marks": 50},
		{"component": "External", "component_name": "External", "label": "Project", "assessment_type": "Proj", "maximum_marks": 10},
		{"component": "Internal", "component_name": "Internal (Custom)", "label": "Summative Assessment", "assessment_type": "SA", "maximum_marks": 40},
	]
	reexam_cols = [
		{"component": "Re Exam", "component_name": "Re exam", "label": "Supplementary Exam", "assessment_type": "Se", "maximum_marks": 100}
	]

	wb = Workbook()
	ws = wb.active

	groups = []
	group_map = {}
	for col in cols:
		comp = col.get("component") or "__none__"
		if comp not in group_map:
			group_map[comp] = {
				"component": comp,
				"component_name": col.get("component_name") or comp,
				"cols": []
			}
			groups.append(group_map[comp])
		group_map[comp]["cols"].append(col)

	rxgroups = []
	rxgroup_map = {}
	for col in (reexam_cols or []):
		comp = col.get("component") or "__rx_none__"
		if comp not in rxgroup_map:
			rxgroup_map[comp] = {
				"component": comp,
				"component_name": col.get("component_name") or comp,
				"cols": []
			}
			rxgroups.append(rxgroup_map[comp])
		rxgroup_map[comp]["cols"].append(col)

	row1, row2, row3 = [], [], []

	student_headers = ["S.No", "Name", "Student RegistrationId", "EmailId"]
	for h in student_headers:
		row1.append(h)
		row2.append("")
		row3.append("")

	col_keys = []
	for g in groups:
		row1.append(g["component_name"])
		for _ in range(len(g["cols"]) * 2 - 1):
			row1.append("")
		for col in g["cols"]:
			lbl = col.get("label") or col.get("type_name") or col.get("assessment_type") or ""
			maxm = col.get("maximum_marks") or 0
			row2.extend([f"{lbl} (Max: {maxm})", ""])
			row3.extend(["Marks", "Revaluation Marks"])
			col_keys.append((col["component"] or "") + "|" + (col["assessment_type"] or ""))

	row1.extend(["Grade", ""])
	row2.extend(["Total Marks", "Grade"])
	row3.extend(["", ""])

	status_headers = ["Enrollment Status", "Attendance Status", "Fairness Status", "SGPA", "Remarks"]
	row1.append("Overall Status")
	row1.extend([""] * (len(status_headers) - 1))
	for h in status_headers:
		row2.append(h)
		row3.append("")
	
	reexam_keys = []
	for g in rxgroups:
		row1.append(f'{g["component_name"]} (Re-Exam)')
		for _ in range(len(g["cols"]) * 2 - 1):
			row1.append("")
		for col in g["cols"]:
			lbl = col.get("label") or col.get("type_name") or col.get("assessment_type") or ""
			maxm = col.get("maximum_marks") or 0
			row2.extend([f"{lbl} (Max: {maxm})", ""])
			row3.extend(["Marks", "Revaluation Marks"])
			reexam_keys.append((col["component"] or "") + "|" + (col["assessment_type"] or ""))

	row1.extend(["Updated Final Result", ""])
	row2.extend(["Updated Final Marks", "Updated Grade"])
	row3.extend(["", ""])

	ws.append(row1)
	ws.append(row2)
	ws.append(row3)

	ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
	ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
	ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3)
	ws.merge_cells(start_row=1, start_column=4, end_row=3, end_column=4)

	c_idx = 5
	for g in groups:
		span = len(g["cols"]) * 2
		if span > 1:
			ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + span - 1)
		for _ in g["cols"]:
			ws.merge_cells(start_row=2, start_column=c_idx, end_row=2, end_column=c_idx + 1)
			c_idx += 2
	
	ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + 1)
	ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
	ws.merge_cells(start_row=2, start_column=c_idx+1, end_row=3, end_column=c_idx+1)
	c_idx += 2

	ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + 4)
	for _ in range(5):
		ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
		c_idx += 1

	for g in rxgroups:
		span = len(g["cols"]) * 2
		if span > 1:
			ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + span - 1)
		for _ in g["cols"]:
			ws.merge_cells(start_row=2, start_column=c_idx, end_row=2, end_column=c_idx + 1)
			c_idx += 2

	ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + 1)
	ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
	ws.merge_cells(start_row=2, start_column=c_idx+1, end_row=3, end_column=c_idx+1)

	hdr_font = Font(bold=True, color="FFFFFF")
	c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
	thin_border = Border(
		left=Side(style='thin', color='CCCCCC'),
		right=Side(style='thin', color='CCCCCC'),
		top=Side(style='thin', color='CCCCCC'),
		bottom=Side(style='thin', color='CCCCCC')
	)

	fill_r1 = PatternFill("solid", fgColor="4338CA")
	fill_r2 = PatternFill("solid", fgColor="4F46E5") 
	fill_r3 = PatternFill("solid", fgColor="6366F1")

	for r in range(1, 4):
		for c in range(1, len(row1)+1):
			cell = ws.cell(row=r, column=c)
			cell.font = hdr_font
			cell.alignment = c_align
			cell.border = thin_border
			if r == 1:
				cell.fill = fill_r1
			elif r == 2:
				cell.fill = fill_r2
			else:
				cell.fill = fill_r3
			if c <= 4:
				cell.fill = fill_r1

	for c in range(5, len(row1) + 1):
		ws.column_dimensions[get_column_letter(c)].width = 18
	ws.column_dimensions["A"].width = 8
	ws.column_dimensions["B"].width = 30
	ws.column_dimensions["C"].width = 25
	ws.column_dimensions["D"].width = 35

	print(f"Generated {len(row1)} columns.")
