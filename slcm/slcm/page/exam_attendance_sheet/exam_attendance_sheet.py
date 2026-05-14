# Copyright (c) 2026, TFSS and contributors
# For license information, please see license.txt

import frappe


@frappe.whitelist()
def get_exam_plans():
	"""Return list of active exam plans."""
	return frappe.db.get_all(
		"Exam Plan",
		filters={"status": "Active"},
		fields=["name", "exam_name", "term"],
		order_by="exam_name asc",
	)


@frappe.whitelist()
def get_exam_courses(exam_plan):
	"""Return courses scheduled under the given exam plan, with dates and barcode counts."""
	schedules = frappe.db.get_all(
		"Exam Course Schedule",
		filters={"parent": exam_plan, "parenttype": "Exam Plan"},
		fields=["course", "exam_date", "start_time", "end_time", "venue", "hall"],
		order_by="exam_date asc, course asc",
	)

	if not schedules:
		return []

	# Fetch course names and codes
	course_info = {
		r.name: {"course_name": r.course_name, "course_code": r.course_code or r.name}
		for r in frappe.db.get_all(
			"Course",
			filters={"name": ["in", [s["course"] for s in schedules]]},
			fields=["name", "course_name", "course_code"],
		)
	}

	# Count existing barcodes per course using raw SQL (group_by with COUNT)
	barcode_counts = {}
	rows = frappe.db.sql(
		"""
		SELECT course, COUNT(name) AS cnt
		FROM `tabExam Barcode`
		WHERE exam_plan = %s
		GROUP BY course
		""",
		(exam_plan,),
		as_dict=True,
	)
	for r in rows:
		barcode_counts[r.course] = r.cnt

	for s in schedules:
		info = course_info.get(s["course"], {})
		s["course_name"] = info.get("course_name", s["course"])
		s["course_code"] = info.get("course_code", s["course"])
		s["barcode_count"] = barcode_counts.get(s["course"], 0)

	return schedules


@frappe.whitelist()
def generate_barcodes(exam_plan, courses=None):
	"""
	Generate 6-digit unique barcodes for all students enrolled in the given
	courses under the exam plan. Existing barcodes are preserved; only missing
	students get new ones.

	courses: JSON list of course names (optional, defaults to all scheduled).
	"""
	import json
	import random
	import datetime

	if isinstance(courses, str):
		courses = json.loads(courses)

	if not courses:
		schedules = frappe.db.get_all(
			"Exam Course Schedule",
			filters={"parent": exam_plan, "parenttype": "Exam Plan"},
			fields=["course", "exam_date"],
		)
		courses = list({s.course for s in schedules})

	# Build exam_date lookup per course
	date_lookup = {}
	for s in frappe.db.get_all(
		"Exam Course Schedule",
		filters={"parent": exam_plan, "parenttype": "Exam Plan", "course": ["in", courses]},
		fields=["course", "exam_date"],
	):
		date_lookup[s.course] = s.exam_date

	# Collect all barcodes already issued under this plan (for uniqueness)
	used_barcodes = set(
		r.barcode
		for r in frappe.db.get_all("Exam Barcode", filters={"exam_plan": exam_plan}, fields=["barcode"])
	)

	def _new_barcode():
		for _ in range(10000):
			code = str(random.randint(100000, 999999))
			if code not in used_barcodes:
				used_barcodes.add(code)
				return code
		frappe.throw("Could not generate a unique 6-digit barcode. Barcode pool may be exhausted.")

	now = frappe.utils.now_datetime()
	created = 0

	for course in courses:
		students = _get_students_for_course(course, exam_plan)
		if not students:
			continue

		# Find which students already have barcodes
		existing = {
			r.student
			for r in frappe.db.get_all(
				"Exam Barcode",
				filters={"exam_plan": exam_plan, "course": course},
				fields=["student"],
			)
		}

		for st in students:
			if st["student"] in existing:
				continue
			doc = frappe.get_doc({
				"doctype": "Exam Barcode",
				"exam_plan": exam_plan,
				"course": course,
				"exam_date": date_lookup.get(course),
				"student": st["student"],
				"student_name": st["student_name"],
				"registration_id": st["registration_id"],
				"section": st["section"],
				"barcode": _new_barcode(),
				"generated_on": now,
			})
			doc.insert(ignore_permissions=True)
			created += 1

	frappe.db.commit()
	return {"created": created, "message": f"{created} new barcode(s) generated."}


@frappe.whitelist()
def get_barcodes(exam_plan, course=None):
	"""Return all barcodes for the exam plan (optionally filtered by course)."""
	filters = {"exam_plan": exam_plan}
	if course:
		filters["course"] = course

	return frappe.db.get_all(
		"Exam Barcode",
		filters=filters,
		fields=["course", "exam_date", "student", "student_name", "registration_id", "section", "barcode"],
		order_by="course asc, registration_id asc",
	)


@frappe.whitelist()
def get_course_students(exam_plan, course):
	"""
	Return students for a course with their barcode (if generated).
	Source: Student Course Marks (same as Examination Result page).
	"""
	students = _get_students_for_course(course, exam_plan)

	# Existing barcodes
	existing_barcodes = {
		r.student: r.barcode
		for r in frappe.db.get_all(
			"Exam Barcode",
			filters={"exam_plan": exam_plan, "course": course},
			fields=["student", "barcode"],
		)
	}

	result = []
	for st in students:
		result.append({
			"student": st["student"],
			"student_name": st["student_name"],
			"registration_id": st["registration_id"],
			"section": st["section"],
			"barcode": existing_barcodes.get(st["student"], ""),
			"has_barcode": st["student"] in existing_barcodes,
		})

	result.sort(key=lambda x: (x["registration_id"] or "zzz", x["student_name"]))
	return result


@frappe.whitelist()
def export_attendance_excel(exam_plan):
	"""
	Generate the exam attendance Excel with one sheet per exam date.
	Each sheet lists all courses for that date with student rows.
	Returns a base64-encoded .xlsx file.
	"""
	import io
	import base64
	import datetime
	import openpyxl
	from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
	from openpyxl.utils import get_column_letter

	plan_doc = frappe.get_doc("Exam Plan", exam_plan)

	# Group schedules by exam_date
	schedules = frappe.db.get_all(
		"Exam Course Schedule",
		filters={"parent": exam_plan, "parenttype": "Exam Plan"},
		fields=["course", "exam_date", "start_time", "end_time", "venue", "hall"],
		order_by="exam_date asc, course asc",
	)

	if not schedules:
		frappe.throw("No courses scheduled in this Exam Plan.")

	# Fetch course names and codes
	course_name_map = {}
	for r in frappe.db.get_all(
		"Course",
		filters={"name": ["in", list({s["course"] for s in schedules})]},
		fields=["name", "course_name", "course_code"],
	):
		label = r.course_name or r.name
		if r.course_code:
			label = f"{label} [{r.course_code}]"
		course_name_map[r.name] = label

	# Fetch all barcodes for this plan
	barcodes = frappe.db.get_all(
		"Exam Barcode",
		filters={"exam_plan": exam_plan},
		fields=["course", "exam_date", "student", "student_name", "registration_id", "section", "barcode"],
		order_by="registration_id asc",
	)
	# Map: course -> sorted list of student rows
	barcode_map = {}
	for b in barcodes:
		barcode_map.setdefault(b.course, []).append(b)

	# Get institution info from Academic Term if available
	institution = "Institution"
	term_label = ""
	try:
		if plan_doc.term:
			term_doc = frappe.get_doc("Academic Term", plan_doc.term)
			term_label = term_doc.term_name or plan_doc.term
			if hasattr(term_doc, "academic_year"):
				yd = frappe.db.get_value("Academic Year", term_doc.academic_year, "year_name") or term_doc.academic_year
				term_label = f"{term_label}, AY {yd}"
	except Exception:
		term_label = plan_doc.term or ""

	# Group schedules by date
	from collections import defaultdict
	date_groups = defaultdict(list)
	for s in schedules:
		date_groups[s.exam_date].append(s)

	wb = openpyxl.Workbook()
	wb.remove(wb.active)  # remove default sheet

	# Styles
	thin = Side(style="thin")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	def _hdr_font(size=12, bold=False):
		return Font(name="Bookman Old Style", size=size, bold=bold)

	def _cell_font():
		return Font(name="Bookman Old Style", size=11)

	def _center(cell):
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

	def _set_col_widths(ws, widths):
		for i, w in enumerate(widths, 1):
			ws.column_dimensions[get_column_letter(i)].width = w

	sorted_dates = sorted(date_groups.keys(), key=lambda d: (d or ""))

	for exam_date in sorted_dates:
		date_str = ""
		if exam_date:
			try:
				date_str = frappe.utils.formatdate(str(exam_date), "dd/MM/yyyy")
			except Exception:
				date_str = str(exam_date)

		sheet_name = date_str.replace("/", "-") if date_str else "No Date"
		ws = wb.create_sheet(title=sheet_name[:31])

		day_schedules = date_groups[exam_date]
		current_row = 1

		for sched in day_schedules:
			course = sched.course
			course_display = course_name_map.get(course, course)
			students = sorted(barcode_map.get(course, []), key=lambda x: (x.registration_id or ""))

			# ── Header block ────────────────────────────────────────────────
			# Row 1: Institution name
			ws.merge_cells(
				start_row=current_row, start_column=1,
				end_row=current_row, end_column=8
			)
			c = ws.cell(row=current_row, column=1, value=institution)
			c.font = _hdr_font(14, bold=True)
			_center(c)
			ws.row_dimensions[current_row].height = 22
			current_row += 1

			# Row 2: Term / Year info
			ws.merge_cells(
				start_row=current_row, start_column=1,
				end_row=current_row, end_column=8
			)
			c = ws.cell(row=current_row, column=1, value=term_label)
			c.font = _hdr_font(12, bold=True)
			_center(c)
			ws.row_dimensions[current_row].height = 18
			current_row += 1

			# Row 3: Course + Date
			exam_line = f"{course_display} End Term Examination  ({date_str})"
			ws.merge_cells(
				start_row=current_row, start_column=1,
				end_row=current_row, end_column=8
			)
			c = ws.cell(row=current_row, column=1, value=exam_line)
			c.font = _hdr_font(11, bold=True)
			_center(c)
			ws.row_dimensions[current_row].height = 16
			current_row += 1

			# Row 4: Column headers
			col_headers = [
				"Sl.No", "Name", "ID no.", "Sec", "Barcode",
				"Answer Booklet\nnumbers", "Student\nSignature",
				"No. of Additional\nBooklet Used"
			]
			fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
			for col_idx, header in enumerate(col_headers, 1):
				c = ws.cell(row=current_row, column=col_idx, value=header)
				c.font = Font(name="Bookman Old Style", size=11, bold=True)
				c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
				c.border = border
				c.fill = fill
			ws.row_dimensions[current_row].height = 35
			current_row += 1

			# Rows 5+: Student data
			if students:
				for sl, st in enumerate(students, 1):
					row_data = [
						sl,
						st.student_name or "",
						st.registration_id or "",
						st.section or "",
						st.barcode or "",
						"",  # Answer Booklet numbers (blank for faculty)
						"",  # Student Signature
						"",  # No. of Additional Booklet Used
					]
					for col_idx, val in enumerate(row_data, 1):
						c = ws.cell(row=current_row, column=col_idx, value=val)
						c.font = _cell_font()
						c.border = border
						if col_idx in (1, 3, 4, 5, 6, 8):
							_center(c)
						else:
							c.alignment = Alignment(horizontal="left", vertical="center")
					ws.row_dimensions[current_row].height = 25
					current_row += 1
			else:
				# No barcodes generated yet - placeholder row
				ws.merge_cells(
					start_row=current_row, start_column=1,
					end_row=current_row, end_column=8
				)
				c = ws.cell(row=current_row, column=1,
							value="No barcodes generated for this course yet.")
				c.font = Font(name="Bookman Old Style", size=11, italic=True, color="FF0000")
				_center(c)
				current_row += 1

			# Gap between courses on same date
			current_row += 2

		# Set column widths (once per sheet, applies to all courses)
		_set_col_widths(ws, [6, 28, 14, 8, 10, 16, 16, 16])

	# Save to buffer
	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)
	file_b64 = base64.b64encode(buf.read()).decode()
	filename = f"Exam_Attendance_{exam_plan.replace(' ', '_')}.xlsx"
	return {"file_content": file_b64, "filename": filename}


# ── Internal helpers ────────────────────────────────────────────────────────────

def _get_students_for_course(course, exam_plan):
	"""
	Return list of dicts {student, student_name, registration_id, section}
	sourced from Student Course Marks — the same table used by Examination Result.
	Students are sorted by registration_id ascending.
	"""
	rows = frappe.db.sql(
		"""
		SELECT
			scm.student,
			CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
			COALESCE(NULLIF(sm.registration_id, ''), sm.name) AS registration_id,
			COALESCE(cc.section, '') AS section
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		LEFT JOIN (
			SELECT cs.student, cc2.section, cc2.course
			FROM `tabClass Student` cs
			INNER JOIN `tabClass Configuration` cc2 ON cc2.name = cs.parent
		) cc ON cc.student = scm.student AND cc.course = scm.course
		WHERE scm.exam_plan = %(exam_plan)s
		  AND scm.course = %(course)s
		ORDER BY COALESCE(NULLIF(sm.registration_id, ''), sm.name) ASC
		""",
		{"exam_plan": exam_plan, "course": course},
		as_dict=True,
	)

	# Deduplicate — a student could appear in multiple class configurations
	seen = set()
	result = []
	for r in rows:
		if r.student in seen:
			continue
		seen.add(r.student)
		result.append({
			"student": r.student,
			"student_name": r.student_name or r.student,
			"registration_id": r.registration_id or "",
			"section": r.section or "",
		})

	return result
