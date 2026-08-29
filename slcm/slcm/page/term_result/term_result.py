# Copyright (c) 2026, TFSS and contributors
# For license information, please see license.txt

import frappe
from collections import defaultdict


# ── Helpers ────────────────────────────────────────────────────────────────────

def _build_student_filter(exam_plan, search="", inst_programmes="", inst_batches=""):
	"""Build the WHERE clause fragment + params for filtering students in an exam plan."""
	params = {"exam_plan": exam_plan}
	extra_cond = ""
	if search:
		extra_cond += (
			" AND (sm.registration_id LIKE %(search)s"
			" OR sm.first_name LIKE %(search)s"
			" OR sm.last_name LIKE %(search)s)"
		)
		params["search"] = f"%{search}%"

	f_programmes = frappe.parse_json(inst_programmes) if inst_programmes else []
	f_batches    = frappe.parse_json(inst_batches)    if inst_batches    else []
	if f_programmes:
		placeholders = ",".join([f"%(prog_{i})s" for i in range(len(f_programmes))])
		extra_cond += f" AND sm.programme_of_study IN ({placeholders})"
		for i, v in enumerate(f_programmes):
			params[f"prog_{i}"] = v
	if f_batches:
		placeholders = ",".join([f"%(batch_{i})s" for i in range(len(f_batches))])
		extra_cond += f" AND sm.batch_year IN ({placeholders})"
		for i, v in enumerate(f_batches):
			params[f"batch_{i}"] = v

	return extra_cond, params


def _get_matching_student_ids(exam_plan, search="", inst_programmes="", inst_batches=""):
	"""Return all student ids enrolled in the exam plan that match the given filters (no pagination)."""
	extra_cond, params = _build_student_filter(exam_plan, search, inst_programmes, inst_batches)
	rows = frappe.db.sql(
		f"""
		SELECT DISTINCT scm.student AS student
		FROM `tabStudent Course Marks` scm
		INNER JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.exam_plan = %(exam_plan)s
		{extra_cond}
		""",
		params, as_dict=True,
	)
	return [r["student"] for r in rows]


def _compute_term_gpa(exam_plan, student_names, students, force_recompute=False):
	"""Compute SGPA, term percentage and cumulative GPA for each student in-place."""
	if not student_names:
		return

	placeholders = ",".join(["%s"] * len(student_names))

	marks_rows = frappe.db.sql(
		f"""
		SELECT
			scm.student,
			scm.course,
			COALESCE(NULLIF(scm.updated_grade, ''), NULLIF(scm.grade, '')) AS final_grade,
			CASE WHEN scm.updated_grade IS NOT NULL AND scm.updated_grade != '' THEN scm.updated_final_marks ELSE scm.total_marks END AS final_marks,
			scm.consider_for_sgpa,
			scm.enrollment_status,
			COALESCE(c.credit_value, 0)                                     AS credit_value,
			COALESCE(gs.maximum_marks, 100)                                 AS maximum_marks,
			gsc.grade_point                                                 AS grade_point,
			COALESCE(gsc.failed, 0)                                         AS is_failed
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabCourse` c ON c.name = scm.course
		LEFT JOIN `tabCourse Schema Assignment` csa
			ON csa.exam_plan = %s AND csa.course = scm.course
		LEFT JOIN `tabGrading Schema` gs ON gs.name = csa.grade_schema
		LEFT JOIN `tabGrading Schema Component` gsc
			ON gsc.parent = csa.grade_schema
			AND gsc.parentfield = 'grades'
			AND gsc.grade = COALESCE(NULLIF(scm.updated_grade,''), NULLIF(scm.grade,''))
		WHERE scm.exam_plan = %s
		  AND scm.student IN ({placeholders})
		  AND COALESCE(scm.enrollment_status,'') NOT IN ('Dropped','Detained','Migrated')
		""",
		[exam_plan, exam_plan] + list(student_names),
		as_dict=True,
	)

	# Group marks by student
	student_marks = defaultdict(list)
	for m in marks_rows:
		student_marks[m.student].append(m)

	student_map = {s["student"]: s for s in students}

	# First, get saved results from Student Result Publish
	saved_results = frappe.db.sql("""
		SELECT student, term_gpa, term_percentage, cumulative_gpa, cumulative_percentage
		FROM `tabStudent Result Publish`
		WHERE exam_plan = %s AND student IN ({})
	""".format(",".join(["%s"] * len(student_names))), [exam_plan] + list(student_names), as_dict=True)
	
	saved_map = {r["student"]: r for r in saved_results}

	for student_id, marks in student_marks.items():
		s = student_map.get(student_id)
		if not s:
			continue

		# If a saved Student Result Publish record exists, use its values directly.
		# NULL in DB → None → "Not Generated"; 0.0 in DB → 0.0 → shows "0.00"
		# (0.0 is a valid GPA/percentage for a student who failed all courses).
		if not force_recompute and student_id in saved_map:
			sr = saved_map[student_id]
			s["term_gpa"] = sr.get("term_gpa")
			s["term_percentage"] = sr.get("term_percentage")
			s["cumulative_gpa"] = sr.get("cumulative_gpa")
			s["cumulative_percentage"] = sr.get("cumulative_percentage")
			continue

		weighted_gp   = 0.0
		total_credits = 0.0
		all_weighted_gp = 0.0   # fallback: all graded courses
		all_credits     = 0.0
		total_marks   = 0.0
		total_max     = 0.0
		graded_count  = 0
		total_count   = len(marks)

		for m in marks:
			# grade_point is NULL when the grade couldn't be resolved against the
			# course's Grading Schema Component (e.g. no Course Schema Assignment,
			# or the grade string doesn't match any component) — such a course must
			# be excluded from the weighted average rather than silently treated as
			# grade_point 0, which would drag the GPA down to a false low value.
			if m["consider_for_sgpa"] and m["final_grade"] and m["credit_value"] and m["grade_point"] is not None:
				weighted_gp   += float(m["grade_point"]) * float(m["credit_value"])
				total_credits += float(m["credit_value"])
			if m["final_grade"] and m["credit_value"] and m["grade_point"] is not None:
				all_weighted_gp += float(m["grade_point"]) * float(m["credit_value"])
				all_credits     += float(m["credit_value"])
			if m["final_grade"]:
				graded_count += 1
			total_marks += float(m["final_marks"] or 0)
			total_max   += float(m["maximum_marks"] or 100)

		all_graded = graded_count == total_count and total_count > 0

		if total_credits > 0 and all_graded:
			term_gpa = round(weighted_gp / total_credits, 2)
		elif all_credits > 0 and all_graded:
			# No courses flagged consider_for_sgpa — compute from all graded courses
			term_gpa = round(all_weighted_gp / all_credits, 2)
		else:
			term_gpa = None

		s["term_gpa"]        = term_gpa
		s["term_percentage"] = round((total_marks / total_max) * 100, 2) if (total_max > 0 and all_graded) else None
		s["cumulative_gpa"]  = s.get("current_cgpa") or None
		s["cumulative_percentage"] = s.get("cumulative_percentage") or None

# ── Actions ───────────────────────────────────────────────────────────────────

def _compute_cumulative_stats(student_id):
	marks_rows = frappe.db.sql(
		"""
		SELECT
			scm.course,
			COALESCE(NULLIF(scm.updated_grade, ''), NULLIF(scm.grade, '')) AS final_grade,
			CASE WHEN scm.updated_grade IS NOT NULL AND scm.updated_grade != '' THEN scm.updated_final_marks ELSE scm.total_marks END AS final_marks,
			scm.consider_for_sgpa,
			scm.enrollment_status,
			COALESCE(c.credit_value, 0)                                     AS credit_value,
			COALESCE(gs.maximum_marks, 100)                                 AS maximum_marks,
			gsc.grade_point                                                 AS grade_point
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabCourse` c ON c.name = scm.course
		LEFT JOIN `tabCourse Schema Assignment` csa
			ON csa.exam_plan = scm.exam_plan AND csa.course = scm.course
		LEFT JOIN `tabGrading Schema` gs ON gs.name = csa.grade_schema
		LEFT JOIN `tabGrading Schema Component` gsc
			ON gsc.parent = csa.grade_schema
			AND gsc.parentfield = 'grades'
			AND gsc.grade = COALESCE(NULLIF(scm.updated_grade,''), NULLIF(scm.grade,''))
		WHERE scm.student = %s
		  AND COALESCE(scm.enrollment_status,'') NOT IN ('Dropped','Detained','Migrated')
		""",
		(student_id,), as_dict=True
	)

	weighted_gp = 0.0
	total_credits = 0.0
	total_marks = 0.0
	total_max = 0.0

	for m in marks_rows:
		if m["consider_for_sgpa"] and m["final_grade"] and m["credit_value"] and m["grade_point"] is not None:
			weighted_gp += float(m["grade_point"]) * float(m["credit_value"])
			total_credits += float(m["credit_value"])
		if m["final_grade"]:
			# only include graded courses in total percentage
			total_marks += float(m["final_marks"] or 0)
			total_max += float(m["maximum_marks"] or 100)

	cgpa = round(weighted_gp / total_credits, 2) if total_credits > 0 else None
	cpct = round((total_marks / total_max) * 100, 2) if total_max > 0 else None

	# If CGPA can't be computed from marks, read the saved value from Student Master
	if cgpa is None:
		try:
			sm_val = frappe.db.get_value("Student Master", student_id, "current_cgpa")
			if sm_val and float(sm_val) > 0:
				cgpa = round(float(sm_val), 2)
		except Exception:
			pass

	return cgpa, cpct

@frappe.whitelist()
def generate_term_results(exam_plan, student_names, action, select_all=0,
                           exclude_students="[]", search="", inst_programmes="", inst_batches=""):
	import json

	if frappe.utils.cint(select_all):
		exclude_list = set(json.loads(exclude_students) or [])
		student_list = [
			s for s in _get_matching_student_ids(exam_plan, search, inst_programmes, inst_batches)
			if s not in exclude_list
		]
		if not student_list:
			return
	else:
		student_list = json.loads(student_names)
		if not student_list:
			student_list = [r["student"] for r in frappe.db.sql("SELECT DISTINCT student FROM `tabStudent Course Marks` WHERE exam_plan=%s", exam_plan, as_dict=True)]
			if not student_list:
				return

	# compute dynamic term gpa
	students = frappe.db.sql(
		"SELECT name as student, current_cgpa, cumulative_percentage FROM `tabStudent Master` WHERE name in %s",
		(tuple(student_list),), as_dict=True
	)
	student_map = {s["student"]: s for s in students}
	
	# only compute term if term action
	if action in ["term_gpa", "term_percentage"]:
		# force recalculation for map logic
		_compute_term_gpa(exam_plan, student_list, students, force_recompute=True)

	for student_id in student_list:
		# Ensure a Student Result Publish record exists, then use db.set_value
		# directly for all actions — bypasses before_save which would overwrite
		# values with marks-based recalculation (returns 0 when no consider_for_sgpa courses).
		doc_name = frappe.db.get_value(
			"Student Result Publish",
			{"exam_plan": exam_plan, "student": student_id},
			"name",
		)
		if not doc_name:
			doc = frappe.new_doc("Student Result Publish")
			doc.exam_plan = exam_plan
			doc.student = student_id
			doc.flags.ignore_permissions = True
			doc.insert(ignore_permissions=True)
			doc_name = doc.name
			# term_gpa/term_percentage/cumulative_gpa/cumulative_percentage are Float
			# columns, which Frappe always creates as NOT NULL DEFAULT 0 — they can't
			# hold NULL, so a freshly inserted record already reads as 0.0 for every
			# field until the action below sets a real value.

		s_data = student_map.get(student_id, {})

		if action == "term_gpa":
			val = s_data.get("term_gpa")
			if val is not None:
				frappe.db.set_value("Student Result Publish", doc_name, "term_gpa", val, update_modified=False)
		elif action == "term_percentage":
			val = s_data.get("term_percentage")
			if val is not None:
				frappe.db.set_value("Student Result Publish", doc_name, "term_percentage", val, update_modified=False)
		elif action in ["cumulative_gpa", "cumulative_percentage"]:
			cgpa, cpct = _compute_cumulative_stats(student_id)

			if action == "cumulative_gpa":
				frappe.db.set_value("Student Result Publish", doc_name, "cumulative_gpa", cgpa or 0, update_modified=False)
				if cgpa is not None and cgpa > 0:
					frappe.db.set_value("Student Master", student_id, "current_cgpa", cgpa, update_modified=False)
			else:
				from slcm.slcm.doctype.cgpa_percentage_scale.cgpa_percentage_scale import (
					lookup_percentage_for_cgpa,
				)
				scale_pct = lookup_percentage_for_cgpa(cgpa) if cgpa is not None else None
				final_pct = scale_pct if scale_pct is not None else cpct
				frappe.db.set_value("Student Result Publish", doc_name, "cumulative_percentage", final_pct or 0, update_modified=False)
				if final_pct is not None and final_pct > 0:
					frappe.db.set_value("Student Master", student_id, "cumulative_percentage", final_pct, update_modified=False)
	
	frappe.db.commit()
	return "Success"

@frappe.whitelist()
def download_consolidated_report(exam_plan="", search="", inst_programmes="", inst_batches="",
		course="", course_offering="", academic_year="", programme="", trimester="", batch="", year=""):

	f_programmes = frappe.parse_json(inst_programmes) if inst_programmes else []
	f_batches    = frappe.parse_json(inst_batches)    if inst_batches    else []

	params     = {}
	where_cond = "1=1"

	if exam_plan:
		where_cond += " AND scm.exam_plan = %(exam_plan)s"
		params["exam_plan"] = exam_plan

	# course_offering is the modern, precise filter (Course now only lives on
	# the offering); "course" is kept for backwards compatibility only.
	if course_offering:
		where_cond += " AND scm.course_offering = %(course_offering)s"
		params["course_offering"] = course_offering
	elif course:
		where_cond += " AND scm.course = %(course)s"
		params["course"] = course

	if academic_year:
		where_cond += " AND sm.academic_year = %(academic_year)s"
		params["academic_year"] = academic_year

	if trimester:
		where_cond += " AND sm.academic_term = %(trimester)s"
		params["trimester"] = trimester

	if batch:
		where_cond += " AND sm.batch = %(batch)s"
		params["batch"] = batch

	if year:
		where_cond += " AND sm.term_year = %(year)s"
		params["year"] = year

	if programme:
		where_cond += " AND coh.program = %(programme)s"
		params["programme"] = programme

	if search:
		where_cond += (
			" AND (sm.registration_id LIKE %(search)s"
			" OR sm.first_name LIKE %(search)s"
			" OR sm.last_name LIKE %(search)s)"
		)
		params["search"] = f"%{search}%"

	if f_programmes:
		placeholders = ",".join([f"%(prog_{i})s" for i in range(len(f_programmes))])
		where_cond += f" AND sm.programme_of_study IN ({placeholders})"
		for i, v in enumerate(f_programmes):
			params[f"prog_{i}"] = v

	if f_batches:
		placeholders = ",".join([f"%(batch_{i})s" for i in range(len(f_batches))])
		where_cond += f" AND sm.batch_year IN ({placeholders})"
		for i, v in enumerate(f_batches):
			params[f"batch_{i}"] = v

	if not params:
		frappe.throw("Please select at least one filter to download the report.")

	rows = frappe.db.sql(
		f"""
		SELECT
			sm.registration_id,
			TRIM(CONCAT_WS(' ', sm.first_name, COALESCE(NULLIF(sm.middle_name,''), NULL), sm.last_name)) AS student_name,
			sm.academic_year,
			coh.program AS programme_name,
			sm.specialisation,
			sm.batch AS batch_year,
			sm.academic_term AS trimester,
			c.course_code,
			c.course_name,
			'' AS course_registration_type,
			'' AS exam_type,
			csa.evaluation_schema,
			csa.grade_schema,
			scm.total_marks,
			scm.grade,
			gsc.grade_point,
			gsc.failed AS is_failed,
			scm.attendance_status,
			scm.consider_for_sgpa,
			scm.updated_final_marks,
			scm.updated_grade,
			gsc2.grade_point AS updated_grade_point,
			gsc2.failed AS updated_is_failed,
			'' AS cgpa_sgpa_rule,
			srp.term_gpa AS sgpa,
			srp.cumulative_gpa AS cgpa
		FROM `tabStudent Course Marks` scm
		INNER JOIN `tabStudent Master` sm ON sm.name = scm.student
		LEFT JOIN `tabBatch` coh ON coh.name = sm.batch
		LEFT JOIN `tabCourse` c ON c.name = scm.course
		LEFT JOIN `tabCourse Schema Assignment` csa ON csa.exam_plan = scm.exam_plan AND csa.course = scm.course
		LEFT JOIN `tabGrading Schema Component` gsc ON gsc.parent = csa.grade_schema AND gsc.grade = COALESCE(NULLIF(scm.grade,''), NULL)
		LEFT JOIN `tabGrading Schema Component` gsc2 ON gsc2.parent = csa.grade_schema AND gsc2.grade = COALESCE(NULLIF(scm.updated_grade,''), NULL)
		LEFT JOIN `tabStudent Result Publish` srp ON srp.exam_plan = scm.exam_plan AND srp.student = scm.student
		WHERE {where_cond}
		ORDER BY c.course_name ASC, sm.registration_id ASC
		""", params, as_dict=True)

	if not rows:
		frappe.throw("No records found for the selected filters. Try widening the Exam Plan, Academic Year, Programme, Trimester or Batch selection.")

	headers = [
		"Registration ID", "Student Name", "Academic Year", "Programme Name",
		"Programme Specialization", "Batch", "Trimester",
		"Course Code", "Course Name", "Course Registration Type", "Exam Type",
		"Evaluation Schema", "Grade Schema", "Total Marks", "Grade", "Grade Points",
		"Is Failed", "Attendance Status", "Consider For SGPA Calculation",
		"Re Exam Total Marks", "Re Exam Grade", "Re Exam Grade Point",
		"Is Failed For Re Exam", "CGPA SGPA Rule", "SGPA", "CGPA"
	]

	def _make_row(r):
		return [
			r.get("registration_id"),
			r.get("student_name"),
			r.get("academic_year"),
			r.get("programme_name"),
			r.get("specialisation"),
			r.get("batch_year"),
			r.get("trimester"),
			r.get("course_code"),
			r.get("course_name"),
			r.get("course_registration_type"),
			r.get("exam_type"),
			r.get("evaluation_schema"),
			r.get("grade_schema"),
			r.get("total_marks"),
			r.get("grade"),
			r.get("grade_point"),
			1 if r.get("is_failed") else 0,
			r.get("attendance_status"),
			1 if r.get("consider_for_sgpa") else 0,
			r.get("updated_final_marks"),
			r.get("updated_grade"),
			r.get("updated_grade_point"),
			1 if r.get("updated_is_failed") else 0,
			r.get("cgpa_sgpa_rule"),
			r.get("sgpa"),
			r.get("cgpa"),
		]

	# ── Build Excel with Overall sheet + one sheet per course ─────────────────
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment
	except ImportError:
		frappe.throw("openpyxl is required. Run: bench pip install openpyxl")

	import io
	from collections import OrderedDict

	wb = openpyxl.Workbook()
	hdr_font  = Font(bold=True, color="FFFFFF")
	hdr_fill  = PatternFill("solid", fgColor="B24040")
	hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

	def _write_sheet(ws, data_rows):
		ws.append(headers)
		for cell in ws[1]:
			cell.font  = hdr_font
			cell.fill  = hdr_fill
			cell.alignment = hdr_align
		for row_data in data_rows:
			ws.append(row_data)
		for col in ws.columns:
			ws.column_dimensions[col[0].column_letter].width = 18

	# Overall sheet
	ws_all = wb.active
	ws_all.title = "Overall"
	_write_sheet(ws_all, [_make_row(r) for r in rows])

	# Per-course sheets (only when more than one course present)
	course_groups = OrderedDict()
	for r in rows:
		cn = (r.get("course_name") or "Unknown")[:31]
		course_groups.setdefault(cn, []).append(r)

	if len(course_groups) > 1:
		for cn, course_rows_list in course_groups.items():
			ws = wb.create_sheet(title=cn)
			_write_sheet(ws, [_make_row(r) for r in course_rows_list])

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	# Stream the file directly so the browser downloads it immediately
	import hashlib, time
	suffix = hashlib.md5(str(time.time()).encode()).hexdigest()[:6]
	fname  = f"Consolidated_Report_{suffix}.xlsx"

	frappe.response.filename    = fname
	frappe.response.filecontent = buf.read()
	frappe.response.type        = "download"
	frappe.response.display_content_as = "attachment"



# ── Public API ─────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_term_stats(exam_plan):
	"""Return summary statistics for the exam plan."""
	if not exam_plan:
		return {}

	total = frappe.db.sql(
		"SELECT COUNT(DISTINCT student) AS cnt FROM `tabStudent Course Marks` WHERE exam_plan=%s",
		exam_plan, as_dict=True,
	)[0]["cnt"] or 0

	total_courses = frappe.db.sql(
		"SELECT COUNT(DISTINCT course) AS cnt FROM `tabStudent Course Marks` WHERE exam_plan=%s",
		exam_plan, as_dict=True,
	)[0]["cnt"] or 0

	# Graded = every course for that student has a non-empty grade
	graded_row = frappe.db.sql(
		"""
		SELECT COUNT(*) AS cnt FROM (
			SELECT student
			FROM `tabStudent Course Marks`
			WHERE exam_plan = %s
			GROUP BY student
			HAVING SUM(CASE WHEN COALESCE(NULLIF(grade,''),NULL) IS NULL THEN 1 ELSE 0 END) = 0
		) t
		""",
		exam_plan, as_dict=True,
	)
	graded = graded_row[0]["cnt"] if graded_row else 0

	# Average SGPA (only from students who have it calculated)
	avg_gpa = frappe.db.sql(
		"""
		SELECT AVG(current_cgpa) AS avg_gpa
		FROM `tabStudent Master` sm
		INNER JOIN `tabStudent Course Marks` scm ON scm.student = sm.name
		WHERE scm.exam_plan = %s AND sm.current_cgpa IS NOT NULL AND sm.current_cgpa > 0
		""",
		exam_plan, as_dict=True,
	)
	avg_gpa_val = avg_gpa[0]["avg_gpa"] if avg_gpa else None

	return {
		"total_students": total,
		"total_courses":  total_courses,
		"graded":         graded,
		"not_graded":     total - graded,
		"avg_cgpa":       round(float(avg_gpa_val), 2) if avg_gpa_val else None,
	}


@frappe.whitelist()
def get_exam_plans(search=None):
	"""Return exam plans for the plan selector."""
	filters = {}
	if search:
		filters["exam_name"] = ["like", f"%{search}%"]
	return frappe.get_all(
		"Exam Plan",
		filters=filters,
		fields=["name", "exam_name", "term", "status"],
		order_by="creation desc",
	)


@frappe.whitelist()
def get_term_inst_filter_options(exam_plan):
	"""Return distinct programmes and batches for students in this exam plan."""
	if not exam_plan:
		return {"programmes": [], "batches": []}
	rows = frappe.db.sql(
		"""
		SELECT DISTINCT sm.programme_of_study AS programme, sm.batch_year
		FROM `tabStudent Course Marks` scm
		INNER JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.exam_plan = %(exam_plan)s
		""",
		{"exam_plan": exam_plan},
		as_dict=True,
	)
	programmes = sorted(set(r["programme"] for r in rows if r.get("programme")))
	batches    = sorted(set(str(r["batch_year"]) for r in rows if r.get("batch_year")), reverse=True)
	return {"programmes": programmes, "batches": batches}


@frappe.whitelist()
def get_term_students(exam_plan, search="", page=1, page_length=20,
                      sort_by="registration_id", sort_order="asc",
                      inst_programmes="", inst_batches=""):
	"""Return paginated students enrolled in the exam plan with term result data."""
	if not exam_plan:
		return {"students": [], "total": 0}

	page        = int(page)
	page_length = int(page_length)
	offset      = (page - 1) * page_length
	sort_dir    = "DESC" if sort_order == "desc" else "ASC"

	sort_col_map = {
		"registration_id": "sm.registration_id",
		"name":            "CONCAT_WS(' ', sm.first_name, sm.last_name)",
		"programme":       "sm.programme_of_study",
	}
	sort_col = sort_col_map.get(sort_by, "sm.registration_id")

	extra_cond, params = _build_student_filter(exam_plan, search, inst_programmes, inst_batches)

	students = frappe.db.sql(
		f"""
		SELECT
			sm.name                                                             AS student,
			sm.registration_id,
			TRIM(CONCAT_WS(' ', sm.first_name,
				COALESCE(NULLIF(sm.middle_name,''), NULL),
				sm.last_name))                                                  AS student_name,
			sm.student_status,
			sm.programme_of_study AS programme,
			sm.batch_year,
			sm.current_cgpa,
			sm.cumulative_percentage,
			sm.passport_size_photo                                              AS image,
			sm.email,
			COUNT(DISTINCT scm.course)                                          AS course_count
		FROM `tabStudent Course Marks` scm
		INNER JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.exam_plan = %(exam_plan)s
		{extra_cond}
		GROUP BY scm.student
		ORDER BY {sort_col} {sort_dir}
		LIMIT %(lim)s OFFSET %(off)s
		""",
		{**params, "lim": page_length, "off": offset},
		as_dict=True,
	)

	total_row = frappe.db.sql(
		f"""
		SELECT COUNT(DISTINCT scm.student) AS cnt
		FROM `tabStudent Course Marks` scm
		INNER JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.exam_plan = %(exam_plan)s
		{extra_cond}
		""",
		params,
		as_dict=True,
	)
	total = total_row[0]["cnt"] if total_row else 0

	if students:
		_compute_term_gpa(exam_plan, [s["student"] for s in students], students)

	return {"students": students, "total": total}


@frappe.whitelist()
def get_student_courses(exam_plan, student):
	"""Return detailed course marks for a student in the exam plan (View popup)."""
	if not exam_plan or not student:
		return {"courses": [], "student": {}}

	# ── Student profile info ──────────────────────────────────────────────────
	sm = frappe.db.get_value(
		"Student Master",
		student,
		["first_name", "middle_name", "last_name", "registration_id",
		 "email", "programme_of_study", "batch_year", "passport_size_photo", "specialisation"],
		as_dict=True,
	) or {}
	if sm:
		sm["programme"] = sm.pop("programme_of_study", "")

	# ── Course marks with grading schema data ─────────────────────────────────
	rows = frappe.db.sql(
		"""
		SELECT
			scm.course,
			c.course_name,
			c.course_code,
			COALESCE(c.credit_value, 0)                                      AS credit_value,
			c.course_type,
			scm.grade                                                         AS regular_grade,
			COALESCE(scm.total_marks, 0)                                      AS regular_marks,
			scm.moderated_grade,
			CASE WHEN COALESCE(scm.re_exam_grade, '') != ''
			     THEN COALESCE(scm.updated_grade, '') ELSE '' END              AS reexam_grade,
			CASE WHEN COALESCE(scm.re_exam_grade, '') != ''
			     THEN COALESCE(scm.updated_final_marks, 0) ELSE 0 END          AS reexam_marks,
			scm.consider_for_sgpa,
			scm.enrollment_status,
			scm.attendance_status,
			scm.status                                                        AS result_status,
			COALESCE(gs.maximum_marks, 100)                                   AS max_marks,
			COALESCE(gsc.failed, 0)                                           AS is_failed
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabCourse` c ON c.name = scm.course
		LEFT JOIN `tabCourse Schema Assignment` csa
			ON csa.exam_plan = %(exam_plan)s AND csa.course = scm.course
		LEFT JOIN `tabGrading Schema` gs ON gs.name = csa.grade_schema
		LEFT JOIN `tabGrading Schema Component` gsc
			ON gsc.parent = csa.grade_schema
			AND gsc.grade = COALESCE(NULLIF(scm.updated_grade,''), NULLIF(scm.grade,''))
		WHERE scm.exam_plan = %(exam_plan)s
		  AND scm.student   = %(student)s
		ORDER BY c.course_code
		""",
		{"exam_plan": exam_plan, "student": student},
		as_dict=True,
	)

	# ── Per-course moderation marks sum ───────────────────────────────────────
	mod_rows = frappe.db.sql(
		"""
		SELECT scm.course, SUM(sme.moderated_marks) AS moderation_marks
		FROM `tabStudent Course Marks` scm
		INNER JOIN `tabStudent Marks Entry` sme ON sme.parent = scm.name
		WHERE scm.exam_plan = %(exam_plan)s
		  AND scm.student   = %(student)s
		  AND sme.moderated_marks IS NOT NULL
		  AND sme.moderated_marks != 0
		GROUP BY scm.course
		""",
		{"exam_plan": exam_plan, "student": student},
		as_dict=True,
	)
	mod_map = {r["course"]: r["moderation_marks"] for r in mod_rows}

	for row in rows:
		row["moderation_marks"] = mod_map.get(row["course"])

	return {"courses": rows, "student": sm}
