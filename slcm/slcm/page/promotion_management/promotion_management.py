# Copyright (c) 2026, TFSS and contributors
# For license information, please see license.txt

import frappe
from frappe.utils import now_datetime, flt, cint


# ── Filter option helpers ──────────────────────────────────────────────────────

@frappe.whitelist()
def get_programs():
	return frappe.db.get_all("Programme", fields=["name", "program_name"], order_by="name asc")


@frappe.whitelist()
def get_academic_years():
	return frappe.db.get_all(
		"Academic Year", fields=["name", "academic_year_name"],
		order_by="year_start_date desc"
	)


@frappe.whitelist()
def get_policies_for_filters(program, academic_year):
	"""Return matching Active policies for given program + academic_year."""
	if not program or not academic_year:
		return []
	return frappe.db.get_all(
		"Promotion Policy",
		filters={"program": program, "academic_year": academic_year, "status": "Active"},
		fields=["name", "title", "from_year", "to_year",
		        "enable_cgpa_check", "min_cgpa",
		        "enable_backlog_check", "max_backlogs_allowed",
		        "enable_attendance_check", "min_attendance_percent",
		        "enable_course_shortage_check", "max_shortage_courses",
		        "enable_cf_check", "max_cf_fa_shortage",
		        "conditional_promotion_action", "auto_update_student_year"],
		order_by="from_year asc",
	)


# ── Core promotion engine ──────────────────────────────────────────────────────

def _evaluate_student(student_row, policy_dict):
	cgpa_result            = "Not Checked"
	backlog_result         = "Not Checked"
	attendance_result      = "Not Checked"
	shortage_course_result = "Not Checked"
	cf_result              = "Not Checked"
	failures               = 0

	if policy_dict.get("enable_cgpa_check"):
		cgpa = flt(student_row.get("current_cgpa") or 0)
		if cgpa >= flt(policy_dict.get("min_cgpa") or 0):
			cgpa_result = "Pass"
		else:
			cgpa_result = "Fail"
			failures += 1

	if policy_dict.get("enable_backlog_check"):
		backlogs = cint(student_row.get("backlog_count") or 0)
		if backlogs <= cint(policy_dict.get("max_backlogs_allowed") or 0):
			backlog_result = "Pass"
		else:
			backlog_result = "Fail"
			failures += 1

	if policy_dict.get("enable_attendance_check"):
		att = flt(student_row.get("attendance_percent") or 0)
		if att >= flt(policy_dict.get("min_attendance_percent") or 0):
			attendance_result = "Pass"
		else:
			attendance_result = "Fail"
			failures += 1

	# Criterion 3: No more than N courses with attendance shortage
	if policy_dict.get("enable_course_shortage_check"):
		shortage_count = cint(student_row.get("shortage_course_count") or 0)
		if shortage_count <= cint(policy_dict.get("max_shortage_courses") or 2):
			shortage_course_result = "Pass"
		else:
			shortage_course_result = "Fail"
			failures += 1

	# Criterion 4: Carry-forward courses with FA applied but still not exam-eligible
	if policy_dict.get("enable_cf_check"):
		cf_shortage = cint(student_row.get("cf_fa_shortage_count") or 0)
		if cf_shortage <= cint(policy_dict.get("max_cf_fa_shortage") or 0):
			cf_result = "Pass"
		else:
			cf_result = "Fail"
			failures += 1

	total_checks = sum([
		bool(policy_dict.get("enable_cgpa_check")),
		bool(policy_dict.get("enable_backlog_check")),
		bool(policy_dict.get("enable_attendance_check")),
		bool(policy_dict.get("enable_course_shortage_check")),
		bool(policy_dict.get("enable_cf_check")),
	])

	if failures == 0:
		status = "Promoted"
	elif failures == total_checks:
		status = "Not Promoted"
	else:
		cond_action = policy_dict.get("conditional_promotion_action") or "Not Promoted"
		status = "Conditional" if "Conditional" in cond_action else "Not Promoted"

	return {
		"cgpa_result":            cgpa_result,
		"backlog_result":         backlog_result,
		"attendance_result":      attendance_result,
		"shortage_course_result": shortage_course_result,
		"cf_result":              cf_result,
		"promotion_status":       status,
	}


def _get_students_raw(program, academic_year, from_year):
	"""
	Fetch active students for the given program + academic_year + from_year
	along with their CGPA, backlog count, and attendance average.
	"""
	from_year_str = str(from_year)

	students = frappe.db.sql(
		"""
		SELECT
			sm.name          AS student,
			sm.first_name    AS first_name,
			sm.last_name     AS last_name,
			sm.current_cgpa  AS current_cgpa,
			sm.programme     AS programme,
			sm.batch_year    AS batch_year,
			sm.current_year  AS current_year,
			c.cohort_name    AS cohort_name
		FROM `tabStudent Master` sm
		INNER JOIN `tabCohort` c ON c.name = sm.programme
		WHERE
			c.program         = %(program)s
			AND c.academic_year = %(academic_year)s
			AND sm.current_year  = %(from_year)s
			AND sm.student_status = 'Active'
		ORDER BY sm.first_name, sm.last_name
		""",
		{"program": program, "academic_year": academic_year, "from_year": from_year_str},
		as_dict=True,
	)

	if not students:
		return []

	student_names = [s["student"] for s in students]

	# Backlog counts (failed courses in exam plans of this academic year)
	backlog_rows = frappe.db.sql(
		"""
		SELECT scm.student, COUNT(*) AS backlog_count
		FROM `tabStudent Course Marks` scm
		INNER JOIN `tabExam Plan` ep ON ep.name = scm.exam_plan
		INNER JOIN `tabAcademic Term` at2 ON at2.name = ep.term
		WHERE scm.student IN %(students)s
		  AND at2.academic_year = %(academic_year)s
		  AND scm.status = 'Fail'
		GROUP BY scm.student
		""",
		{"students": student_names, "academic_year": academic_year},
		as_dict=True,
	) if student_names else []
	backlog_map = {r["student"]: r["backlog_count"] for r in backlog_rows}

	# Attendance average (overall %)
	att_rows = frappe.db.sql(
		"""
		SELECT student, AVG(attendance_percentage) AS avg_attendance
		FROM `tabAttendance Summary`
		WHERE student IN %(students)s
		  AND academic_year = %(academic_year)s
		GROUP BY student
		""",
		{"students": student_names, "academic_year": academic_year},
		as_dict=True,
	) if student_names else []
	att_map = {r["student"]: flt(r["avg_attendance"]) for r in att_rows}

	# Criterion 3: Count courses where student has attendance shortage
	# (attendance_percentage < minimum_required_percentage for the course)
	shortage_rows = frappe.db.sql(
		"""
		SELECT student, COUNT(*) AS shortage_count
		FROM `tabAttendance Summary`
		WHERE student IN %(students)s
		  AND academic_year = %(academic_year)s
		  AND attendance_percentage < minimum_required_percentage
		GROUP BY student
		""",
		{"students": student_names, "academic_year": academic_year},
		as_dict=True,
	) if student_names else []
	shortage_map = {r["student"]: cint(r["shortage_count"]) for r in shortage_rows}

	# Criterion 4: Carry-forward FA + shortage check
	# Counts courses where FA/MFA condonation was applied (total_fa_mfa_hours > 0)
	# but student is still not exam-eligible (eligible_for_exam = 0).
	# This identifies carry-forward scenarios where attendance was already condoned
	# yet the student still falls short of the required minimum.
	cf_rows = frappe.db.sql(
		"""
		SELECT student, COUNT(*) AS cf_count
		FROM `tabAttendance Summary`
		WHERE student IN %(students)s
		  AND academic_year = %(academic_year)s
		  AND total_fa_mfa_hours > 0
		  AND eligible_for_exam = 0
		GROUP BY student
		""",
		{"students": student_names, "academic_year": academic_year},
		as_dict=True,
	) if student_names else []
	cf_map = {r["student"]: cint(r["cf_count"]) for r in cf_rows}

	for s in students:
		s["backlog_count"]         = backlog_map.get(s["student"], 0)
		s["attendance_percent"]    = att_map.get(s["student"], 0.0)
		s["shortage_course_count"] = shortage_map.get(s["student"], 0)
		s["cf_fa_shortage_count"]  = cf_map.get(s["student"], 0)
		s["student_name"]          = (
			(s.get("first_name") or "") + " " + (s.get("last_name") or "")
		).strip()

	return students


# ── Public APIs ────────────────────────────────────────────────────────────────

@frappe.whitelist()
def fetch_students(program, academic_year, from_year, policy_name=None):
	"""
	Fetch students for filters and optionally evaluate against a policy.
	If no policy_name, all students are returned as 'Promoted' (no criteria active).
	"""
	students = _get_students_raw(program, academic_year, from_year)

	if not students:
		return {"students": [], "counts": {"total": 0, "promoted": 0, "not_promoted": 0, "conditional": 0}}

	policy_dict = {}
	if policy_name:
		p = frappe.get_doc("Promotion Policy", policy_name)
		policy_dict = p.as_dict()

	results = []
	counts  = {"total": 0, "promoted": 0, "not_promoted": 0, "conditional": 0}

	for s in students:
		evaluation = _evaluate_student(s, policy_dict)
		row = {**s, **evaluation}
		results.append(row)
		counts["total"] += 1
		st = evaluation["promotion_status"]
		if st == "Promoted":
			counts["promoted"] += 1
		elif st == "Not Promoted":
			counts["not_promoted"] += 1
		else:
			counts["conditional"] += 1

	return {"students": results, "counts": counts}


@frappe.whitelist()
def confirm_promotion(program, academic_year, from_year, to_year, policy_name=None):
	"""
	Save promotion results:
	- Auto-creates or reuses a Promotion Policy matching the filters.
	- Deletes old Student Promotion records for this policy (idempotent).
	- Creates new Student Promotion records.
	- Updates term_year on Student Master for Promoted students (if policy allows).
	Returns summary counts.
	"""
	frappe.has_permission("Student Promotion", "create", throw=True)

	from_year = cint(from_year)
	to_year   = cint(to_year)

	# Find or create a matching policy
	if policy_name:
		policy = frappe.get_doc("Promotion Policy", policy_name)
	else:
		existing = frappe.db.get_value(
			"Promotion Policy",
			{"program": program, "academic_year": academic_year,
			 "from_year": from_year, "to_year": to_year},
			"name",
		)
		if existing:
			policy = frappe.get_doc("Promotion Policy", existing)
		else:
			policy = frappe.new_doc("Promotion Policy")
			policy.title        = f"{program} | {academic_year} | Yr {from_year}→{to_year}"
			policy.program      = program
			policy.academic_year = academic_year
			policy.from_year    = from_year
			policy.to_year      = to_year
			policy.status       = "Active"
			policy.auto_update_student_year = 1
			policy.insert(ignore_permissions=True)
			frappe.db.commit()

	policy_dict = policy.as_dict()
	students    = _get_students_raw(program, academic_year, from_year)

	if not students:
		frappe.throw("No active students found for the selected filters.")

	# Remove old records for this policy
	old = frappe.db.get_all("Student Promotion", filters={"promotion_policy": policy.name}, pluck="name")
	for name in old:
		frappe.delete_doc("Student Promotion", name, ignore_permissions=True, force=True)

	promoted_count     = 0
	not_promoted_count = 0
	conditional_count  = 0
	now = now_datetime()

	for s in students:
		ev     = _evaluate_student(s, policy_dict)
		status = ev["promotion_status"]

		doc = frappe.new_doc("Student Promotion")
		doc.promotion_policy   = policy.name
		doc.student            = s["student"]
		doc.student_name       = s["student_name"]
		doc.batch_year         = s.get("batch_year") or ""
		doc.programme          = s.get("programme") or ""
		doc.current_year       = str(from_year)
		doc.target_year        = str(to_year)
		doc.current_cgpa            = flt(s.get("current_cgpa") or 0)
		doc.backlog_count           = cint(s.get("backlog_count") or 0)
		doc.attendance_percent      = flt(s.get("attendance_percent") or 0)
		doc.shortage_course_count   = cint(s.get("shortage_course_count") or 0)
		doc.cf_fa_shortage_count    = cint(s.get("cf_fa_shortage_count") or 0)
		doc.cgpa_result             = ev["cgpa_result"]
		doc.backlog_result          = ev["backlog_result"]
		doc.attendance_result       = ev["attendance_result"]
		doc.shortage_course_result  = ev["shortage_course_result"]
		doc.cf_result               = ev["cf_result"]
		doc.promotion_status        = status
		doc.processed_by       = frappe.session.user
		doc.processed_on       = now
		doc.insert(ignore_permissions=True)

		if status == "Promoted":
			promoted_count += 1
		elif status == "Not Promoted":
			not_promoted_count += 1
		else:
			conditional_count += 1

	# Update Student Master year for promoted students
	if policy.auto_update_student_year:
		promoted_students = frappe.db.get_all(
			"Student Promotion",
			filters={"promotion_policy": policy.name, "promotion_status": "Promoted"},
			pluck="student",
		)
		for sname in promoted_students:
			frappe.db.set_value("Student Master", sname, "current_year", str(to_year), update_modified=False)

	frappe.db.commit()

	return {
		"policy_name":  policy.name,
		"total":        len(students),
		"promoted":     promoted_count,
		"not_promoted": not_promoted_count,
		"conditional":  conditional_count,
	}


@frappe.whitelist()
def get_saved_results_by_filters(program, academic_year, from_year, to_year):
	"""Return existing saved Student Promotion records for these filters."""
	policy_name = frappe.db.get_value(
		"Promotion Policy",
		{"program": program, "academic_year": academic_year,
		 "from_year": cint(from_year), "to_year": cint(to_year)},
		"name",
	)
	if not policy_name:
		return {"records": [], "policy_name": None}

	records = frappe.db.get_all(
		"Student Promotion",
		filters={"promotion_policy": policy_name},
		fields=[
			"name", "student", "student_name", "batch_year", "programme",
			"current_year", "target_year", "promotion_status",
			"current_cgpa", "backlog_count", "attendance_percent",
			"shortage_course_count", "cf_fa_shortage_count",
			"cgpa_result", "backlog_result", "attendance_result",
			"shortage_course_result", "cf_result",
			"manual_override", "override_reason", "processed_on",
		],
		order_by="student_name asc",
	)
	return {"records": records, "policy_name": policy_name}


@frappe.whitelist()
def save_override(record_name, new_status, reason):
	frappe.has_permission("Student Promotion", "write", throw=True)
	doc = frappe.get_doc("Student Promotion", record_name)
	doc.promotion_status = new_status
	doc.manual_override  = 1
	doc.override_reason  = reason
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"ok": True}


@frappe.whitelist()
def download_promotion_list(policy_name, list_type):
	"""
	Download Excel for:
	  list_type = 'promoted' | 'not_promoted' | 'conditional' | 'all'
	"""
	import io
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
	except ImportError:
		frappe.throw("openpyxl is not installed. Run: bench pip install openpyxl")

	status_map = {
		"promoted":     ["Promoted", "Override - Promoted"],
		"not_promoted": ["Not Promoted", "Override - Not Promoted"],
		"conditional":  ["Conditional"],
		"all":          ["Promoted", "Not Promoted", "Conditional",
		                 "Override - Promoted", "Override - Not Promoted"],
	}
	statuses = status_map.get(list_type, ["Promoted"])
	policy   = frappe.get_doc("Promotion Policy", policy_name)

	records = frappe.db.get_all(
		"Student Promotion",
		filters=[
			["promotion_policy", "=", policy_name],
			["promotion_status", "in", statuses],
		],
		fields=[
			"student", "student_name", "batch_year", "current_year", "target_year",
			"promotion_status", "current_cgpa", "backlog_count", "attendance_percent",
			"shortage_course_count", "cf_fa_shortage_count",
			"cgpa_result", "backlog_result", "attendance_result",
			"shortage_course_result", "cf_result",
			"manual_override", "override_reason",
		],
		order_by="promotion_status asc, student_name asc",
	)

	label_map = {
		"promoted":     "Promoted List",
		"not_promoted": "Not Promoted List",
		"conditional":  "Conditional List",
		"all":          "All Students",
	}

	wb  = openpyxl.Workbook()
	ws  = wb.active
	ws.title = label_map.get(list_type, "List")

	hdr_fill = PatternFill("solid", fgColor="1E293B")
	hdr_font = Font(bold=True, color="FFFFFF", size=11)
	ctr      = Alignment(horizontal="center", vertical="center")
	bdr      = Border(
		left=Side(style="thin"), right=Side(style="thin"),
		top=Side(style="thin"),  bottom=Side(style="thin"),
	)
	row_colors = {
		"Promoted":                "D1FAE5",
		"Override - Promoted":     "A7F3D0",
		"Not Promoted":            "FEE2E2",
		"Override - Not Promoted": "FECACA",
		"Conditional":             "FEF3C7",
	}

	# Title
	ws.merge_cells("A1:P1")
	t = ws["A1"]
	t.value = (f"{label_map.get(list_type)} — {policy.title}  "
	           f"({policy.program} | {policy.academic_year} | "
	           f"Year {policy.from_year} → Year {policy.to_year})")
	t.font      = Font(bold=True, size=13, color="0F172A")
	t.alignment = ctr
	ws.row_dimensions[1].height = 22

	ws.merge_cells("A2:P2")
	ws["A2"].value = (f"Promoted: {sum(1 for r in records if 'Promoted' in (r.promotion_status or ''))}   |   "
	                  f"Not Promoted: {sum(1 for r in records if 'Not Promoted' in (r.promotion_status or ''))}   |   "
	                  f"Conditional: {sum(1 for r in records if r.promotion_status == 'Conditional')}   |   "
	                  f"Total: {len(records)}")
	ws["A2"].font = Font(italic=True, size=10, color="475569")
	ws.row_dimensions[2].height = 16

	headers   = ["#", "Student ID", "Student Name", "Batch", "Current Year", "Target Year",
	             "CGPA", "Backlogs", "Attendance %", "Shortage Courses", "CF FA+Shortage",
	             "CGPA Check", "Backlog Check", "Attendance Check", "Shortage Check", "CF Check",
	             "Promotion Status", "Override Reason"]
	col_widths = [5, 18, 28, 14, 14, 14, 10, 12, 14, 16, 16, 14, 15, 18, 16, 14, 22, 30]

	for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
		cell         = ws.cell(row=3, column=ci, value=h)
		cell.fill    = hdr_fill
		cell.font    = hdr_font
		cell.alignment = ctr
		cell.border  = bdr
		ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
	ws.row_dimensions[3].height = 18

	for ri, rec in enumerate(records, 1):
		rn   = ri + 3
		rfil = PatternFill("solid", fgColor=row_colors.get(rec.promotion_status, "FFFFFF"))
		vals = [
			ri,
			rec.student,
			rec.student_name,
			rec.batch_year or "",
			rec.current_year or "",
			rec.target_year or "",
			round(flt(rec.current_cgpa), 2),
			cint(rec.backlog_count),
			str(round(flt(rec.attendance_percent), 1)) + "%",
			cint(rec.shortage_course_count),
			cint(rec.cf_fa_shortage_count),
			rec.cgpa_result or "Not Checked",
			rec.backlog_result or "Not Checked",
			rec.attendance_result or "Not Checked",
			rec.shortage_course_result or "Not Checked",
			rec.cf_result or "Not Checked",
			rec.promotion_status,
			rec.override_reason or "",
		]
		for ci, v in enumerate(vals, 1):
			cell         = ws.cell(row=rn, column=ci, value=v)
			cell.fill    = rfil
			cell.border  = bdr
			if ci == 1:
				cell.alignment = ctr

	ws.freeze_panes = "A4"

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.response.filename    = f"{list_type}_{policy_name}.xlsx"
	frappe.response.filecontent = output.read()
	frappe.response.type        = "download"


@frappe.whitelist()
def download_formatted_promotion_list(program, academic_year, university_name=None):
	"""
	Download NLS-style formatted Excel:
	  - One sheet per year-level (all Active policies for program + academic_year)
	  - Each sheet: Promoted section + Re-admitted section
	  - Term-wise failed/shortage courses as columns
	  - Improvement courses column for re-admitted students
	"""
	import io
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
	except ImportError:
		frappe.throw("openpyxl is not installed. Run: bench pip install openpyxl")

	policies = frappe.db.get_all(
		"Promotion Policy",
		filters={"program": program, "academic_year": academic_year, "status": ["in", ["Active", "Draft"]]},
		fields=["name", "title", "from_year", "to_year", "status"],
		order_by="from_year asc",
	)

	# When no policies exist yet, fall back to raw student data grouped by year level.
	# Do NOT filter by academic_year here — cohort.academic_year may differ from the
	# academic_year the user selected (e.g. "2026-27" vs "2025-2026").
	import re as _re
	def _year_num(val, fallback=1):
		"""Extract the first integer from strings like 'Year 1', '2', etc."""
		m = _re.search(r'\d+', str(val or ""))
		return int(m.group()) if m else fallback

	use_raw_fallback = not policies
	if use_raw_fallback:
		raw_year_rows = frappe.db.sql("""
			SELECT DISTINCT sm.current_year
			FROM `tabStudent Master` sm
			INNER JOIN `tabCohort` c ON c.name = sm.programme
			WHERE c.program = %(program)s
			  AND sm.student_status = 'Active'
			  AND sm.current_year IS NOT NULL AND sm.current_year != ''
			ORDER BY sm.current_year
		""", {"program": program}, as_dict=True)
		if not raw_year_rows:
			frappe.throw(f"No active students found for <b>{program}</b>.")
		# Store raw current_year string (e.g. "Year 1") so we can query it back exactly
		policies = [
			frappe._dict(name=None, raw_year=str(r.current_year))
			for r in raw_year_rows
		]

	prog_name = frappe.db.get_value("Programme", program, "program_name") or program
	univ      = (university_name or "").strip()

	terms = frappe.db.get_all(
		"Academic Term",
		filters={"academic_year": academic_year},
		fields=["name", "term_name", "sequence"],
		order_by="sequence asc",
	)
	term_names  = [t.name for t in terms]
	term_labels = [t.term_name for t in terms] if terms else []
	num_terms   = len(term_names)

	def ord_suffix(n):
		n = cint(n)
		if n == 1:   return "1st"
		if n == 2:   return "2nd"
		if n == 3:   return "3rd"
		return f"{n}th"

	thin   = Side(style="thin", color="CBD5E1")
	bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
	ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
	top_l  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

	hdr_fill  = PatternFill("solid", fgColor="1E293B")
	hdr_font  = Font(bold=True, color="FFFFFF", size=10)
	pro_fill  = PatternFill("solid", fgColor="F0FDF4")
	re_fill   = PatternFill("solid", fgColor="FFF7ED")
	sec_pro   = PatternFill("solid", fgColor="DCFCE7")
	sec_re    = PatternFill("solid", fgColor="FEF3C7")

	col_letter = openpyxl.utils.get_column_letter

	def set_row_height(ws, row_num, max_items):
		ws.row_dimensions[row_num].height = max(18, 15 * max(1, max_items))

	def write_section_headers(ws, row_num, include_improvement):
		hdrs = ["Sl No", "Id No", "Student Name", "Email id", "CGPA"] + term_labels
		if include_improvement:
			first_t = term_labels[0]  if term_labels else "Term 1"
			last_t  = term_labels[-1] if term_labels else "Last Term"
			hdrs.append(f"C,C+ (Improvement Course {first_t} to {last_t}, if any)")
		for ci, h in enumerate(hdrs, 1):
			cell = ws.cell(row=row_num, column=ci, value=h)
			cell.fill = hdr_fill; cell.font = hdr_font
			cell.alignment = ctr; cell.border = bdr
		ws.row_dimensions[row_num].height = 30

	def build_course_map(student_ids):
		cm = {sid: {tn: [] for tn in term_names} for sid in student_ids}
		if not student_ids or not term_names:
			return cm
		fail_rows = frappe.db.sql("""
			SELECT scm.student, at2.name AS term_name, c.course_name
			FROM `tabStudent Course Marks` scm
			INNER JOIN `tabExam Plan` ep ON ep.name = scm.exam_plan
			INNER JOIN `tabAcademic Term` at2 ON at2.name = ep.term
			INNER JOIN `tabCourse` c ON c.name = scm.course
			WHERE scm.student IN %(students)s
			  AND at2.academic_year = %(ay)s
			  AND scm.status = 'Fail'
			ORDER BY at2.sequence, c.course_name
		""", {"students": student_ids, "ay": academic_year}, as_dict=True)
		fail_set = set()
		for r in fail_rows:
			tn = r.term_name
			if tn in cm.get(r.student, {}):
				entry = f"{r.course_name} (F)"
				if entry not in cm[r.student][tn]:
					cm[r.student][tn].append(entry)
				fail_set.add((r.student, tn, r.course_name))
		shortage_rows = frappe.db.sql("""
			SELECT att.student, at2.name AS term_name, c.course_name
			FROM `tabAttendance Summary` att
			INNER JOIN `tabCourse` c ON c.name = att.course
			INNER JOIN `tabAcademic Term` at2
			  ON at2.term_name = att.term_name AND at2.academic_year = %(ay)s
			WHERE att.student IN %(students)s
			  AND att.academic_year = %(ay)s
			  AND att.attendance_percentage < att.minimum_required_percentage
			ORDER BY at2.sequence, c.course_name
		""", {"students": student_ids, "ay": academic_year}, as_dict=True)
		for r in shortage_rows:
			tn = r.term_name
			if tn in cm.get(r.student, {}):
				if (r.student, tn, r.course_name) not in fail_set:
					entry = f"{r.course_name} (AS)"
					if entry not in cm[r.student][tn]:
						cm[r.student][tn].append(entry)
		return cm

	wb = openpyxl.Workbook()
	wb.remove(wb.active)

	for policy_idx, policy in enumerate(policies):
		if not use_raw_fallback:
			p = frappe.get_doc("Promotion Policy", policy.name)
			# from_year/to_year may store calendar years (e.g. 2026) instead of
			# study year levels (1, 2, 3). Use sequential index when value > 10.
			if cint(p.from_year) <= 10:
				year_level    = cint(p.from_year)
				year_level_to = cint(p.to_year)
			else:
				year_level    = policy_idx + 1
				year_level_to = policy_idx + 2
		else:
			year_level    = _year_num(policy.raw_year, fallback=policy_idx + 1)
			year_level_to = year_level + 1

		from_ord   = ord_suffix(year_level)
		to_ord     = ord_suffix(year_level_to)
		sheet_name = f"{from_ord} Year"[:31]
		ws = wb.create_sheet(title=sheet_name)

		if use_raw_fallback:
			# No promotion run yet — fetch students directly by program + current_year.
			# Intentionally no academic_year filter: cohort.academic_year may differ
			# from the user-selected academic_year.
			raw_students = frappe.db.sql("""
				SELECT sm.name AS student,
				       sm.first_name, sm.last_name, sm.current_cgpa
				FROM `tabStudent Master` sm
				INNER JOIN `tabCohort` c ON c.name = sm.programme
				WHERE c.program = %(program)s
				  AND sm.current_year = %(yr)s
				  AND sm.student_status = 'Active'
				ORDER BY sm.first_name, sm.last_name
			""", {"program": program, "yr": policy.raw_year}, as_dict=True)
			if not raw_students:
				ws.cell(row=1, column=1, value="No students found for this year level.")
				continue

			student_ids = [s["student"] for s in raw_students]
			email_rows  = frappe.db.sql(
				"SELECT name, email FROM `tabStudent Master` WHERE name IN %(ids)s",
				{"ids": student_ids}, as_dict=True,
			)
			email_map  = {r.name: r.email or "" for r in email_rows}
			course_map = build_course_map(student_ids)

			total_cols = 5 + num_terms
			r = 1
			if univ:
				ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
				c = ws.cell(row=r, column=1, value=univ)
				c.font = Font(bold=True, size=14, color="0F172A"); c.alignment = ctr
				ws.row_dimensions[r].height = 22; r += 1

			ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
			c = ws.cell(row=r, column=1,
			            value=f"Student List — {prog_name} — {from_ord} Year ({academic_year})")
			c.font = Font(bold=True, size=12, color="0F172A"); c.alignment = ctr
			ws.row_dimensions[r].height = 18; r += 1

			ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
			c = ws.cell(row=r, column=1, value="(Promotion not yet run — showing enrolled students)")
			c.font = Font(bold=False, size=10, color="92400E"); c.alignment = ctr
			ws.row_dimensions[r].height = 14; r += 1

			ws.row_dimensions[r].height = 8; r += 1

			hdrs = ["Sl No", "Id No", "Student Name", "Email id", "CGPA"] + term_labels
			for ci, h in enumerate(hdrs, 1):
				cell = ws.cell(row=r, column=ci, value=h)
				cell.fill = hdr_fill; cell.font = hdr_font
				cell.alignment = ctr; cell.border = bdr
			ws.row_dimensions[r].height = 30; r += 1

			for si, s in enumerate(raw_students, 1):
				sname = f"{s.get('first_name', '')} {s.get('last_name', '')}".strip()
				vals  = [si, s["student"], sname,
				         email_map.get(s["student"], ""),
				         round(flt(s.get("current_cgpa") or 0), 2)]
				for tn in term_names:
					courses = course_map.get(s["student"], {}).get(tn, [])
					vals.append("\n".join(courses) if courses else "")
				max_lines = max(1, max(
					(len(course_map.get(s["student"], {}).get(tn, [])) for tn in term_names),
					default=1
				))
				for ci, v in enumerate(vals, 1):
					cell = ws.cell(row=r, column=ci, value=v)
					cell.fill = pro_fill; cell.border = bdr
					cell.alignment = ctr if ci == 1 else top_l
				set_row_height(ws, r, max_lines); r += 1

			col_widths = [6, 18, 28, 32, 10] + [22] * num_terms
			for ci, w in enumerate(col_widths[:total_cols], 1):
				ws.column_dimensions[col_letter(ci)].width = w
			ws.freeze_panes = "A6" if not univ else "A7"
			continue  # skip policy-based rendering below

		# ── Policy-based rendering (promotion already run) ────────────────────
		records = frappe.db.get_all(
			"Student Promotion",
			filters=[
				["promotion_policy", "=", policy.name],
				["promotion_status", "in", [
					"Promoted", "Override - Promoted",
					"Not Promoted", "Override - Not Promoted",
					"Conditional",
				]],
			],
			fields=["student", "student_name", "promotion_status", "current_cgpa"],
			order_by="student_name asc",
		)

		student_ids = [r.student for r in records]
		if not student_ids:
			ws.cell(row=1, column=1, value="No confirmed records found.")
			continue

		email_rows = frappe.db.sql(
			"SELECT name, email FROM `tabStudent Master` WHERE name IN %(ids)s",
			{"ids": student_ids}, as_dict=True,
		)
		email_map  = {r.name: r.email or "" for r in email_rows}
		course_map = build_course_map(student_ids)

		promoted_recs = [r for r in records if "Promoted" in (r.promotion_status or "")]
		readmit_recs  = [r for r in records if r.promotion_status not in
		                 ("Promoted", "Override - Promoted")]

		total_cols = 5 + num_terms + 1  # Sl,ID,Name,Email,CGPA + terms + improvement

		r = 1
		if univ:
			ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
			c = ws.cell(row=r, column=1, value=univ)
			c.font = Font(bold=True, size=14, color="0F172A")
			c.alignment = ctr
			ws.row_dimensions[r].height = 22
			r += 1

		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
		c = ws.cell(row=r, column=1,
		            value=f"Promotion List of {prog_name} — {from_ord} Year ({academic_year})")
		c.font = Font(bold=True, size=12, color="0F172A")
		c.alignment = ctr
		ws.row_dimensions[r].height = 18
		r += 1

		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
		c = ws.cell(row=r, column=1,
		            value=f"(Promoted to {to_ord} Year AY {academic_year})")
		c.font = Font(bold=False, size=11, color="374151")
		c.alignment = ctr
		ws.row_dimensions[r].height = 16
		r += 1

		ws.row_dimensions[r].height = 8
		r += 1

		# ── Promoted section ──────────────────────────────────────────────────
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
		c = ws.cell(row=r, column=1,
		            value=f"Promoted to {to_ord} Year AY {academic_year}  "
		                  f"({len(promoted_recs)} students)")
		c.font = Font(bold=True, size=11, color="166534")
		c.fill = sec_pro
		c.alignment = ctr
		ws.row_dimensions[r].height = 18
		r += 1

		write_section_headers(ws, r, include_improvement=False)
		r += 1

		if promoted_recs:
			for si, rec in enumerate(promoted_recs, 1):
				vals = [si, rec.student, rec.student_name,
				        email_map.get(rec.student, ""),
				        round(flt(rec.current_cgpa), 2)]
				for tn in term_names:
					courses = course_map.get(rec.student, {}).get(tn, [])
					vals.append("\n".join(courses) if courses else "")
				max_lines = max(1, max(
					(len(course_map.get(rec.student, {}).get(tn, [])) for tn in term_names),
					default=1
				))
				for ci, v in enumerate(vals, 1):
					cell = ws.cell(row=r, column=ci, value=v)
					cell.fill = pro_fill; cell.border = bdr
					cell.alignment = ctr if ci == 1 else top_l
				set_row_height(ws, r, max_lines)
				r += 1
		else:
			ws.cell(row=r, column=1, value="— None —").alignment = ctr
			r += 1

		ws.row_dimensions[r].height = 10
		r += 1

		# ── Re-admitted section ───────────────────────────────────────────────
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
		c = ws.cell(row=r, column=1,
		            value=f"Re-admitted to {from_ord} year AY {academic_year}  "
		                  f"({len(readmit_recs)} students)")
		c.font = Font(bold=True, size=11, color="92400E")
		c.fill = sec_re
		c.alignment = ctr
		ws.row_dimensions[r].height = 18
		r += 1

		write_section_headers(ws, r, include_improvement=True)
		r += 1

		if readmit_recs:
			for si, rec in enumerate(readmit_recs, 1):
				all_issues = []
				seen_imp = set()
				for tn in term_names:
					for item in course_map.get(rec.student, {}).get(tn, []):
						if item not in seen_imp:
							seen_imp.add(item)
							all_issues.append(item)

				vals = [si, rec.student, rec.student_name,
				        email_map.get(rec.student, ""),
				        round(flt(rec.current_cgpa), 2)]
				for tn in term_names:
					courses = course_map.get(rec.student, {}).get(tn, [])
					vals.append("\n".join(courses) if courses else "")
				vals.append("\n".join(all_issues))

				max_lines = max(1,
					max((len(course_map.get(rec.student, {}).get(tn, [])) for tn in term_names),
					    default=1),
					len(all_issues),
				)
				for ci, v in enumerate(vals, 1):
					cell = ws.cell(row=r, column=ci, value=v)
					cell.fill = re_fill; cell.border = bdr
					cell.alignment = ctr if ci == 1 else top_l
				set_row_height(ws, r, max_lines)
				r += 1
		else:
			ws.cell(row=r, column=1, value="— None —").alignment = ctr
			r += 1

		col_widths = [6, 18, 28, 32, 10] + [22] * num_terms + [40]
		for ci, w in enumerate(col_widths[:total_cols], 1):
			ws.column_dimensions[col_letter(ci)].width = w
		ws.freeze_panes = "A7" if univ else "A6"

	if not wb.sheetnames:
		frappe.throw("No data to export.")

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	safe_prog = program.replace(" ", "_")
	safe_ay   = academic_year.replace(" ", "").replace("-", "_")
	frappe.response.filename    = f"Promotion_List_{safe_prog}_{safe_ay}.xlsx"
	frappe.response.filecontent = output.read()
	frappe.response.type        = "download"
