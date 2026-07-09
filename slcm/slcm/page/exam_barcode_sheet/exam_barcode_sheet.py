# Copyright (c) 2026, TFSS and contributors
# For license information, please see license.txt

import frappe


@frappe.whitelist()
def get_exam_plans():
	"""Return list of active exam plans."""
	return frappe.db.get_all(
		"Exam Plan",
		filters={"status": "Active"},
		fields=["name", "exam_name", "term"],
		order_by="exam_name asc",
	)


@frappe.whitelist()
def get_exam_courses(exam_plan):
	"""Return courses scheduled under the given exam plan, with dates and barcode counts."""
	schedules = frappe.db.get_all(
		"Exam Course Schedule",
		filters={"parent": exam_plan, "parenttype": "Exam Plan"},
		fields=["course", "exam_date", "start_time", "end_time", "venue", "hall"],
		order_by="exam_date asc, course asc",
	)

	if not schedules:
		return []

	# Fetch course names and codes
	course_info = {
		r.name: {"course_name": r.course_name, "course_code": r.course_code or r.name}
		for r in frappe.db.get_all(
			"Course",
			filters={"name": ["in", [s["course"] for s in schedules]]},
			fields=["name", "course_name", "course_code"],
		)
	}

	# Count existing barcodes per course using raw SQL (group_by with COUNT)
	barcode_counts = {}
	rows = frappe.db.sql(
		"""
		SELECT course, COUNT(name) AS cnt
		FROM `tabExam Barcode`
		WHERE exam_plan = %s
		GROUP BY course
		""",
		(exam_plan,),
		as_dict=True,
	)
	for r in rows:
		barcode_counts[r.course] = r.cnt

	for s in schedules:
		info = course_info.get(s["course"], {})
		s["course_name"] = info.get("course_name", s["course"])
		s["course_code"] = info.get("course_code", s["course"])
		s["barcode_count"] = barcode_counts.get(s["course"], 0)

	return schedules


@frappe.whitelist()
def generate_barcodes(exam_plan, courses=None):
	"""
	Generate 6-digit unique barcodes for all students enrolled in the given
	courses under the exam plan. Existing barcodes are preserved; only missing
	students get new ones.

	courses: JSON list of course names (optional, defaults to all scheduled).
	"""
	import json
	import random
	import datetime

	if isinstance(courses, str):
		courses = json.loads(courses)

	if not courses:
		schedules = frappe.db.get_all(
			"Exam Course Schedule",
			filters={"parent": exam_plan, "parenttype": "Exam Plan"},
			fields=["course", "exam_date"],
		)
		courses = list({s.course for s in schedules})

	# Build exam_date lookup per course
	date_lookup = {}
	for s in frappe.db.get_all(
		"Exam Course Schedule",
		filters={"parent": exam_plan, "parenttype": "Exam Plan", "course": ["in", courses]},
		fields=["course", "exam_date"],
	):
		date_lookup[s.course] = s.exam_date

	# Collect all barcodes already issued under this plan (for uniqueness)
	used_barcodes = set(
		r.barcode
		for r in frappe.db.get_all("Exam Barcode", filters={"exam_plan": exam_plan}, fields=["barcode"])
	)

	def _new_barcode():
		for _ in range(10000):
			code = str(random.randint(100000, 999999))
			if code not in used_barcodes:
				used_barcodes.add(code)
				return code
		frappe.throw("Could not generate a unique 6-digit barcode. Barcode pool may be exhausted.")

	now = frappe.utils.now_datetime()
	created = 0

	for course in courses:
		students = _get_students_for_course(course, exam_plan)
		if not students:
			continue

		# Find which students already have barcodes
		existing = {
			r.student
			for r in frappe.db.get_all(
				"Exam Barcode",
				filters={"exam_plan": exam_plan, "course": course},
				fields=["student"],
			)
		}

		for st in students:
			if st["student"] in existing:
				continue
			doc = frappe.get_doc({
				"doctype": "Exam Barcode",
				"exam_plan": exam_plan,
				"course": course,
				"exam_date": date_lookup.get(course),
				"student": st["student"],
				"student_name": st["student_name"],
				"registration_id": st["registration_id"],
				"section": st["section"],
				"barcode": _new_barcode(),
				"generated_on": now,
			})
			doc.insert(ignore_permissions=True)
			created += 1

	frappe.db.commit()
	return {"created": created, "message": f"{created} new barcode(s) generated."}


@frappe.whitelist()
def get_barcodes(exam_plan, course=None):
	"""Return all barcodes for the exam plan (optionally filtered by course)."""
	filters = {"exam_plan": exam_plan}
	if course:
		filters["course"] = course

	return frappe.db.get_all(
		"Exam Barcode",
		filters=filters,
		fields=["course", "exam_date", "student", "student_name", "registration_id", "section", "barcode"],
		order_by="course asc, registration_id asc",
	)


@frappe.whitelist()
def get_course_students(exam_plan, course):
	"""
	Return students for a course with their barcode (if generated).
	Source: Student Course Marks (same as Examination Result page).
	"""
	students = _get_students_for_course(course, exam_plan)

	# Existing barcodes
	existing_barcodes = {
		r.student: r.barcode
		for r in frappe.db.get_all(
			"Exam Barcode",
			filters={"exam_plan": exam_plan, "course": course},
			fields=["student", "barcode"],
		)
	}

	result = []
	for st in students:
		result.append({
			"student": st["student"],
			"student_name": st["student_name"],
			"registration_id": st["registration_id"],
			"section": st["section"],
			"barcode": existing_barcodes.get(st["student"], ""),
			"has_barcode": st["student"] in existing_barcodes,
		})

	result.sort(key=lambda x: (x["registration_id"] or "zzz", x["student_name"]))
	return result


@frappe.whitelist()
def export_attendance_excel(exam_plan, course=None, mode="by_date",
							from_date=None, to_date=None, selected_courses=None):
	"""
	Generate exam attendance Excel.

	course (str):            single-course export (one sheet).
	mode:
	  'by_date'   — one sheet per exam date (default).
	  'by_course' — one sheet per course.
	from_date / to_date:     filter date range (ISO strings); only used in by_date mode.
	selected_courses (JSON): list of course names; filter for by_course download.
	"""
	import io, base64, json as _json
	import openpyxl
	from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
	from openpyxl.utils import get_column_letter
	from collections import defaultdict

	# Normalise optional args
	course           = course or None
	from_date        = from_date or None
	to_date          = to_date   or None
	if isinstance(selected_courses, str):
		selected_courses = _json.loads(selected_courses) if selected_courses.strip() else None

	plan_doc = frappe.get_doc("Exam Plan", exam_plan)

	# ── Fetch schedules ───────────────────────────────────────────────────────
	sch_filters = {"parent": exam_plan, "parenttype": "Exam Plan"}
	if course:
		sch_filters["course"] = course

	schedules = frappe.db.get_all(
		"Exam Course Schedule",
		filters=sch_filters,
		fields=["course", "exam_date", "start_time", "end_time", "venue", "hall"],
		order_by="exam_date asc, course asc",
	)

	# ── Apply date-range filter (by_date mode) ────────────────────────────────
	if not course and (from_date or to_date):
		fd = frappe.utils.getdate(from_date) if from_date else None
		td = frappe.utils.getdate(to_date)   if to_date   else None
		filtered = []
		for s in schedules:
			if not s.exam_date:
				continue  # exclude undated courses when a range is given
			d = frappe.utils.getdate(str(s.exam_date))
			if fd and d < fd:
				continue
			if td and d > td:
				continue
			filtered.append(s)
		schedules = filtered
		if not schedules:
			frappe.throw("No courses found within the selected date range.")

	# ── Apply course-list filter (by_course mode) ─────────────────────────────
	if not course and mode == "by_course" and selected_courses:
		schedules = [s for s in schedules if s.course in selected_courses]
		if not schedules:
			frappe.throw("None of the selected courses have schedules in this Exam Plan.")

	if not schedules:
		frappe.throw("No courses scheduled in this Exam Plan.")

	# ── Course name / code maps ───────────────────────────────────────────────
	course_name_map = {}   # plain course name (used in Excel row 3 heading)
	course_code_map = {}   # course code (used for sheet tab names)
	for r in frappe.db.get_all(
		"Course",
		filters={"name": ["in", list({s["course"] for s in schedules})]},
		fields=["name", "course_name", "course_code"],
	):
		course_name_map[r.name] = r.course_name or r.name
		course_code_map[r.name] = r.course_code or r.name

	# ── Barcode map: course -> list of student rows ───────────────────────────
	bc_filters = {"exam_plan": exam_plan}
	if course:
		bc_filters["course"] = course
	barcodes = frappe.db.get_all(
		"Exam Barcode",
		filters=bc_filters,
		fields=["course", "exam_date", "student", "student_name", "registration_id", "section", "barcode"],
		order_by="registration_id asc",
	)
	barcode_map = {}
	for b in barcodes:
		barcode_map.setdefault(b.course, []).append(b)

	# ── Institution name — from Institution Settings (configured by admin) ────
	institution = "Institution"
	try:
		inst = frappe.get_single("Institution Settings")
		if inst.institution_name:
			institution = inst.institution_name
	except Exception:
		pass
	if institution == "Institution":
		# Fallback: default company name
		try:
			co = frappe.defaults.get_global_default("company")
			if co:
				institution = frappe.db.get_value("Company", co, "company_name") or co
		except Exception:
			pass

	# ── Term / Academic Year label ────────────────────────────────────────────
	term_label = ""
	try:
		if plan_doc.term:
			term_doc = frappe.get_doc("Academic Term", plan_doc.term)
			term_label = term_doc.term_name or plan_doc.term
			if hasattr(term_doc, "academic_year") and term_doc.academic_year:
				yd = frappe.db.get_value("Academic Year", term_doc.academic_year, "year_name") or term_doc.academic_year
				term_label = f"{term_label}, AY {yd}"
	except Exception:
		term_label = plan_doc.term or ""

	# ── Shared style objects ──────────────────────────────────────────────────
	thin   = Side(style="thin")
	medium = Side(style="medium")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	# Thicker outer border for the header block (rows 1-3)
	hdr_border = Border(left=medium, right=medium, top=medium, bottom=medium)
	header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
	# Column widths: Sl.No | Name | ID no. | Sec | Barcode | Ans Booklet | Signature | Add. Booklet
	COL_WIDTHS = [6, 30, 16, 8, 12, 18, 18, 18]

	def _font(size=11, bold=False, italic=False, color=None):
		kw = dict(name="Bookman Old Style", size=size, bold=bold, italic=italic)
		if color:
			kw["color"] = color
		return Font(**kw)

	def _center(cell, wrap=True):
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)

	def _left(cell):
		cell.alignment = Alignment(horizontal="left", vertical="center")

	def _set_col_widths(ws):
		for i, w in enumerate(COL_WIDTHS, 1):
			ws.column_dimensions[get_column_letter(i)].width = w

	def _write_course_block(ws, start_row, sched, course_name, students):
		"""Write one course attendance block. Returns the next free row."""
		date_str = "DD/MM/YYYY"
		if sched.exam_date:
			try:
				date_str = frappe.utils.formatdate(str(sched.exam_date), "dd/MM/yyyy")
			except Exception:
				date_str = str(sched.exam_date)

		row = start_row

		# ── Row 1: Institution name ───────────────────────────────────────────
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
		c = ws.cell(row=row, column=1, value=institution)
		c.font = _font(14, bold=True)
		_center(c)
		ws.row_dimensions[row].height = 26
		row += 1

		# ── Row 2: Term / Academic Year ───────────────────────────────────────
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
		c = ws.cell(row=row, column=1, value=term_label)
		c.font = _font(12, bold=True)
		_center(c)
		ws.row_dimensions[row].height = 20
		row += 1

		# ── Row 3: Course name + Exam type + Date ─────────────────────────────
		exam_line = f"{course_name} End term Examination  ({date_str})"
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
		c = ws.cell(row=row, column=1, value=exam_line)
		c.font = _font(11, bold=True)
		_center(c)
		ws.row_dimensions[row].height = 18
		row += 1

		# ── Row 4: Column headers ─────────────────────────────────────────────
		col_headers = [
			"Sl.No", "Name", "ID no.", "Sec", "Barcode",
			"Answer Booklet\nnumbers", "Student\nSignature",
			"No. of Additional\nBooklet Used",
		]
		for col_idx, header in enumerate(col_headers, 1):
			c = ws.cell(row=row, column=col_idx, value=header)
			c.font = _font(11, bold=True)
			c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			c.border = border
			c.fill = header_fill
		ws.row_dimensions[row].height = 38
		row += 1

		# ── Student rows ──────────────────────────────────────────────────────
		sorted_students = sorted(students, key=lambda x: (x.registration_id or ""))
		if sorted_students:
			for sl, st in enumerate(sorted_students, 1):
				row_data = [
					sl,
					st.student_name or "",
					st.registration_id or "",
					st.section or "",
					st.barcode or "",
					"", "", "",  # Answer Booklet / Signature / Additional Booklet
				]
				for col_idx, val in enumerate(row_data, 1):
					c = ws.cell(row=row, column=col_idx, value=val)
					c.font = _font(11)
					c.border = border
					if col_idx in (1, 3, 4, 5, 6, 8):
						_center(c)
					else:
						_left(c)
				# Tall rows give space for handwritten signatures
				ws.row_dimensions[row].height = 40
				row += 1
		else:
			ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
			c = ws.cell(row=row, column=1, value="No barcodes generated for this course yet.")
			c.font = _font(11, italic=True, color="FF0000")
			_center(c)
			row += 1

		return row + 2  # blank gap between course blocks

	# ── Build workbook ────────────────────────────────────────────────────────
	wb = openpyxl.Workbook()
	wb.remove(wb.active)

	if course:
		# ── Single-course mode ────────────────────────────────────────────────
		sched = schedules[0]
		course_name = course_name_map.get(sched.course, sched.course)
		students = barcode_map.get(sched.course, [])
		sheet_name = course_code_map.get(sched.course, sched.course)[:31]
		ws = wb.create_sheet(title=sheet_name)
		_write_course_block(ws, 1, sched, course_name, students)
		_set_col_widths(ws)
		safe = sched.course.replace(" ", "_").replace("/", "-")[:30]
		filename = f"Attendance_{safe}.xlsx"

	elif mode == "by_course":
		# ── One sheet per course ──────────────────────────────────────────────
		for sched in schedules:
			course_name = course_name_map.get(sched.course, sched.course)
			students = barcode_map.get(sched.course, [])
			sheet_name = course_code_map.get(sched.course, sched.course)[:31]
			ws = wb.create_sheet(title=sheet_name)
			_write_course_block(ws, 1, sched, course_name, students)
			_set_col_widths(ws)
		filename = f"Exam_Attendance_ByCourse_{exam_plan.replace(' ', '_')}.xlsx"

	else:
		# ── By date (default) — one sheet per exam date ───────────────────────
		date_groups = defaultdict(list)
		for s in schedules:
			date_groups[s.exam_date].append(s)

		for exam_date in sorted(date_groups.keys(), key=lambda d: (str(d) if d else "")):
			date_str = ""
			if exam_date:
				try:
					date_str = frappe.utils.formatdate(str(exam_date), "dd/MM/yyyy")
				except Exception:
					date_str = str(exam_date)
			sheet_name = (date_str.replace("/", "-") if date_str else "No Date")[:31]
			ws = wb.create_sheet(title=sheet_name)
			current_row = 1
			for sched in date_groups[exam_date]:
				course_name = course_name_map.get(sched.course, sched.course)
				students = barcode_map.get(sched.course, [])
				current_row = _write_course_block(ws, current_row, sched, course_name, students)
			_set_col_widths(ws)
		filename = f"Exam_Attendance_{exam_plan.replace(' ', '_')}.xlsx"

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)
	file_b64 = base64.b64encode(buf.read()).decode()
	return {"file_content": file_b64, "filename": filename}


@frappe.whitelist()
def get_student_details(student, exam_plan=None, course=None):
	"""Return full Student Master details plus barcode for the exam context."""
	sm = frappe.get_doc("Student Master", student)

	def _get(*field_names, default=""):
		for fn in field_names:
			val = getattr(sm, fn, None)
			if val:
				return val
		return default

	barcode = ""
	if exam_plan and course:
		barcode = frappe.db.get_value(
			"Exam Barcode",
			{"student": student, "exam_plan": exam_plan, "course": course},
			"barcode",
		) or ""

	dob_str = ""
	dob = _get("date_of_birth")
	if dob:
		try:
			dob_str = frappe.utils.formatdate(str(dob))
		except Exception:
			dob_str = str(dob)

	full_name = f"{_get('first_name')} {_get('last_name')}".strip() or sm.name

	return {
		"student": sm.name,
		"student_name": full_name,
		"first_name": _get("first_name"),
		"middle_name": _get("middle_name"),
		"last_name": _get("last_name"),
		"registration_id": _get("registration_id") or sm.name,
		"gender": _get("gender"),
		"date_of_birth": dob_str,
		"blood_group": _get("blood_group"),
		"email": _get("student_email_id", "email_id", "email"),
		"phone": _get("student_mobile_number", "mobile_no", "phone", "contact_number"),
		"program": _get("program"),
		"academic_year": _get("academic_year"),
		"joining_date": str(_get("joining_date")) if _get("joining_date") else "",
		"student_status": _get("student_status", "status"),
		"image": _get("student_image", "image"),
		"barcode": barcode,
		"course": course or "",
		"exam_plan": exam_plan or "",
	}


# ── Internal helpers ────────────────────────────────────────────────────────────

def _get_students_for_course(course, exam_plan):
	"""
	Return list of dicts {student, student_name, registration_id, section}.
	Primary source: Student Course Marks — the same table used by Examination Result.
	Falls back to Student Enrollment (Student Enrollment Course + Course Offering, same
	join as exam_plan_api.get_courses_for_plan) when no Student Course Marks rows exist
	yet for this exam plan/course — i.e. before anyone has run the marks sync.
	Students are sorted by registration_id ascending.
	"""
	rows = frappe.db.sql(
		"""
		SELECT
			scm.student,
			CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
			COALESCE(NULLIF(sm.registration_id, ''), sm.name) AS registration_id,
			COALESCE(cc.section, '') AS section
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		LEFT JOIN (
			SELECT cs.student, cc2.section, cc2.course
			FROM `tabClass Student` cs
			INNER JOIN `tabClass Configuration` cc2 ON cc2.name = cs.parent
		) cc ON cc.student = scm.student AND cc.course = scm.course
		WHERE scm.exam_plan = %(exam_plan)s
		  AND scm.course = %(course)s
		ORDER BY COALESCE(NULLIF(sm.registration_id, ''), sm.name) ASC
		""",
		{"exam_plan": exam_plan, "course": course},
		as_dict=True,
	)

	if not rows:
		term_name = frappe.db.get_value("Exam Plan", exam_plan, "term")
		rows = frappe.db.sql(
			"""
			SELECT
				se.student AS student,
				CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
				COALESCE(NULLIF(sm.registration_id, ''), sm.name) AS registration_id,
				COALESCE(cc.section, '') AS section
			FROM `tabStudent Enrollment` se
			INNER JOIN `tabStudent Enrollment Course` sec ON sec.parent = se.name
			INNER JOIN `tabCourse Offering` co ON co.name = sec.course_offering
			LEFT JOIN `tabStudent Master` sm ON sm.name = se.student
			LEFT JOIN (
				SELECT cs.student, cc2.section, cc2.course
				FROM `tabClass Student` cs
				INNER JOIN `tabClass Configuration` cc2 ON cc2.name = cs.parent
			) cc ON cc.student = se.student AND cc.course = co.course_title
			WHERE co.course_title = %(course)s
			  AND sec.status = 'Enrolled'
			  AND se.status = 'Enrolled'
			  AND (%(term_name)s IS NULL OR se.term_name = %(term_name)s)
			GROUP BY se.student
			ORDER BY COALESCE(NULLIF(sm.registration_id, ''), sm.name) ASC
			""",
			{"course": course, "term_name": term_name},
			as_dict=True,
		)

	# Deduplicate — a student could appear in multiple class configurations
	seen = set()
	result = []
	for r in rows:
		if r.student in seen:
			continue
		seen.add(r.student)
		result.append({
			"student": r.student,
			"student_name": r.student_name or r.student,
			"registration_id": r.registration_id or "",
			"section": r.section or "",
		})

	return result
