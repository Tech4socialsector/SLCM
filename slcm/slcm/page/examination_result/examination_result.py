# Copyright (c) 2026, TFSS and contributors
# For license information, please see license.txt

import frappe


# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_or_create_access(exam_plan, course):
	"""Return the Access Result Settings doc for the given exam_plan+course, creating if absent."""
	existing = frappe.db.get_value(
		"Access Result Settings",
		{"exam_plan": exam_plan, "course": course},
		"name",
	)
	if existing:
		return frappe.get_doc("Access Result Settings", existing)
	doc = frappe.new_doc("Access Result Settings")
	doc.exam_plan    = exam_plan
	doc.course       = course
	doc.status       = "UNLOCKED"
	doc.view_access  = 1
	doc.edit_access  = 1
	doc.insert(ignore_permissions=True)
	return doc


def _access_map(exam_plan):
	"""Return a dict of course → Access Result Settings row (without child tables)."""
	rows = frappe.db.sql(
		"""
		SELECT
			cra.name, cra.course,
			cra.view_access, cra.view_deadline, cra.edit_access, cra.edit_deadline,
			cra.auto_generate_grade_access, cra.edit_grade_access,
			cra.relative_grading_access, cra.mask_student_info,
			cra.generate_grade_report, cra.moderation_policy_access,
			cra.status
		FROM `tabAccess Result Settings` cra
		WHERE cra.exam_plan = %(exam_plan)s
		""",
		{"exam_plan": exam_plan},
		as_dict=True,
	)
	return {r["course"]: r for r in rows}


def _evaluator_map(exam_plan):
	"""Return a dict of course → list of evaluator dicts for the given exam plan.

	evaluator_name is a Link to Faculty; we fetch the faculty's full name and
	email from the Faculty doctype for display.
	"""
	rows = frappe.db.sql(
		"""
		SELECT
			cra.course,
			rce.evaluator_type,
			rce.evaluator_name,
			rce.evaluator_email,
			f.faculty_name
		FROM `tabResult Course Evaluator` rce
		INNER JOIN `tabAccess Result Settings` cra ON rce.parent = cra.name
		LEFT JOIN `tabFaculty` f ON f.name = rce.evaluator_name
		WHERE cra.exam_plan = %(exam_plan)s
		""",
		{"exam_plan": exam_plan},
		as_dict=True,
	)
	result = {}
	for r in rows:
		result.setdefault(r["course"], []).append({
			"evaluator_type":  r["evaluator_type"],
			"evaluator_name":  r["evaluator_name"],
			"evaluator_display": r.get("faculty_name") or r["evaluator_name"] or "",
			"evaluator_email": r["evaluator_email"],
		})
	return result


def _visibility_map(exam_plan):
	"""Return a dict of course → list of exam type dicts for the given exam plan.

	exam_type is a Link to Exam Assessment Type; we fetch type_name for display.
	"""
	rows = frappe.db.sql(
		"""
		SELECT cra.course, rev.exam_type, eat.type_name
		FROM `tabResult Exam Visibility` rev
		INNER JOIN `tabAccess Result Settings` cra ON rev.parent = cra.name
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = rev.exam_type
		WHERE cra.exam_plan = %(exam_plan)s
		""",
		{"exam_plan": exam_plan},
		as_dict=True,
	)
	result = {}
	for r in rows:
		result.setdefault(r["course"], []).append({
			"name":      r["exam_type"],
			"type_name": r.get("type_name") or r["exam_type"] or "",
		})
	return result


# ── Public API ─────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_exam_plans(search=None):
	"""Return all exam plans for the term selector."""
	filters = {}
	if search:
		filters["exam_name"] = ["like", f"%{search}%"]
	return frappe.get_all(
		"Exam Plan",
		filters=filters,
		fields=["name", "exam_name", "term", "status"],
		order_by="creation desc",
	)


@frappe.whitelist()
def get_courses_for_result(exam_plan, search=""):
	"""
	Return all courses with their result-access settings for the given exam plan.
	Mirrors the Access Results table in the UI.
	"""
	if not exam_plan:
		frappe.throw("Exam Plan is required.")

	page_length = 500

	if search:
		courses = frappe.db.sql(
			"""
			SELECT name, course_name, course_code, department_name
			FROM `tabCourse`
			WHERE course_name LIKE %(s)s OR course_code LIKE %(s)s
			ORDER BY course_name ASC
			LIMIT %(limit)s
			""",
			{"s": f"%{search}%", "limit": page_length},
			as_dict=True,
		)
	else:
		courses = frappe.get_all(
			"Course",
			fields=["name", "course_name", "course_code", "department_name"],
			order_by="course_name asc",
			page_length=page_length,
		)

	amap = _access_map(exam_plan)
	emap = _evaluator_map(exam_plan)
	vmap = _visibility_map(exam_plan)

	for c in courses:
		acc = amap.get(c["name"], {})
		c["evaluators"]                 = emap.get(c["name"], [])
		c["view_access"]                = int(acc.get("view_access", 1))
		c["view_deadline"]              = str(acc.get("view_deadline") or "")
		c["edit_access"]                = int(acc.get("edit_access", 1))
		c["edit_deadline"]              = str(acc.get("edit_deadline") or "")
		c["auto_generate_grade_access"] = int(acc.get("auto_generate_grade_access", 0))
		c["edit_grade_access"]          = int(acc.get("edit_grade_access", 0))
		c["relative_grading_access"]    = int(acc.get("relative_grading_access", 0))
		c["mask_student_info"]          = int(acc.get("mask_student_info", 0))
		c["generate_grade_report"]      = int(acc.get("generate_grade_report", 0))
		c["moderation_policy_access"]   = int(acc.get("moderation_policy_access", 0))
		c["status"]                     = acc.get("status", "UNLOCKED")
		# list of {name, type_name} dicts — name is the Exam Assessment Type link
		c["visible_exams"]              = vmap.get(c["name"], [])

	return courses


@frappe.whitelist()
def save_access_settings(exam_plan, courses, settings):
	"""
	Bulk-update access settings for the selected courses.

	courses  – JSON list of course names
	settings – JSON dict with fields:
	           view_access, view_deadline, edit_access, edit_deadline,
	           auto_generate_grade_access, edit_grade_access,
	           relative_grading_access, mask_student_info,
	           generate_grade_report, moderation_policy_access
	"""
	import json

	if isinstance(courses, str):
		courses = json.loads(courses)
	if isinstance(settings, str):
		settings = json.loads(settings)

	if not courses:
		frappe.throw("No courses selected.")

	_ALLOWED = {
		"view_access", "view_deadline", "edit_access", "edit_deadline",
		"auto_generate_grade_access", "edit_grade_access",
		"relative_grading_access", "mask_student_info",
		"generate_grade_report", "moderation_policy_access",
	}

	for course in courses:
		doc = _get_or_create_access(exam_plan, course)
		for key, val in settings.items():
			if key in _ALLOWED:
				setattr(doc, key, val)
		doc.save(ignore_permissions=True)

	frappe.db.commit()
	return True


@frappe.whitelist()
def assign_evaluator(exam_plan, courses, evaluators):
	"""
	Add one or more evaluators to the selected courses (append, do not replace).

	courses    – JSON list of course names
	evaluators – JSON list of dicts:
	             [{ evaluator_type, evaluator_name (Faculty name/link), evaluator_email }, ...]
	             evaluator_email is auto-fetched from the Faculty record when evaluator_type
	             is 'Custom'; pass it explicitly when available.
	"""
	import json

	if isinstance(courses, str):
		courses = json.loads(courses)
	if isinstance(evaluators, str):
		evaluators = json.loads(evaluators)

	if not courses:
		frappe.throw("No courses selected.")
	if not evaluators:
		frappe.throw("No evaluators provided.")

	for course in courses:
		doc = _get_or_create_access(exam_plan, course)

		# Existing Faculty name set (Link field) to avoid duplicates
		existing_names = {
			(row.evaluator_name or "").strip()
			for row in doc.evaluators
		}

		for ev in evaluators:
			ev_type  = ev.get("evaluator_type", "Class Faculty")
			fac_name = (ev.get("evaluator_name") or "").strip()

			# Deduplicate by Faculty name for Custom; one Class Faculty entry per course
			if ev_type == "Class Faculty":
				if any(r.evaluator_type == "Class Faculty" for r in doc.evaluators):
					continue
			else:
				if fac_name and fac_name in existing_names:
					continue

			# Resolve email from Faculty record if not provided
			email = (ev.get("evaluator_email") or "").strip()
			if not email and fac_name:
				email = frappe.db.get_value("Faculty", fac_name, "email") or ""

			doc.append("evaluators", {
				"evaluator_type":  ev_type,
				"evaluator_name":  fac_name,   # Link → Faculty
				"evaluator_email": email,
			})
			if fac_name:
				existing_names.add(fac_name)

		doc.save(ignore_permissions=True)

	frappe.db.commit()
	return True


@frappe.whitelist()
def remove_evaluator(exam_plan, course, evaluator_name):
	"""
	Remove a specific evaluator by Faculty name (Link value) from a course.
	For Class Faculty rows with no name, pass evaluator_name='' and
	evaluator_type='Class Faculty' — removes the first Class Faculty row.
	"""
	name = frappe.db.get_value(
		"Access Result Settings",
		{"exam_plan": exam_plan, "course": course},
		"name",
	)
	if not name:
		return True

	doc = frappe.get_doc("Access Result Settings", name)
	target = (evaluator_name or "").strip()

	if target:
		# Remove by Faculty link name
		doc.set(
			"evaluators",
			[row for row in doc.evaluators if (row.evaluator_name or "").strip() != target],
		)
	else:
		# Remove the first Class Faculty entry (no link)
		removed = False
		kept = []
		for row in doc.evaluators:
			if not removed and row.evaluator_type == "Class Faculty" and not row.evaluator_name:
				removed = True
				continue
			kept.append(row)
		doc.set("evaluators", kept)

	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return True


@frappe.whitelist()
def save_exam_visibility(exam_plan, courses, exam_types):
	"""
	Set the visible exam types for the selected courses (replaces existing).

	exam_types – JSON list of Exam Assessment Type names (the Link field value),
	             e.g. ["End Term Exam", "Mid-Term Exam"]
	"""
	import json

	if isinstance(courses, str):
		courses = json.loads(courses)
	if isinstance(exam_types, str):
		exam_types = json.loads(exam_types)

	if not courses:
		frappe.throw("No courses selected.")

	# Validate each exam_type exists in Exam Assessment Type
	valid_types = set(
		frappe.get_all("Exam Assessment Type", pluck="name")
	)

	for course in courses:
		doc = _get_or_create_access(exam_plan, course)
		doc.set("visible_exams", [])
		for et in exam_types:
			if et and et in valid_types:
				doc.append("visible_exams", {"exam_type": et})
		doc.save(ignore_permissions=True)

	frappe.db.commit()
	return True


@frappe.whitelist()
def get_faculty_list(search=""):
	"""
	Return Faculty records for the Assign Evaluator custom search box.
	evaluator_name is a Link to Faculty, so we search the Faculty doctype.
	"""
	filters = {}
	if search:
		filters["faculty_name"] = ["like", f"%{search}%"]

	return frappe.get_all(
		"Faculty",
		filters=filters,
		fields=["name", "faculty_name", "email", "image"],
		order_by="faculty_name asc",
		page_length=20,
	)


@frappe.whitelist()
def get_exam_types(search=""):
	"""
	Return Exam Assessment Type records for the Exam Visibility modal.
	exam_type is a Link to Exam Assessment Type, so we always load from that doctype.
	Returns: [{ name, type_name, assessment_type }, ...]
	         where `name` is the Link value to store in visible_exams.exam_type.
	"""
	filters = {"is_active": 1}
	if search:
		filters["type_name"] = ["like", f"%{search}%"]

	records = frappe.get_all(
		"Exam Assessment Type",
		filters=filters,
		fields=["name", "type_name", "assessment_type"],
		order_by="type_name asc",
	)
	return records


@frappe.whitelist()
def get_departments(search="", exam_plan=None):
	"""Return departments for the Course Results page department filter."""
	if exam_plan:
		search_clause = "AND d.department_name LIKE %(search)s" if search else ""
		return frappe.db.sql(
			f"""
			SELECT DISTINCT d.name, d.department_name
			FROM `tabDepartment` d
			JOIN `tabCourse` c ON c.department = d.name
			JOIN `tabCourse Schema Assignment` csa ON csa.course = c.name AND csa.exam_plan = %(exam_plan)s
			WHERE d.status = 'Active'
			{search_clause}
			ORDER BY d.department_name ASC
			LIMIT 100
			""",
			{"exam_plan": exam_plan, "search": f"%{search}%"},
			as_dict=True,
		)
	filters = {"status": "Active"}
	if search:
		filters["department_name"] = ["like", f"%{search}%"]
	return frappe.get_all(
		"Department",
		filters=filters,
		fields=["name", "department_name"],
		order_by="department_name asc",
		page_length=100,
	)


@frappe.whitelist()
def get_courses_by_department(department, exam_plan=None, search=""):
	"""Return courses in a department, optionally filtered by exam plan."""
	if exam_plan:
		search_clause = "AND c.course_name LIKE %(search)s" if search else ""
		return frappe.db.sql(
			f"""
			SELECT DISTINCT c.name, c.course_name, c.course_code, c.credit_value
			FROM `tabCourse` c
			JOIN `tabCourse Schema Assignment` csa ON csa.course = c.name AND csa.exam_plan = %(exam_plan)s
			WHERE c.department = %(department)s
			{search_clause}
			ORDER BY c.course_name ASC
			LIMIT 200
			""",
			{"exam_plan": exam_plan, "department": department, "search": f"%{search}%"},
			as_dict=True,
		)
	course_filters = {"department": department}
	if search:
		course_filters["course_name"] = ["like", f"%{search}%"]
	return frappe.get_all(
		"Course",
		filters=course_filters,
		fields=["name", "course_name", "course_code", "credit_value"],
		order_by="course_name asc",
		page_length=200,
	)


# ── New Course Results Page API ────────────────────────────────────────────────

def _get_marks_columns(evaluation_schema):
	"""Return ordered assessment config columns for a given evaluation schema."""
	rows = frappe.db.sql(
		"""
		SELECT sac.name, sac.component, ec.component_name,
		       sac.assessment_type, eat.type_name,
		       sac.label, sac.maximum_marks, sac.effective_marks,
		       sac.consider_for_pass_fail, sac.enrollment, sac.idx
		FROM `tabSchema Assessment Config` sac
		LEFT JOIN `tabExam Component` ec ON ec.name = sac.component
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = sac.assessment_type
		WHERE sac.parent = %(schema)s
		ORDER BY sac.idx ASC
		""",
		{"schema": evaluation_schema},
		as_dict=True,
	)
	# Ensure type_name falls back to assessment_type when JOIN returns NULL
	for row in rows:
		if not row.get("type_name"):
			row["type_name"] = row.get("assessment_type") or ""
		if not row.get("component_name"):
			row["component_name"] = row.get("component") or ""
	return rows


def _get_reexam_columns(evaluation_schema):
	"""Return ordered reexam config columns for a given evaluation schema."""
	rows = frappe.db.sql(
		"""
		SELECT src.name, src.component, ec.component_name,
		       src.assessment_type, eat.type_name,
		       src.label, src.maximum_marks, src.effective_marks,
		       src.enrollment, src.re_exam_type_category, src.idx
		FROM `tabSchema Reexam Config` src
		LEFT JOIN `tabExam Component` ec ON ec.name = src.component
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = src.assessment_type
		WHERE src.parent = %(schema)s
		ORDER BY src.idx ASC
		""",
		{"schema": evaluation_schema},
		as_dict=True,
	)
	for row in rows:
		if not row.get("type_name"):
			row["type_name"] = row.get("assessment_type") or ""
		if not row.get("component_name"):
			row["component_name"] = row.get("component") or ""
	return rows


@frappe.whitelist()
def get_course_info(course, exam_plan=None):
	"""Return full course info for the Course Results page."""
	course_doc = frappe.db.get_value(
		"Course", course,
		["course_name", "course_code", "credit_value", "department_name", "department"],
		as_dict=True,
	) or {}

	# Use provided exam plan or find best (Active first, then most recent)
	if exam_plan:
		assignment = frappe.db.sql(
			"""
			SELECT csa.exam_plan, csa.evaluation_schema, csa.grade_schema,
			       ep.exam_name, ep.status AS plan_status
			FROM `tabCourse Schema Assignment` csa
			JOIN `tabExam Plan` ep ON ep.name = csa.exam_plan
			WHERE csa.course = %(course)s AND csa.exam_plan = %(exam_plan)s
			LIMIT 1
			""",
			{"course": course, "exam_plan": exam_plan},
			as_dict=True,
		)
	else:
		assignment = frappe.db.sql(
			"""
			SELECT csa.exam_plan, csa.evaluation_schema, csa.grade_schema,
			       ep.exam_name, ep.status AS plan_status
			FROM `tabCourse Schema Assignment` csa
			JOIN `tabExam Plan` ep ON ep.name = csa.exam_plan
			WHERE csa.course = %(course)s
			ORDER BY FIELD(ep.status, 'Active', 'Inactive') ASC, ep.creation DESC
			LIMIT 1
			""",
			{"course": course},
			as_dict=True,
		)
	assignment = assignment[0] if assignment else {}

	access = frappe.db.get_value(
		"Access Result Settings",
		{"exam_plan": assignment.get("exam_plan"), "course": course},
		["view_access", "edit_access", "edit_deadline", "mask_student_info",
		 "auto_generate_grade_access", "status"],
		as_dict=True,
	) or {}

	count_row = frappe.db.sql(
		"""
		SELECT COUNT(DISTINCT scm.student) AS cnt
		FROM `tabStudent Course Marks` scm
		WHERE scm.course = %(course)s
		  AND scm.exam_plan = %(exam_plan)s
		""",
		{"course": course, "exam_plan": assignment.get("exam_plan", "")},
		as_dict=True,
	)

	columns = []
	reexam_columns = []
	if assignment.get("evaluation_schema"):
		columns = _get_marks_columns(assignment["evaluation_schema"])
		reexam_columns = _get_reexam_columns(assignment["evaluation_schema"])

	# Format edit_deadline for display
	edit_deadline = ""
	raw_dl = access.get("edit_deadline")
	if raw_dl:
		try:
			edit_deadline = frappe.utils.format_datetime(str(raw_dl), "dd MMM yy hh:mm a")
		except Exception:
			edit_deadline = str(raw_dl)

	return {
		"course_name":     course_doc.get("course_name", ""),
		"course_code":     course_doc.get("course_code", ""),
		"credit_value":    course_doc.get("credit_value", 0),
		"department_name": course_doc.get("department_name", ""),
		"exam_plan":       assignment.get("exam_plan", ""),
		"exam_name":       assignment.get("exam_name", ""),
		"evaluation_schema": assignment.get("evaluation_schema", ""),
		"grade_schema":    assignment.get("grade_schema", ""),
		"student_count":   count_row[0]["cnt"] if count_row else 0,
		"view_access":     int(access.get("view_access", 1)),
		"edit_access":     int(access.get("edit_access", 1)),
		"edit_deadline":   edit_deadline,
		"mask_student_info": int(access.get("mask_student_info", 0)),
		"status":          access.get("status", "UNLOCKED"),
		"columns":         columns,
		"reexam_columns":  reexam_columns,
	}


@frappe.whitelist()
def get_course_students_paged(course, exam_plan="", search="", page=1, page_length=20,
                               sort_by="registration_id", sort_order="asc",
                               status_filter="", inst_programmes="",
                               inst_batches="", inst_course_types="",
                               grade_filter="", pass_filter=""):
	"""Return paginated students from Student Course Marks for a course."""
	import json
	page        = int(page)
	page_length = int(page_length)
	offset      = (page - 1) * page_length

	# Parse institutional filter lists (JSON arrays)
	f_programmes  = json.loads(inst_programmes)  if inst_programmes  else []
	f_batches     = json.loads(inst_batches)     if inst_batches     else []
	f_course_types= json.loads(inst_course_types)if inst_course_types else []

	# Resolve exam plan and grade schema
	if not exam_plan:
		assignment = frappe.db.sql(
			"""
			SELECT csa.exam_plan, csa.grade_schema
			FROM `tabCourse Schema Assignment` csa
			JOIN `tabExam Plan` ep ON ep.name = csa.exam_plan
			WHERE csa.course = %(course)s
			ORDER BY FIELD(ep.status, 'Active', 'Inactive') ASC, ep.creation DESC
			LIMIT 1
			""",
			{"course": course},
			as_dict=True,
		)
		exam_plan    = assignment[0]["exam_plan"]    if assignment else ""
		grade_schema = assignment[0]["grade_schema"] if assignment else ""
	else:
		row = frappe.db.get_value(
			"Course Schema Assignment",
			{"course": course, "exam_plan": exam_plan},
			"grade_schema",
		)
		grade_schema = row or ""

	# Determine failed grades for pass/fail filter
	failed_grades = []
	if pass_filter and grade_schema:
		rows = frappe.db.sql(
			"SELECT grade FROM `tabGrading Schema Component` WHERE parent = %s AND failed = 1",
			grade_schema,
			as_list=True,
		)
		failed_grades = [r[0] for r in rows]

	sort_col = {
		"total_marks": "scm.total_marks",
		"name":        "CONCAT_WS(' ', sm.first_name, sm.last_name)",
	}.get(sort_by, "sm.registration_id")
	sort_dir = "DESC" if sort_order == "desc" else "ASC"

	extra_cond = ""
	params = {"course": course, "exam_plan": exam_plan, "lim": page_length, "off": offset}
	if search:
		extra_cond += (
			" AND (sm.registration_id LIKE %(search)s "
			"OR sm.first_name LIKE %(search)s "
			"OR sm.last_name LIKE %(search)s)"
		)
		params["search"] = f"%{search}%"
	if status_filter:
		extra_cond += " AND sm.student_status = %(status_filter)s"
		params["status_filter"] = status_filter
	if f_programmes:
		placeholders = ",".join([f"%(prog_{i})s" for i in range(len(f_programmes))])
		extra_cond += f" AND sm.programme IN ({placeholders})"
		for i, v in enumerate(f_programmes):
			params[f"prog_{i}"] = v
	if f_batches:
		placeholders = ",".join([f"%(batch_{i})s" for i in range(len(f_batches))])
		extra_cond += f" AND sm.batch_year IN ({placeholders})"
		for i, v in enumerate(f_batches):
			params[f"batch_{i}"] = v
	# course_type filter skipped — field lives on Program Enrollment, not Student Course Marks
	if grade_filter:
		extra_cond += " AND scm.grade = %(grade_filter)s"
		params["grade_filter"] = grade_filter
	if pass_filter == "failed" and failed_grades:
		placeholders = ",".join([f"%(fg_{i})s" for i in range(len(failed_grades))])
		extra_cond += f" AND scm.grade IN ({placeholders})"
		for i, v in enumerate(failed_grades):
			params[f"fg_{i}"] = v
	elif pass_filter == "passed" and failed_grades:
		placeholders = ",".join([f"%(fg_{i})s" for i in range(len(failed_grades))])
		extra_cond += f" AND (scm.grade IS NOT NULL AND scm.grade != '' AND scm.grade NOT IN ({placeholders}))"
		for i, v in enumerate(failed_grades):
			params[f"fg_{i}"] = v
	elif pass_filter == "graded":
		extra_cond += " AND scm.grade IS NOT NULL AND scm.grade != ''"
	elif pass_filter == "not_graded":
		extra_cond += " AND (scm.grade IS NULL OR scm.grade = '')"

	students = frappe.db.sql(
		f"""
		SELECT DISTINCT
			sm.name               AS student,
			sm.registration_id,
			CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
			sm.student_status,
			sm.account_status,
			sm.programme,
			sm.batch_year,
			sm.intake,
			sm.department,
			sm.passport_size_photo,
			NULL AS section
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s
		  AND scm.exam_plan = %(exam_plan)s
		{extra_cond}
		ORDER BY {sort_col} {sort_dir}
		LIMIT %(lim)s OFFSET %(off)s
		""",
		params,
		as_dict=True,
	)

	total_row = frappe.db.sql(
		f"""
		SELECT COUNT(DISTINCT scm.student) AS cnt
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s
		  AND scm.exam_plan = %(exam_plan)s
		{extra_cond}
		""",
		params,
		as_dict=True,
	)

	return {
		"students": students,
		"total":    total_row[0]["cnt"] if total_row else 0,
	}


@frappe.whitelist()
def get_institutional_filter_options(course):
	"""Return distinct programme, batch_year, and course_type values for students in this course."""
	assignment = frappe.db.sql(
		"""
		SELECT csa.exam_plan
		FROM `tabCourse Schema Assignment` csa
		JOIN `tabExam Plan` ep ON ep.name = csa.exam_plan
		WHERE csa.course = %(course)s
		ORDER BY FIELD(ep.status, 'Active', 'Inactive') ASC, ep.creation DESC
		LIMIT 1
		""",
		{"course": course},
		as_dict=True,
	)
	exam_plan = assignment[0]["exam_plan"] if assignment else ""

	rows = frappe.db.sql(
		"""
		SELECT DISTINCT sm.programme, sm.batch_year
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)

	programmes = sorted(set(r["programme"]        for r in rows if r.get("programme")))
	batches    = sorted(set(str(r["batch_year"])   for r in rows if r.get("batch_year")), reverse=True)

	# course_type lives on Program Enrollment, not Student Course Marks.
	# Return static options for display; SQL filtering is handled via programme/batch.
	course_types = ["Regular", "Backlog"]

	return {
		"programmes":   programmes,
		"batches":      batches,
		"course_types": course_types,
	}


@frappe.whitelist()
def sync_students_from_enrollment(course):
	"""Create Student Course Marks records for all students enrolled in this course via Student Enrollment."""
	# Get the active exam plan and evaluation schema for the course
	assignment = frappe.db.sql(
		"""
		SELECT csa.exam_plan, csa.evaluation_schema
		FROM `tabCourse Schema Assignment` csa
		JOIN `tabExam Plan` ep ON ep.name = csa.exam_plan
		WHERE csa.course = %(course)s
		ORDER BY FIELD(ep.status, 'Active', 'Inactive') ASC, ep.creation DESC
		LIMIT 1
		""",
		{"course": course},
		as_dict=True,
	)
	if not assignment:
		frappe.throw("No Exam Plan assigned for this course. Please set up a Course Schema Assignment first.")

	exam_plan         = assignment[0]["exam_plan"]
	evaluation_schema = assignment[0]["evaluation_schema"] or ""

	# Find students enrolled in this course via Student Enrollment
	enrolled_students = frappe.db.sql(
		"""
		SELECT DISTINCT se.student
		FROM `tabStudent Enrollment` se
		INNER JOIN `tabProgram Enrollment` pe ON pe.parent = se.name
		WHERE pe.course = %(course)s
		  AND se.status = 'Enrolled'
		  AND se.student IS NOT NULL
		  AND se.student != ''
		""",
		{"course": course},
		as_dict=True,
	)

	if not enrolled_students:
		frappe.throw("No enrolled students found for this course in Student Enrollment.")

	added = 0
	skipped = 0
	for row in enrolled_students:
		student = row["student"]
		exists = frappe.db.exists("Student Course Marks", {
			"course":     course,
			"exam_plan":  exam_plan,
			"student":    student,
		})
		if exists:
			skipped += 1
			continue
		doc = frappe.new_doc("Student Course Marks")
		doc.course             = course
		doc.exam_plan          = exam_plan
		doc.student            = student
		doc.evaluation_schema  = evaluation_schema
		doc.status             = "Draft"
		doc.consider_for_sgpa  = 1
		doc.insert(ignore_permissions=True)
		added += 1

	frappe.db.commit()
	return {"added": added, "skipped": skipped, "total": len(enrolled_students)}


@frappe.whitelist()
def sync_students_from_class_config(course, class_config, course_type):
	"""Create Student Course Marks records for all students in the given Class Configuration."""
	# Validate the class config belongs to this course
	cc_course = frappe.db.get_value("Class Configuration", class_config, "course")
	if cc_course != course:
		frappe.throw("The selected Class does not belong to this course.")

	# Get active exam plan and evaluation schema
	assignment = frappe.db.sql(
		"""
		SELECT csa.exam_plan, csa.evaluation_schema
		FROM `tabCourse Schema Assignment` csa
		JOIN `tabExam Plan` ep ON ep.name = csa.exam_plan
		WHERE csa.course = %(course)s
		ORDER BY FIELD(ep.status, 'Active', 'Inactive') ASC, ep.creation DESC
		LIMIT 1
		""",
		{"course": course},
		as_dict=True,
	)
	if not assignment:
		frappe.throw("No Exam Plan assigned for this course. Please set up a Course Schema Assignment first.")

	exam_plan         = assignment[0]["exam_plan"]
	evaluation_schema = assignment[0]["evaluation_schema"] or ""

	# Get all students from the Class Configuration
	class_students = frappe.db.sql(
		"""
		SELECT cs.student
		FROM `tabClass Student` cs
		WHERE cs.parent = %(class_config)s
		  AND cs.student IS NOT NULL
		  AND cs.student != ''
		""",
		{"class_config": class_config},
		as_dict=True,
	)

	if not class_students:
		frappe.throw("No students found in the selected Class Configuration.")

	added = 0
	skipped = 0
	for row in class_students:
		student = row["student"]
		exists = frappe.db.exists("Student Course Marks", {
			"course":    course,
			"exam_plan": exam_plan,
			"student":   student,
		})
		if exists:
			skipped += 1
			continue
		doc = frappe.new_doc("Student Course Marks")
		doc.course            = course
		doc.exam_plan         = exam_plan
		doc.student           = student
		doc.evaluation_schema = evaluation_schema
		doc.status            = "Draft"
		doc.consider_for_sgpa = 1
		doc.insert(ignore_permissions=True)
		added += 1

		# Update course_type in Program Enrollment if present
		pe = frappe.db.get_value(
			"Program Enrollment",
			{"parent": ["like", "%"], "course": course, "parenttype": "Student Enrollment"},
			"name",
		)
		# Find program enrollment rows for this student and course
		pe_rows = frappe.db.sql(
			"""
			SELECT pe.name
			FROM `tabProgram Enrollment` pe
			INNER JOIN `tabStudent Enrollment` se ON se.name = pe.parent
			WHERE pe.course = %(course)s AND se.student = %(student)s
			LIMIT 1
			""",
			{"course": course, "student": student},
			as_dict=True,
		)
		if pe_rows and course_type:
			frappe.db.set_value("Program Enrollment", pe_rows[0]["name"], "course_type", course_type)

	frappe.db.commit()
	return {"added": added, "skipped": skipped, "total": len(class_students)}


@frappe.whitelist()
def get_student_hover_info(student, course):
	"""Return student profile + section for the hover popup."""
	sm = frappe.db.get_value(
		"Student Master", student,
		["registration_id", "first_name", "last_name",
		 "official_email_id", "email", "programme",
		 "batch_year", "current_year", "current_term", "intake", "department"],
		as_dict=True,
	)
	if not sm:
		return {}

	cohort_name = ""
	if sm.get("programme"):
		cohort_name = (
			frappe.db.get_value("Cohort", sm["programme"], "cohort_name")
			or sm["programme"]
		)

	# Batch: prefer batch_year, fall back to current_year
	batch_val = sm.get("batch_year") or sm.get("current_year") or ""

	section_row = frappe.db.sql(
		"""
		SELECT cc.section
		FROM `tabClass Student` cs
		INNER JOIN `tabClass Configuration` cc ON cs.parent = cc.name
		WHERE cs.student = %(student)s AND cc.course = %(course)s
		LIMIT 1
		""",
		{"student": student, "course": course},
		as_dict=True,
	)

	return {
		"student_name":    " ".join(filter(None, [sm.first_name, sm.last_name])),
		"registration_id": sm.registration_id or student,
		"email":           sm.official_email_id or sm.email or "",
		"programme":       cohort_name,
		"department":      sm.department or "",
		"batch":           batch_val,
		"current_term":    sm.current_term or "",
		"intake":          sm.intake or "",
		"section":         section_row[0]["section"] if section_row else "",
	}


@frappe.whitelist()
def get_marks_for_students(course, exam_plan, student_ids):
	"""Return marks data keyed by student → {entries: {comp|atype → marks}, totals, status fields}."""
	import json as _json
	if isinstance(student_ids, str):
		student_ids = _json.loads(student_ids)
	if not student_ids or not exam_plan:
		return {}

	# Fetch header-level data (totals, grade, status fields)
	header_rows = frappe.db.sql(
		"""
		SELECT scm.student, scm.total_marks, scm.grade,
		       scm.status, scm.enrollment_status, scm.attendance_status,
		       scm.mfa, scm.fairness_status, scm.consider_for_sgpa, scm.remark,
		       scm.updated_final_marks, scm.updated_grade
		FROM `tabStudent Course Marks` scm
		WHERE scm.course = %(course)s
		  AND scm.exam_plan = %(exam_plan)s
		  AND scm.student IN %(students)s
		""",
		{"course": course, "exam_plan": exam_plan, "students": tuple(student_ids)},
		as_dict=True,
	)

	# Fetch Approved FA MFA Applications for this course
	fa_mfa_apps = frappe.db.sql(
		"""
		SELECT student
		FROM `tabFA MFA Application`
		WHERE course = %(course)s AND docstatus = 1 AND status = 'Approved'
		""",
		{"course": course},
		as_dict=True
	)
	approved_fa_mfa_students = {row["student"] for row in fa_mfa_apps}

	result = {}
	for row in header_rows:
		s = row["student"]
		result[s] = {
			"total":               row["total_marks"],
			"grade":               row["grade"] or "",
			"status":              row["status"] or "",
			"enrollment_status":   row["enrollment_status"] or "",
			"attendance_status":   row["attendance_status"] or "",
			"mfa":                 "Yes" if s in approved_fa_mfa_students else (row.get("mfa") or "No"),
			"fairness_status":     row["fairness_status"] or "",
			"consider_for_sgpa":   int(row["consider_for_sgpa"] or 0),
			"remark":              row["remark"] or "",
			"updated_final_marks": row["updated_final_marks"],
			"updated_grade":       row["updated_grade"] or "",
			"entries":             {},
		}

	# Fetch per-assessment marks
	entry_rows = frappe.db.sql(
		"""
		SELECT scm.student, sme.component, sme.assessment_type,
		       sme.marks, sme.revaluation_marks
		FROM `tabStudent Course Marks` scm
		JOIN `tabStudent Marks Entry` sme ON sme.parent = scm.name
		WHERE scm.course = %(course)s
		  AND scm.exam_plan = %(exam_plan)s
		  AND scm.student IN %(students)s
		""",
		{"course": course, "exam_plan": exam_plan, "students": tuple(student_ids)},
		as_dict=True,
	)

	for row in entry_rows:
		s = row["student"]
		if s not in result:
			result[s] = {"total": None, "grade": "",
			             "status": "", "enrollment_status": "", "attendance_status": "",
			             "mfa": "Yes" if s in approved_fa_mfa_students else "No",
			             "fairness_status": "", "consider_for_sgpa": 1, "remark": "",
			             "updated_final_marks": None, "updated_grade": "", "entries": {}}
		key = (row["component"] or "") + "|" + (row["assessment_type"] or "")
		result[s]["entries"][key] = {
			"marks":             row["marks"],
			"revaluation_marks": row["revaluation_marks"],
		}

	return result


@frappe.whitelist()
def get_course_overview(exam_plan, course):
	"""Return course info panel data for the Course Results page."""
	course_doc = frappe.db.get_value(
		"Course",
		course,
		["course_name", "course_code", "credit_value", "department_name"],
		as_dict=True,
	) or {}

	schema = frappe.db.get_value(
		"Course Schema Assignment",
		{"exam_plan": exam_plan, "course": course},
		["evaluation_schema", "grade_schema"],
		as_dict=True,
	) or {}

	access = frappe.db.get_value(
		"Access Result Settings",
		{"exam_plan": exam_plan, "course": course},
		["view_access", "edit_access", "mask_student_info", "status",
		 "auto_generate_grade_access", "edit_grade_access"],
		as_dict=True,
	) or {}

	count_row = frappe.db.sql(
		"""
		SELECT COUNT(DISTINCT cs.student) AS cnt
		FROM `tabClass Student` cs
		INNER JOIN `tabClass Configuration` cc ON cs.parent = cc.name
		WHERE cc.course = %(course)s
		""",
		{"course": course},
		as_dict=True,
	)

	return {
		"course_name":              course_doc.get("course_name", ""),
		"course_code":              course_doc.get("course_code", ""),
		"credit_value":             course_doc.get("credit_value", 0),
		"department_name":          course_doc.get("department_name", ""),
		"evaluation_schema":        schema.get("evaluation_schema", ""),
		"grade_schema":             schema.get("grade_schema", ""),
		"view_access":              int(access.get("view_access", 1)),
		"edit_access":              int(access.get("edit_access", 1)),
		"mask_student_info":        int(access.get("mask_student_info", 0)),
		"auto_generate_grade_access": int(access.get("auto_generate_grade_access", 0)),
		"edit_grade_access":        int(access.get("edit_grade_access", 0)),
		"status":                   access.get("status", "UNLOCKED"),
		"student_count":            (count_row[0]["cnt"] if count_row else 0),
	}


@frappe.whitelist()
def get_course_students_with_marks(exam_plan, course, search="", page=1, page_length=20):
	"""
	Return paginated students enrolled in the course via Class Configuration,
	along with placeholder marks structure per exam type.
	"""
	page        = int(page)
	page_length = int(page_length)
	offset      = (page - 1) * page_length

	search_cond = ""
	params      = {"course": course, "lim": page_length, "off": offset}

	if search:
		search_cond    = "AND (sm.registration_id LIKE %(search)s OR sm.first_name LIKE %(search)s OR sm.last_name LIKE %(search)s)"
		params["search"] = f"%{search}%"

	students = frappe.db.sql(
		f"""
		SELECT DISTINCT
			sm.name           AS student,
			sm.registration_id,
			CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
			sm.email,
			sm.official_email_id,
			sm.student_status,
			sm.account_status,
			sm.programme,
			sm.batch_year,
			sm.intake,
			sm.department
		FROM `tabClass Student` cs
		INNER JOIN `tabClass Configuration` cc ON cs.parent = cc.name
		LEFT JOIN `tabStudent Master` sm ON sm.name = cs.student
		WHERE cc.course = %(course)s
		{search_cond}
		ORDER BY sm.registration_id ASC
		LIMIT %(lim)s OFFSET %(off)s
		""",
		params,
		as_dict=True,
	)

	total_row = frappe.db.sql(
		f"""
		SELECT COUNT(DISTINCT cs.student) AS cnt
		FROM `tabClass Student` cs
		INNER JOIN `tabClass Configuration` cc ON cs.parent = cc.name
		LEFT JOIN `tabStudent Master` sm ON sm.name = cs.student
		WHERE cc.course = %(course)s
		{search_cond}
		""",
		params,
		as_dict=True,
	)

	# Exam types visible for this course (from Result Exam Visibility)
	visible_types = frappe.db.sql(
		"""
		SELECT rev.exam_type, eat.type_name
		FROM `tabResult Exam Visibility` rev
		INNER JOIN `tabAccess Result Settings` cra ON rev.parent = cra.name
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = rev.exam_type
		WHERE cra.exam_plan = %(exam_plan)s AND cra.course = %(course)s
		""",
		{"exam_plan": exam_plan, "course": course},
		as_dict=True,
	)

	return {
		"students":      students,
		"total":         (total_row[0]["cnt"] if total_row else 0),
		"page":          page,
		"page_length":   page_length,
		"exam_types":    visible_types,
	}


@frappe.whitelist()
def get_student_profile(student):
	"""Return detailed student info for hover popup on Course Results page."""
	sm = frappe.db.get_value(
		"Student Master",
		student,
		["registration_id", "first_name", "last_name", "email", "official_email_id",
		 "programme", "batch_year", "intake", "student_status", "account_status",
		 "phone", "department"],
		as_dict=True,
	)
	if not sm:
		frappe.throw("Student not found.")
	# Get cohort/programme label
	cohort_name = ""
	if sm.get("programme"):
		cohort_name = frappe.db.get_value("Cohort", sm["programme"], "cohort_name") or sm["programme"]
	sm["cohort_name"] = cohort_name
	return sm


@frappe.whitelist()
def get_calc_settings(evaluation_schema):
	"""Return calculation settings for the evaluation schema."""
	doc = frappe.db.get_value(
		"Evaluation Schema",
		evaluation_schema,
		["calc_higher_revaluation", "calc_higher_makeup", "calc_higher_reexam"],
		as_dict=True,
	) or {}
	return {
		"calc_higher_revaluation": int(doc.get("calc_higher_revaluation") or 0),
		"calc_higher_makeup":      int(doc.get("calc_higher_makeup") or 0),
		"calc_higher_reexam":      int(doc.get("calc_higher_reexam") or 0),
	}


@frappe.whitelist()
def save_calc_settings(evaluation_schema, calc_higher_revaluation, calc_higher_makeup, calc_higher_reexam):
	"""Save calculation settings to the evaluation schema."""
	doc = frappe.get_doc("Evaluation Schema", evaluation_schema)
	doc.calc_higher_revaluation = int(calc_higher_revaluation)
	doc.calc_higher_makeup      = int(calc_higher_makeup)
	doc.calc_higher_reexam      = int(calc_higher_reexam)
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return True


@frappe.whitelist()
def get_evaluation_schema_details(name):
	"""Return full evaluation schema for the popup view."""
	doc = frappe.get_doc("Evaluation Schema", name)

	components = []
	for row in doc.schema_components:
		comp_name = frappe.db.get_value("Exam Component", row.component, "component_name") or row.component
		components.append({
			"component":       row.component,
			"component_name":  comp_name,
			"effective_max_marks": row.effective_max_marks,
			"weightage":       row.weightage,
			"passing_marks":   row.passing_marks,
			"consider_for_pass_fail": int(row.consider_for_pass_fail or 0),
		})

	assessments = []
	for row in doc.assessment_configs:
		comp_name = frappe.db.get_value("Exam Component", row.component, "component_name") or row.component
		type_name = frappe.db.get_value("Exam Assessment Type", row.assessment_type, "type_name") or row.assessment_type
		assessments.append({
			"component":       row.component,
			"component_name":  comp_name,
			"assessment_type": row.assessment_type,
			"type_name":       type_name,
			"label":           row.label,
			"maximum_marks":   row.maximum_marks,
			"effective_marks": row.effective_marks,
			"passing_marks":   row.passing_marks,
			"consider_for_pass_fail": int(row.consider_for_pass_fail or 0),
			"enrollment":      row.enrollment,
		})

	return {
		"schema_name":   doc.schema_name,
		"description":   doc.description or "",
		"total_marks":   doc.total_marks,
		"passing_marks": doc.passing_marks,
		"components":    components,
		"assessments":   assessments,
		"calc_higher_revaluation": int(doc.calc_higher_revaluation or 0),
		"calc_higher_makeup":      int(doc.calc_higher_makeup or 0),
		"calc_higher_reexam":      int(doc.calc_higher_reexam or 0),
	}


@frappe.whitelist()
def get_grading_schema_details(name):
	"""Return full grading schema for the popup view."""
	doc = frappe.get_doc("Grading Schema", name)

	grades = []
	for row in doc.grades:
		grades.append({
			"grade":               row.grade,
			"qualitative_meaning": row.qualitative_meaning or "",
			"from_operator":       row.from_operator or ">=",
			"marks_from":          row.marks_from,
			"to_operator":         row.to_operator or "<",
			"marks_to":            row.marks_to,
			"grade_point":         row.grade_point,
			"failed":              int(row.failed or 0),
			"consider_for_sgpa":   int(row.consider_for_sgpa or 0),
		})

	return {
		"grading_schema_name": doc.grading_schema_name,
		"schema_name":         doc.schema_name or "",
		"description":         doc.description or "",
		"maximum_marks":       doc.maximum_marks,
		"grading_type":        doc.grading_type,
		"grades":              grades,
	}


@frappe.whitelist()
def get_result_summary(exam_plan):
	"""Return high-level counts for the result settings dashboard."""
	total      = frappe.db.count("Course", {})
	configured = frappe.db.count("Access Result Settings", {"exam_plan": exam_plan})
	locked     = frappe.db.count("Access Result Settings", {"exam_plan": exam_plan, "status": "LOCKED"})
	unlocked   = configured - locked
	return {
		"total_courses": total,
		"configured":    configured,
		"locked":        locked,
		"unlocked":      unlocked,
		"unconfigured":  total - configured,
	}


@frappe.whitelist()
def download_grade_sample(course, exam_plan, include_students=1):
	"""Generate a sample Excel file for bulk grade upload."""
	import io
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment
	except ImportError:
		frappe.throw("openpyxl is required. Run: bench pip install openpyxl")

	include_students = frappe.utils.cint(include_students)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Grade Upload"

	header_font  = Font(bold=True, color="FFFFFF")
	header_fill  = PatternFill("solid", fgColor="C0392B")
	center_align = Alignment(horizontal="center")

	headers = ["Registration ID", "Email ID", "Student Name", "Grade"]
	for col, h in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col, value=h)
		cell.font  = header_font
		cell.fill  = header_fill
		cell.alignment = center_align

	ws.column_dimensions["A"].width = 20
	ws.column_dimensions["B"].width = 30
	ws.column_dimensions["C"].width = 30
	ws.column_dimensions["D"].width = 15

	if include_students:
		students = frappe.db.sql(
			"""
			SELECT sm.registration_id, sm.official_email_id,
			       CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name
			FROM `tabStudent Course Marks` scm
			LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
			WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
			ORDER BY sm.registration_id ASC
			""",
			{"course": course, "exam_plan": exam_plan},
			as_dict=True,
		)
		for row_idx, s in enumerate(students, 2):
			ws.cell(row=row_idx, column=1, value=s.registration_id or "")
			ws.cell(row=row_idx, column=2, value=s.official_email_id or "")
			ws.cell(row=row_idx, column=3, value=s.student_name or "")
			ws.cell(row=row_idx, column=4, value="")

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	file_name = f"grade_upload_{course}.xlsx"
	file_doc = frappe.get_doc({
		"doctype":    "File",
		"file_name":  file_name,
		"is_private": 1,
		"content":    buf.read(),
	})
	file_doc.save(ignore_permissions=True)
	return {"file_url": file_doc.file_url}


@frappe.whitelist()
def bulk_upload_grades(course, exam_plan, file_url):
	"""Process uploaded Excel file and update grades in Student Course Marks."""
	import io
	try:
		import openpyxl
	except ImportError:
		frappe.throw("openpyxl is required. Run: bench pip install openpyxl")

	# Fetch the file
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	file_path = file_doc.get_full_path()

	wb = openpyxl.load_workbook(file_path, data_only=True)
	ws = wb.active

	headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

	def col_idx(name):
		try:
			return headers.index(name)
		except ValueError:
			return -1

	reg_col   = col_idx("Registration ID")
	email_col = col_idx("Email ID")
	grade_col = col_idx("Grade")

	if grade_col == -1:
		frappe.throw("Column 'Grade' not found in the uploaded file.")

	updated = 0
	errors  = []

	for row in ws.iter_rows(min_row=2, values_only=True):
		reg_id = str(row[reg_col]).strip() if reg_col >= 0 and row[reg_col] else ""
		email  = str(row[email_col]).strip() if email_col >= 0 and row[email_col] else ""
		grade  = str(row[grade_col]).strip() if row[grade_col] else ""

		if not grade:
			continue

		# Find student
		student = None
		if reg_id:
			student = frappe.db.get_value("Student Master", {"registration_id": reg_id}, "name")
		if not student and email:
			student = frappe.db.get_value(
				"Student Master",
				[["official_email_id", "=", email], ["email", "=", email]],
				"name",
				or_filters=True,
			) or frappe.db.get_value("Student Master", {"official_email_id": email}, "name") \
			  or frappe.db.get_value("Student Master", {"email": email}, "name")

		if not student:
			errors.append(f"Student not found: {reg_id or email}")
			continue

		scm_name = frappe.db.get_value(
			"Student Course Marks",
			{"course": course, "exam_plan": exam_plan, "student": student},
			"name",
		)
		if not scm_name:
			errors.append(f"No result record for: {reg_id or email}")
			continue

		frappe.db.set_value("Student Course Marks", scm_name, "grade", grade)
		updated += 1

	frappe.db.commit()
	return {"updated": updated, "errors": errors}


@frappe.whitelist()
def save_student_remark(course, exam_plan, student, remark):
	"""Save a remark on a Student Course Marks record."""
	scm_name = frappe.db.get_value(
		"Student Course Marks",
		{"course": course, "exam_plan": exam_plan, "student": student},
		"name",
	)
	if not scm_name:
		frappe.throw("Result record not found for this student.")
	frappe.db.set_value("Student Course Marks", scm_name, "remark", remark)
	frappe.db.commit()
	return {"success": True}


# ── Marks Entry & Grade Calculation ───────────────────────────────────────────

@frappe.whitelist()
def save_status(course, exam_plan, student, field, value):
	"""Save/update a status field on a Student Course Marks record."""
	VALID_FIELDS = {"enrollment_status", "attendance_status", "fairness_status", "mfa", "grade", "updated_grade"}
	if field not in VALID_FIELDS:
		frappe.throw(f"Invalid field: {field}")

	scm_name = frappe.db.get_value(
		"Student Course Marks",
		{"course": course, "exam_plan": exam_plan, "student": student},
		"name",
	)
	if not scm_name:
		frappe.throw("Result record not found for this student.")

	frappe.db.set_value("Student Course Marks", scm_name, field, value)
	# When marking absent, force grade = "Ab" and clear total
	if field == "attendance_status" and (value or "").lower() == "absent":
		frappe.db.set_value("Student Course Marks", scm_name, {"grade": "Ab", "total_marks": 0})
	frappe.db.commit()
	return True

@frappe.whitelist()
def save_marks(course, exam_plan, student, component, assessment_type, marks_field, value):
	"""Save/update a single marks entry for a student assessment. Returns updated total and grade."""
	VALID_FIELDS = {"marks", "revaluation_marks"}
	if marks_field not in VALID_FIELDS:
		frappe.throw(f"Invalid marks_field: {marks_field}")

	fvalue = frappe.utils.flt(value) if value not in (None, "", "null", "--") else None

	scm_name = frappe.db.get_value(
		"Student Course Marks",
		{"course": course, "exam_plan": exam_plan, "student": student},
		"name",
	)
	if not scm_name:
		frappe.throw("Result record not found for this student.")

	# Check lock status
	access = frappe.db.get_value(
		"Access Result Settings",
		{"exam_plan": exam_plan, "course": course},
		["edit_access", "status"],
		as_dict=True,
	) or {}
	if access.get("status") == "LOCKED":
		frappe.throw("Result entry is locked for this course.")
	if not int(access.get("edit_access") or 1):
		frappe.throw("Edit access is disabled for this course.")

	# Update or create the marks entry row
	sme_name = frappe.db.get_value(
		"Student Marks Entry",
		{"parent": scm_name, "component": component, "assessment_type": assessment_type},
		"name",
	)
	if sme_name:
		frappe.db.set_value("Student Marks Entry", sme_name, marks_field, fvalue if fvalue is not None else 0.0)
	else:
		if fvalue is not None:
			# marks_field is already validated against VALID_FIELDS — safe to use in f-string
			frappe.db.sql(
				f"""
				INSERT INTO `tabStudent Marks Entry`
				(name, creation, modified, modified_by, owner,
				 parent, parenttype, parentfield, component, assessment_type, {marks_field})
				VALUES (%(name_val)s, NOW(), NOW(), %(user)s, %(user)s,
				        %(parent)s, 'Student Course Marks', 'marks_entries',
				        %(comp)s, %(atype)s, %(val)s)
				""",
				{
					"name_val": frappe.generate_hash("", 10),
					"user":     frappe.session.user,
					"parent":   scm_name,
					"comp":     component,
					"atype":    assessment_type,
					"val":      fvalue,
				},
			)

	frappe.db.commit()
	return _recalculate_student_marks(scm_name, course, exam_plan)


def _recalculate_student_marks(scm_name, course, exam_plan):
	"""Recalculate total marks and grade for a student. Updates SCM and returns new values.

	Calculation rules (in priority order for each assessment entry):
	  1. moderated_marks — set by moderation workflow; takes precedence over marks.
	  2. If calc_higher_revaluation is enabled on the Evaluation Schema, the
	     effective per-entry marks = max(moderated_marks or marks, revaluation_marks).
	  3. Otherwise effective per-entry marks = moderated_marks or marks.

	Re-exam total is computed similarly using Schema Reexam Config rows with
	calc_higher_reexam controlling whether revaluation_marks are considered.

	The updated_final_marks = max(regular_total, reexam_total).
	updated_grade is looked up using the Re Exam Composition of the grading schema
	(when use_reexam_composition is enabled on that schema).
	"""
	# Absent students get grade "Ab" — skip all calculations
	attendance_status = frappe.db.get_value("Student Course Marks", scm_name, "attendance_status") or ""
	if attendance_status.lower() == "absent":
		frappe.db.set_value("Student Course Marks", scm_name, {"grade": "Ab", "total_marks": 0})
		return {"total": 0, "grade": "Ab", "updated_final_marks": 0, "updated_grade": "Ab"}

	csa = frappe.db.get_value(
		"Course Schema Assignment",
		{"course": course, "exam_plan": exam_plan},
		["evaluation_schema", "grade_schema"],
		as_dict=True,
	) or {}
	eval_schema  = csa.get("evaluation_schema")
	grade_schema = csa.get("grade_schema")

	if not eval_schema:
		# No evaluation schema — fall back to looking up the grade from the stored
		# total_marks so that existing marks are at least graded correctly.
		if grade_schema:
			stored_total = frappe.utils.flt(
				frappe.db.get_value("Student Course Marks", scm_name, "total_marks") or 0
			)
			grade         = _lookup_grade(grade_schema, stored_total)
			updated_grade = _lookup_grade(grade_schema, stored_total, use_reexam=True)
			frappe.db.set_value(
				"Student Course Marks",
				scm_name,
				{"grade": grade, "updated_grade": updated_grade},
			)
			return {
				"total":               stored_total,
				"grade":               grade,
				"updated_final_marks": stored_total,
				"updated_grade":       updated_grade,
			}
		return {"total": None, "grade": "", "updated_final_marks": None, "updated_grade": ""}

	# ── Fetch calc settings from evaluation schema ────────────────────────────
	calc = frappe.db.get_value(
		"Evaluation Schema",
		eval_schema,
		["calc_higher_revaluation", "calc_higher_reexam"],
		as_dict=True,
	) or {}
	calc_higher_revaluation = int(calc.get("calc_higher_revaluation") or 0)
	calc_higher_reexam      = int(calc.get("calc_higher_reexam") or 0)

	# ── Schema Assessment Config (regular exams) ──────────────────────────────
	configs = frappe.db.sql(
		"""
		SELECT sac.component, sac.assessment_type, sac.maximum_marks, sac.effective_marks,
		       sac.label, eat.type_name
		FROM `tabSchema Assessment Config` sac
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = sac.assessment_type
		WHERE sac.parent = %(schema)s
		""",
		{"schema": eval_schema},
		as_dict=True,
	)
	if not configs:
		return {"total": None, "grade": "", "updated_final_marks": None, "updated_grade": ""}

	# ── Fetch all marks entries for this student ──────────────────────────────
	entries = frappe.db.sql(
		"""
		SELECT sme.component, sme.assessment_type,
		       sme.marks, sme.moderated_marks, sme.revaluation_marks
		FROM `tabStudent Marks Entry` sme
		WHERE sme.parent = %(scm)s
		""",
		{"scm": scm_name},
		as_dict=True,
	)
	entry_map = {}
	for e in entries:
		key = (e["component"] or "") + "|" + (e["assessment_type"] or "")
		entry_map[key] = e

	# ── Regular total ─────────────────────────────────────────────────────────
	total = 0.0
	for cfg in configs:
		key   = (cfg["component"] or "") + "|" + (cfg["assessment_type"] or "")
		e     = entry_map.get(key, {})
		# moderated_marks takes priority over raw marks
		base  = frappe.utils.flt(e.get("moderated_marks") or e.get("marks") or 0)
		is_project = (cfg.get("type_name") or cfg.get("label") or "").lower() == "project"
		if is_project:
			# For Project type, revaluation_marks stores the deduction (marks - deduction)
			deduction = frappe.utils.flt(e.get("revaluation_marks") or 0)
			raw_m = max(0.0, base - min(deduction, base))
		elif calc_higher_revaluation:
			reval = frappe.utils.flt(e.get("revaluation_marks") or 0)
			raw_m = max(base, reval)
		else:
			raw_m = base
		max_m = frappe.utils.flt(cfg["maximum_marks"] or 0)
		eff_m = frappe.utils.flt(cfg["effective_marks"] or 0)
		if max_m > 0 and eff_m > 0:
			total += raw_m * (eff_m / max_m)
		else:
			total += raw_m

	grade = _lookup_grade(grade_schema, total) if grade_schema else ""

	# ── Re-exam total ─────────────────────────────────────────────────────────
	reexam_configs = frappe.db.sql(
		"""
		SELECT src.component, src.assessment_type, src.maximum_marks, src.effective_marks
		FROM `tabSchema Reexam Config` src
		WHERE src.parent = %(schema)s
		""",
		{"schema": eval_schema},
		as_dict=True,
	)

	reexam_total = 0.0
	for cfg in (reexam_configs or []):
		key   = (cfg["component"] or "") + "|" + (cfg["assessment_type"] or "")
		e     = entry_map.get(key, {})
		base  = frappe.utils.flt(e.get("moderated_marks") or e.get("marks") or 0)
		if calc_higher_reexam:
			reval = frappe.utils.flt(e.get("revaluation_marks") or 0)
			raw_rx = max(base, reval)
		else:
			raw_rx = base
		max_rx = frappe.utils.flt(cfg["maximum_marks"] or 0)
		eff_rx = frappe.utils.flt(cfg["effective_marks"] or 0)
		if max_rx > 0 and eff_rx > 0:
			reexam_total += raw_rx * (eff_rx / max_rx)
		else:
			reexam_total += raw_rx

	updated_final_marks = max(total, reexam_total)
	# Updated grade uses Re Exam Composition when the schema is configured for it
	updated_grade = _lookup_grade(grade_schema, updated_final_marks, use_reexam=True) if grade_schema else ""

	frappe.db.set_value("Student Course Marks", scm_name, {
		"total_marks":        round(total, 2),
		"grade":              grade,
		"updated_final_marks": round(updated_final_marks, 2),
		"updated_grade":      updated_grade,
	})
	frappe.db.commit()
	return {
		"total":              round(total, 2),
		"grade":              grade,
		"updated_final_marks": round(updated_final_marks, 2),
		"updated_grade":      updated_grade,
	}


def _lookup_grade(grade_schema_name, total_marks, use_reexam=False):
	"""Return grade string for given total from a Grading Schema.

	Args:
		grade_schema_name: Name of the Grading Schema document.
		total_marks: The marks value to look up.
		use_reexam: When True, uses the Re Exam Composition table
		            (reexam_grades) if the schema has use_reexam_composition
		            enabled; otherwise falls back to the regular grades table.
	"""
	try:
		schema = frappe.get_doc("Grading Schema", grade_schema_name)
	except Exception:
		return ""
	total = frappe.utils.flt(total_marks)

	# Choose which grade table to use
	if use_reexam and schema.use_reexam_composition and schema.reexam_grades:
		grade_rows = schema.reexam_grades
	else:
		grade_rows = schema.grades

	for row in grade_rows:
		f       = frappe.utils.flt(row.marks_from)
		t       = frappe.utils.flt(row.marks_to)
		from_op = row.from_operator or ">="
		to_op   = row.to_operator   or "<"
		ok_from = (total >= f) if from_op == ">=" else (total > f)
		ok_to   = (total <= t) if to_op   == "<=" else (total < t)
		if ok_from and ok_to:
			return row.grade
	return ""


# ── Grade Auto-Generation ──────────────────────────────────────────────────────

@frappe.whitelist()
def auto_generate_grades(course, exam_plan, student_ids):
	"""Recalculate and persist grades for one or more students in a course.

	Iterates over each student, calling _recalculate_student_marks so that:
	  - moderated_marks / marks are used correctly (moderated takes priority),
	  - calc_higher_revaluation / calc_higher_reexam are honoured,
	  - both the regular grade and updated_grade (re-exam composition) are written.

	Args:
		course:      Course name.
		exam_plan:   Exam Plan name.
		student_ids: JSON list of Student Master names to process.

	Returns:
		dict mapping student_id → {total, grade, updated_final_marks, updated_grade}
		for every student that was successfully recalculated.
	"""
	import json as _json

	if isinstance(student_ids, str):
		student_ids = _json.loads(student_ids)

	if not student_ids or not course or not exam_plan:
		frappe.throw("course, exam_plan and student_ids are required.")

	# Fetch all SCM names in one query to avoid N individual lookups
	placeholders = ",".join(["%s"] * len(student_ids))
	scm_rows = frappe.db.sql(
		f"""
		SELECT name, student
		FROM `tabStudent Course Marks`
		WHERE course = %s AND exam_plan = %s
		  AND student IN ({placeholders})
		""",
		[course, exam_plan] + list(student_ids),
		as_dict=True,
	)
	scm_map = {r["student"]: r["name"] for r in scm_rows}

	results = {}
	for student_id in student_ids:
		scm_name = scm_map.get(student_id)
		if not scm_name:
			continue
		try:
			result = _recalculate_student_marks(scm_name, course, exam_plan)
			results[student_id] = result
		except Exception:
			frappe.log_error(frappe.get_traceback(), f"auto_generate_grades: {student_id}")

	return results


# ── Lock / Unlock ─────────────────────────────────────────────────────────────

@frappe.whitelist()
def toggle_lock(course, exam_plan):
	"""Toggle LOCKED / UNLOCKED status on the Access Result Settings record."""
	doc        = _get_or_create_access(exam_plan, course)
	new_status = "UNLOCKED" if doc.status == "LOCKED" else "LOCKED"
	frappe.db.set_value("Access Result Settings", doc.name, "status", new_status)
	frappe.db.commit()
	return {"status": new_status}


# ── Excel Import / Export ─────────────────────────────────────────────────────

@frappe.whitelist()
def export_marks_excel(course, exam_plan):
	"""Export component-wise marks to Excel. Returns file_url for download."""
	import io
	try:
		import openpyxl
		from openpyxl.styles import Font, PatternFill, Alignment
		from openpyxl.utils import get_column_letter
		from openpyxl.cell.text import InlineFont
		from openpyxl.cell.rich_text import TextBlock, CellRichText
	except ImportError:
		frappe.throw("openpyxl is required. Run: bench pip install openpyxl")

	csa_row = frappe.db.get_value(
		"Course Schema Assignment",
		{"course": course, "exam_plan": exam_plan},
		"evaluation_schema",
	)
	if not csa_row:
		frappe.throw("No evaluation schema found for this course / exam plan.")

	cols = frappe.db.sql(
		"""
		SELECT sac.component, ec.component_name, sac.assessment_type, eat.type_name,
		       sac.label, sac.maximum_marks, sac.idx
		FROM `tabSchema Assessment Config` sac
		LEFT JOIN `tabExam Component` ec     ON ec.name  = sac.component
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = sac.assessment_type
		WHERE sac.parent = %(schema)s
		ORDER BY sac.idx ASC
		""",
		{"schema": csa_row},
		as_dict=True,
	)

	reexam_cols = frappe.db.sql(
		"""
		SELECT src.component, ec.component_name, src.assessment_type, eat.type_name,
		       src.maximum_marks, src.idx
		FROM `tabSchema Reexam Config` src
		LEFT JOIN `tabExam Component` ec ON ec.name = src.component
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = src.assessment_type
		WHERE src.parent = %(schema)s
		ORDER BY src.idx ASC
		""",
		{"schema": csa_row},
		as_dict=True,
	)

	students = frappe.db.sql(
		"""
		SELECT scm.student, scm.name AS scm_name,
		       COALESCE(sm.registration_id, sm.name) AS registration_id,
		       sm.personal_email AS email_id,
		       CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
		       scm.total_marks, scm.grade, scm.enrollment_status,
		       scm.attendance_status, scm.mfa, scm.fairness_status, scm.consider_for_sgpa,
		       scm.remark, scm.updated_final_marks, scm.updated_grade
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
		ORDER BY sm.registration_id ASC
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)
	if not students:
		frappe.throw("No students found for this course / exam plan.")

	# Fetch Approved FA MFA Applications for this course
	fa_mfa_apps = frappe.db.sql(
		"""
		SELECT student
		FROM `tabFA MFA Application`
		WHERE course = %(course)s AND docstatus = 1 AND status = 'Approved'
		""",
		{"course": course},
		as_dict=True
	)
	approved_fa_mfa_students = {row["student"] for row in fa_mfa_apps}

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Marks"

	groups = []
	group_map = {}
	for col in cols:
		comp = col.get("component") or "__none__"
		if comp not in group_map:
			group_map[comp] = {
				"component": comp,
				"component_name": col.get("component_name") or comp,
				"cols": []
			}
			groups.append(group_map[comp])
		group_map[comp]["cols"].append(col)

	rxgroups = []
	rxgroup_map = {}
	for col in (reexam_cols or []):
		comp = col.get("component") or "__rx_none__"
		if comp not in rxgroup_map:
			rxgroup_map[comp] = {
				"component": comp,
				"component_name": col.get("component_name") or comp,
				"cols": []
			}
			rxgroups.append(rxgroup_map[comp])
		rxgroup_map[comp]["cols"].append(col)

	def _is_project(col):
		return (col.get("type_name") or col.get("label") or "").lower() == "project"

	row1, row2, row3 = [], [], []

	student_headers = ["S.No", "Name", "Student RegistrationId", "EmailId"]
	for h in student_headers:
		row1.append(h)
		row2.append("")
		row3.append("")

	# col_keys: list of (key_str, is_project) — one entry per assessment col
	col_keys = []
	for g in groups:
		# group span = sum of sub-cols per assessment (3 for project, 1 for others)
		g_span = sum(3 if _is_project(c) else 1 for c in g["cols"])
		row1.append(g["component_name"])
		for _ in range(g_span - 1):
			row1.append("")
		for col in g["cols"]:
			lbl  = col.get("label") or col.get("type_name") or col.get("assessment_type") or ""
			maxm = col.get("maximum_marks") or 0
			is_proj = _is_project(col)
			if is_proj:
				row2.extend([f"{lbl} (Max: {maxm})", "", ""])
				row3.extend(["Marks", "Deduction", "Total Marks"])
			else:
				row2.append(f"{lbl} (Max: {maxm})")
				row3.append("Marks")
			col_keys.append(((col["component"] or "") + "|" + (col["assessment_type"] or ""), is_proj))

	row1.extend(["Grade", ""])
	row2.extend(["Total Marks", "Grade"])
	row3.extend(["", ""])

	# Fairness Status removed from exported template
	status_headers = ["Enrollment Status", "Attendance Status", "MFA", "SGPA", "Remarks"]
	row1.append("Overall Status")
	row1.extend([""] * (len(status_headers) - 1))
	for h in status_headers:
		row2.append(h)
		row3.append("")

	reexam_keys = []
	for g in rxgroups:
		g_span = len(g["cols"])  # reexam always 1 sub-col
		row1.append(f'{g["component_name"]} (Re-Exam)')
		for _ in range(g_span - 1):
			row1.append("")
		for col in g["cols"]:
			lbl  = col.get("label") or col.get("type_name") or col.get("assessment_type") or ""
			maxm = col.get("maximum_marks") or 0
			row2.append(f"{lbl} (Max: {maxm})")
			row3.append("Marks")
			reexam_keys.append((col["component"] or "") + "|" + (col["assessment_type"] or ""))

	row1.extend(["Updated Final Result", ""])
	row2.extend(["Updated Final Marks", "Updated Grade"])
	row3.extend(["", ""])

	ws.append(row1)
	ws.append(row2)
	ws.append(row3)

	from openpyxl.styles import Border, Side
	hdr_font     = Font(bold=True, color="FFFFFF")
	readonly_font= Font(bold=True, color="FFFFFF", italic=True)
	c_align      = Alignment(horizontal="center", vertical="center", wrap_text=True)
	thin_border  = Border(
		left=Side(style='thin', color='CCCCCC'),
		right=Side(style='thin', color='CCCCCC'),
		top=Side(style='thin', color='CCCCCC'),
		bottom=Side(style='thin', color='CCCCCC')
	)

	fill_r1      = PatternFill("solid", fgColor="B24040")
	fill_r2      = PatternFill("solid", fgColor="C65959")
	fill_r3      = PatternFill("solid", fgColor="D97373")
	fill_readonly= PatternFill("solid", fgColor="8B7355")  # brownish for read-only cols

	ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
	ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
	ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3)
	ws.merge_cells(start_row=1, start_column=4, end_row=3, end_column=4)

	# Track which columns are read-only (Total Marks for project)
	readonly_cols = set()

	c_idx = 5
	for g in groups:
		g_span = sum(3 if _is_project(c) else 1 for c in g["cols"])
		if g_span > 1:
			ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + g_span - 1)
		for col in g["cols"]:
			is_proj = _is_project(col)
			col_span = 3 if is_proj else 1
			if col_span > 1:
				ws.merge_cells(start_row=2, start_column=c_idx, end_row=2, end_column=c_idx + col_span - 1)
			else:
				ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
			if is_proj:
				# "Total Marks" is the 3rd sub-col — mark as read-only
				readonly_cols.add(c_idx + 2)
			c_idx += col_span

	ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + 1)
	ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
	ws.merge_cells(start_row=2, start_column=c_idx + 1, end_row=3, end_column=c_idx + 1)
	c_idx += 2

	ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + len(status_headers) - 1)
	for _ in range(len(status_headers)):
		ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
		c_idx += 1

	for g in rxgroups:
		g_span = len(g["cols"])
		if g_span > 1:
			ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + g_span - 1)
		for _ in g["cols"]:
			ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
			c_idx += 1

	ws.merge_cells(start_row=1, start_column=c_idx, end_row=1, end_column=c_idx + 1)
	ws.merge_cells(start_row=2, start_column=c_idx, end_row=3, end_column=c_idx)
	ws.merge_cells(start_row=2, start_column=c_idx + 1, end_row=3, end_column=c_idx + 1)

	for r in range(1, 4):
		for c in range(1, len(row1) + 1):
			cell = ws.cell(row=r, column=c)
			cell.alignment = c_align
			cell.border = thin_border
			if c in readonly_cols:
				cell.fill = fill_readonly
				cell.font = readonly_font
			elif c <= 4:
				cell.fill = fill_r1
				cell.font = hdr_font
			elif r == 1:
				cell.fill = fill_r1
				cell.font = hdr_font
			elif r == 2:
				cell.fill = fill_r2
				cell.font = hdr_font
			else:
				cell.fill = fill_r3
				cell.font = hdr_font

	for c in range(5, len(row1) + 1):
		ws.column_dimensions[get_column_letter(c)].width = 18
	ws.column_dimensions["A"].width = 8
	ws.column_dimensions["B"].width = 30
	ws.column_dimensions["C"].width = 25
	ws.column_dimensions["D"].width = 35

	marks_data = frappe.db.sql(
		"""
		SELECT parent, component, assessment_type, marks, revaluation_marks
		FROM `tabStudent Marks Entry`
		WHERE parent IN (
			SELECT name FROM `tabStudent Course Marks`
			WHERE course = %(course)s AND exam_plan = %(exam_plan)s
		)
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)

	marks_map = {}
	for m in marks_data:
		key = (m["parent"] or "") + "|" + (m["component"] or "") + "|" + (m["assessment_type"] or "")
		marks_map[key] = m

	font_super = InlineFont(vertAlign='superscript')

	for ri, s in enumerate(students, 4):
		row_data = [
			ri - 3,
			s.get("student_name") or "",
			s.get("registration_id") or "",
			s.get("email_id") or "",
		]

		for key, is_proj in col_keys:
			m = marks_map.get((s.get("scm_name") or "") + "|" + key, {})
			marks_val = m.get("marks") if m.get("marks") is not None else ""
			if is_proj:
				deduct_val = m.get("revaluation_marks") if m.get("revaluation_marks") is not None else ""
				# Total Marks = marks - deduction (display only, not imported)
				if marks_val != "" and deduct_val != "":
					proj_total = max(0, float(marks_val) - min(float(deduct_val), float(marks_val)))
				elif marks_val != "":
					proj_total = float(marks_val)
				else:
					proj_total = ""
				row_data.extend([marks_val, deduct_val, proj_total])
			else:
				row_data.append(marks_val)

		mfa_val = "Yes" if s.get("student") in approved_fa_mfa_students else (s.get("mfa") or "No")
		grade_val = s.get("grade") or ""
		if mfa_val == "Yes":
			grade_val = CellRichText([grade_val + " ", TextBlock(font_super, "MFA")])

		row_data.extend([
			s.get("total_marks") if s.get("total_marks") is not None else "",
			grade_val,
			s.get("enrollment_status") or "",
			s.get("attendance_status") or "",
			mfa_val,
			"Yes" if s.get("consider_for_sgpa") else "No",
			s.get("remark") or "",
		])

		for key in reexam_keys:
			m = marks_map.get((s.get("scm_name") or "") + "|" + key, {})
			row_data.append(m.get("marks") if m.get("marks") is not None else "")

		ufm_grade_val = s.get("updated_grade") or ""
		if mfa_val == "Yes":
			ufm_grade_val = CellRichText([ufm_grade_val + " ", TextBlock(font_super, "MFA")])

		row_data.extend([
			s.get("updated_final_marks") if s.get("updated_final_marks") is not None else "",
			ufm_grade_val,
		])

		for ci, val in enumerate(row_data, 1):
			ws.cell(row=ri, column=ci, value=val)

	# Grey out read-only Total Marks cells in data rows
	readonly_fill = PatternFill("solid", fgColor="F5F0EB")
	for ri in range(4, len(students) + 4):
		for c in readonly_cols:
			ws.cell(row=ri, column=c).fill = readonly_fill

	buf  = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	course_code = frappe.db.get_value("Course", course, "course_code") or course
	fname       = f"marks_{course_code}_{exam_plan}.xlsx".replace(" ", "_")
	fdoc        = frappe.get_doc({
		"doctype":    "File",
		"file_name":  fname,
		"is_private": 1,
		"content":    buf.read(),
	})
	fdoc.save(ignore_permissions=True)
	return {"file_url": fdoc.file_url}


@frappe.whitelist()
def import_marks_excel(course, exam_plan, file_url):
	"""Import component-wise marks from Excel. Returns {updated, errors}."""
	import io
	try:
		import openpyxl
	except ImportError:
		frappe.throw("openpyxl is required. Run: bench pip install openpyxl")

	file_doc  = frappe.get_doc("File", {"file_url": file_url})
	file_path = file_doc.get_full_path()

	wb = openpyxl.load_workbook(file_path, data_only=True)
	ws = wb.active

	# ── Locate student identifier columns from row 1 ──────────────────────────
	row1_vals = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
	r1_norm   = [h.lower().replace(" ", "").replace("_", "") for h in row1_vals]
	reg_col   = next((i for i, h in enumerate(r1_norm) if "registrationid" in h), None)
	email_col = next((i for i, h in enumerate(r1_norm) if "emailid" in h or h == "email"), None)

	if reg_col is None and email_col is None:
		frappe.throw("'Student RegistrationId' or 'EmailId' column not found in the uploaded file.")

	# ── Get schema columns to build a label → (component, assessment_type) map ─
	csa_row = frappe.db.get_value(
		"Course Schema Assignment",
		{"course": course, "exam_plan": exam_plan},
		"evaluation_schema",
	) or ""
	if not csa_row:
		frappe.throw("No evaluation schema found.")

	schema_cols = frappe.db.sql(
		"""
		SELECT sac.component, sac.assessment_type, eat.type_name, sac.label, sac.maximum_marks
		FROM `tabSchema Assessment Config` sac
		LEFT JOIN `tabExam Assessment Type` eat ON eat.name = sac.assessment_type
		WHERE sac.parent = %(schema)s ORDER BY sac.idx ASC
		""",
		{"schema": csa_row},
		as_dict=True,
	)

	# Build label → schema col lookup (strip/normalise spaces around colon)
	label_to_col = {}
	for col in schema_cols:
		lbl  = col.get("label") or col.get("type_name") or col.get("assessment_type") or ""
		maxm = col.get("maximum_marks") or 0
		# Accept both "Max: N" and "Max:N" formats
		label_to_col[f"{lbl} (Max: {maxm})"] = col
		label_to_col[f"{lbl} (Max:{maxm})"]  = col

	# ── Build per-column mapping from rows 2 and 3 ────────────────────────────
	# col_field_map: col_index (0-based) → (component, assessment_type, db_field)
	# db_field is "marks" or "revaluation_marks"; "total_marks" cols are skipped
	#
	# NOTE: non-project columns have rows 2+3 merged in the exported file, so
	# ws.cell(3,c).value is None for those columns. We handle this by mapping the
	# column directly to "marks" as soon as we see the label in row 2.
	col_field_map = {}
	current_proj_info = None  # (component, assessment_type) — only set for project blocks
	for c in range(1, ws.max_column + 1):
		r2 = str(ws.cell(2, c).value or "").strip()
		r3 = str(ws.cell(3, c).value or "").strip().lower()

		if r2 and r2 in label_to_col:
			sc = label_to_col[r2]
			is_proj = (sc.get("type_name") or sc.get("label") or "").lower() == "project"
			if is_proj:
				# Project block: sub-columns are mapped via row 3 in subsequent cols
				current_proj_info = (sc["component"], sc["assessment_type"])
				# The first sub-col ("Marks") shares the same column as the label
				if r3 in ("marks", ""):
					col_field_map[c - 1] = (sc["component"], sc["assessment_type"], "marks")
			else:
				# Non-project: row2+row3 are merged — map this column directly to "marks"
				col_field_map[c - 1] = (sc["component"], sc["assessment_type"], "marks")
				current_proj_info = None
			continue

		# Continuation columns — only relevant inside a project block
		if current_proj_info is None:
			continue
		comp, atype = current_proj_info
		if r3 == "marks":
			col_field_map[c - 1] = (comp, atype, "marks")
		elif r3 == "deduction":
			col_field_map[c - 1] = (comp, atype, "revaluation_marks")
		elif r3 == "revaluation marks":
			col_field_map[c - 1] = (comp, atype, "revaluation_marks")
		elif r3 == "total marks":
			pass  # read-only, skip
		else:
			# Empty or unrecognised sub-col — end the project block
			current_proj_info = None

	if not col_field_map:
		frappe.throw("No importable marks columns found. Please use the exported template.")

	# ── Build student lookup maps ─────────────────────────────────────────────
	students = frappe.db.sql(
		"""
		SELECT scm.name AS scm_name, scm.student, sm.registration_id, sm.personal_email AS email_id
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)
	reg_map   = {s["registration_id"]: (s["scm_name"], s["student"]) for s in students if s["registration_id"]}
	email_map = {s["email_id"]:        (s["scm_name"], s["student"]) for s in students if s["email_id"]}

	updated = 0
	errors  = []

	# Data starts from row 4 (rows 1-3 are the 3-level headers)
	for row in ws.iter_rows(min_row=4, values_only=True):
		reg_id   = str(row[reg_col]).strip()   if reg_col   is not None and row[reg_col]   else ""
		email_id = str(row[email_col]).strip() if email_col is not None and row[email_col] else ""

		scm_name = _student = None
		if reg_id and reg_id in reg_map:
			scm_name, _student = reg_map[reg_id]
		elif email_id and email_id in email_map:
			scm_name, _student = email_map[email_id]

		if not scm_name:
			if reg_id or email_id:
				errors.append(f"Student '{reg_id or email_id}' not found.")
			continue

		row_changed = False

		for ci, (comp, atype, db_field) in col_field_map.items():
			if ci >= len(row):
				continue
			cell_val = row[ci]
			if cell_val is None or str(cell_val).strip() == "":
				continue
			try:
				fval = float(cell_val)
			except (ValueError, TypeError):
				errors.append(f"Invalid value '{cell_val}' for {reg_id} / col {ci + 1}.")
				continue

			sme_name = frappe.db.get_value(
				"Student Marks Entry",
				{"parent": scm_name, "component": comp, "assessment_type": atype},
				"name",
			)
			if sme_name:
				frappe.db.set_value("Student Marks Entry", sme_name, db_field, fval)
			else:
				frappe.db.sql(
					f"""
					INSERT INTO `tabStudent Marks Entry`
					(name, creation, modified, modified_by, owner,
					 parent, parenttype, parentfield, component, assessment_type, {db_field})
					VALUES (%(nm)s, NOW(), NOW(), %(usr)s, %(usr)s,
					        %(par)s, 'Student Course Marks', 'marks_entries',
					        %(comp)s, %(atype)s, %(val)s)
					""",
					{
						"nm":   frappe.generate_hash("", 10),
						"usr":  frappe.session.user,
						"par":  scm_name,
						"comp": comp,
						"atype": atype,
						"val":  fval,
					},
				)
			row_changed = True

		if row_changed:
			_recalculate_student_marks(scm_name, course, exam_plan)
			updated += 1

	frappe.db.commit()
	return {"updated": updated, "errors": errors}


@frappe.whitelist()
def export_reexam_template(course, exam_plan):
	"""Generate an Excel template for re-exam marks with student details (Registration ID, Name, Email)."""
	try:
		import openpyxl
		from openpyxl.styles import Font, Alignment, PatternFill
	except ImportError:
		frappe.throw("openpyxl is required. Run: bench pip install openpyxl")

	# Fetch students enrolled in this course
	students = frappe.db.sql(
		"""
		SELECT scm.student, sm.registration_id,
		       CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
		       sm.personal_email AS email_id,
		       scm.updated_final_marks, scm.updated_grade
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
		ORDER BY sm.registration_id ASC
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)

	# Create workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Re-Exam Marks"

	# Header styling
	header_font = Font(bold=True, color="FFFFFF", size=11)
	header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")

	# Define headers
	headers = ["Registration ID", "Student Name", "Email ID", "Re Exam Marks", "Current Marks", "Current Grade"]

	# Write headers
	for col_idx, header in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=col_idx)
		cell.value = header
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = header_alignment

	# Set column widths
	ws.column_dimensions['A'].width = 20  # Registration ID
	ws.column_dimensions['B'].width = 30  # Student Name
	ws.column_dimensions['C'].width = 35  # Email ID
	ws.column_dimensions['D'].width = 18  # Re Exam Marks
	ws.column_dimensions['E'].width = 15  # Current Marks
	ws.column_dimensions['F'].width = 15  # Current Grade

	# Write student data
	for row_idx, student in enumerate(students, start=2):
		ws.cell(row=row_idx, column=1).value = student.get("registration_id") or ""
		ws.cell(row=row_idx, column=2).value = student.get("student_name") or ""
		ws.cell(row=row_idx, column=3).value = student.get("email_id") or ""
		ws.cell(row=row_idx, column=4).value = ""  # Empty for user to fill
		ws.cell(row=row_idx, column=5).value = float(student.get("updated_final_marks") or 0) if student.get("updated_final_marks") else ""
		ws.cell(row=row_idx, column=6).value = student.get("updated_grade") or ""

		# Center align marks columns
		ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
		ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
		ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")

	# Freeze header row
	ws.freeze_panes = "A2"

	# Save file
	import io, os
	from frappe.utils.file_manager import save_file

	file_buffer = io.BytesIO()
	wb.save(file_buffer)
	file_buffer.seek(0)

	course_code = frappe.db.get_value("Course", course, "course_code") or course[:10]
	filename = f"ReExam_Template_{course_code}_{exam_plan[:10]}.xlsx"

	file_doc = save_file(
		filename,
		file_buffer.read(),
		"",
		"",
		is_private=1,
	)

	return {"file_url": file_doc.file_url}


@frappe.whitelist()
def import_reexam_marks_excel(course, exam_plan, file_url):
	"""Import re-exam marks from Excel. Updates the updated_final_marks field in Student Course Marks."""
	try:
		import openpyxl
	except ImportError:
		frappe.throw("openpyxl is required. Run: bench pip install openpyxl")

	file_doc = frappe.get_doc("File", {"file_url": file_url})
	file_path = file_doc.get_full_path()

	wb = openpyxl.load_workbook(file_path, data_only=True)
	ws = wb.active

	headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

	# Identify columns
	import_headers = [h.lower().replace(" ", "").replace("_", "") for h in headers]
	reg_col = next((i for i, h in enumerate(import_headers) if "registrationid" in h), None)
	email_col = next((i for i, h in enumerate(import_headers) if "emailid" in h or h == "email"), None)
	reexam_col = next((i for i, h in enumerate(import_headers) if "reexammarks" in h or "reexam" in h), None)

	if reg_col is None and email_col is None:
		frappe.throw("'Registration ID' or 'Email ID' column not found in the uploaded file.")

	if reexam_col is None:
		frappe.throw("'Re Exam Marks' column not found in the uploaded file.")

	# Build student map
	students = frappe.db.sql(
		"""
		SELECT scm.name AS scm_name, scm.student, sm.registration_id, sm.personal_email AS email_id
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)

	reg_map = {s["registration_id"]: s["scm_name"] for s in students if s["registration_id"]}
	email_map = {s["email_id"]: s["scm_name"] for s in students if s["email_id"]}

	updated = 0
	errors = []

	for row in ws.iter_rows(min_row=2, values_only=True):
		if not row or all(cell is None or str(cell).strip() == "" for cell in row):
			continue

		reg_id = str(row[reg_col]).strip() if reg_col is not None and row[reg_col] else ""
		email_id = str(row[email_col]).strip() if email_col is not None and row[email_col] else ""
		reexam_marks = row[reexam_col] if reexam_col is not None and reexam_col < len(row) else None

		# Skip if no re-exam marks provided
		if reexam_marks is None or str(reexam_marks).strip() == "":
			continue

		# Find student
		scm_name = None
		if reg_id and reg_id in reg_map:
			scm_name = reg_map[reg_id]
		elif email_id and email_id in email_map:
			scm_name = email_map[email_id]

		if not scm_name:
			if reg_id or email_id:
				errors.append(f"Student '{reg_id or email_id}' not found.")
			continue

		# Validate marks
		try:
			marks_val = float(reexam_marks)
		except (ValueError, TypeError):
			errors.append(f"Invalid re-exam marks '{reexam_marks}' for student '{reg_id or email_id}'.")
			continue

		# Update updated_final_marks field
		frappe.db.set_value("Student Course Marks", scm_name, "updated_final_marks", marks_val)

		# Also update the re-exam marks in Student Marks Entry (if exists)
		reexam_entry = frappe.db.get_value(
			"Student Marks Entry",
			{
				"parent": scm_name,
				"component": "Re exam",
				"assessment_type": ["like", "%Supplementary%"]
			},
			"name"
		)
		if reexam_entry:
			frappe.db.set_value("Student Marks Entry", reexam_entry, "marks", marks_val)

		# Recalculate grade based on new marks
		_recalculate_grade_for_reexam(scm_name, marks_val, course, exam_plan)
		updated += 1

	frappe.db.commit()
	return {"updated": updated, "errors": errors}


def _recalculate_grade_for_reexam(scm_name, final_marks, course, exam_plan):
	"""Recalculate and update the grade based on re-exam marks."""
	# Get grade schema
	grade_schema = frappe.db.get_value(
		"Course Schema Assignment",
		{"course": course, "exam_plan": exam_plan},
		"grade_schema",
	)

	if not grade_schema:
		return

	# Get grade components
	grade_components = frappe.db.sql(
		"""
		SELECT grade, marks_from, marks_to, from_operator, to_operator
		FROM `tabGrading Schema Component`
		WHERE parent = %(schema)s
		ORDER BY marks_from DESC
		""",
		{"schema": grade_schema},
		as_dict=True,
	)

	# Determine grade
	new_grade = ""
	for component in grade_components:
		marks_from = float(component.get("marks_from") or 0)
		marks_to = float(component.get("marks_to") or 100)
		from_op = component.get("from_operator") or ">="
		to_op = component.get("to_operator") or "<"

		# Check if marks fall within this grade range
		from_condition = False
		to_condition = False

		if from_op == ">=":
			from_condition = final_marks >= marks_from
		elif from_op == ">":
			from_condition = final_marks > marks_from
		else:
			from_condition = final_marks >= marks_from

		if to_op == "<=":
			to_condition = final_marks <= marks_to
		elif to_op == "<":
			to_condition = final_marks < marks_to
		else:
			to_condition = final_marks <= marks_to

		if from_condition and to_condition:
			new_grade = component.get("grade") or ""
			break

	# Update grade
	if new_grade:
		frappe.db.set_value("Student Course Marks", scm_name, "updated_grade", new_grade)


# ── Statistics ────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_course_statistics(course, exam_plan):
	"""Return pass/fail counts, grade distribution, avg marks, and topper for a course."""
	# Fetch ALL students (including those without marks yet) for total/graded/not_graded counts
	rows = frappe.db.sql(
		"""
		SELECT scm.grade, scm.total_marks,
		       CONCAT_WS(' ', sm.first_name, sm.last_name) AS student_name,
		       sm.registration_id
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)

	# Grade schema pass info
	csa = frappe.db.get_value(
		"Course Schema Assignment",
		{"course": course, "exam_plan": exam_plan},
		"grade_schema",
	)
	failed_grades = set()
	if csa:
		try:
			gs = frappe.get_doc("Grading Schema", csa)
			failed_grades = {r.grade for r in gs.grades if r.failed}
		except Exception:
			pass

	grade_dist = {}
	marks_with_values = []
	topper_row = None
	graded     = 0
	not_graded = 0

	for r in rows:
		g = (r["grade"] or "").strip()
		if g:
			graded += 1
			grade_dist[g] = grade_dist.get(g, 0) + 1
		else:
			not_graded += 1

		if r["total_marks"] is not None:
			m = frappe.utils.flt(r["total_marks"])
			marks_with_values.append(m)
			if topper_row is None or m > frappe.utils.flt(topper_row["total_marks"]):
				topper_row = r

	passed = sum(v for g, v in grade_dist.items() if g not in failed_grades)
	failed = sum(v for g, v in grade_dist.items() if g in failed_grades)

	avg = round(sum(marks_with_values) / len(marks_with_values), 2) if marks_with_values else 0

	return {
		"total":      len(rows),
		"graded":     graded,
		"not_graded": not_graded,
		"passed":     passed,
		"failed":     failed,
		"avg_marks":  avg,
		"topper":     {
			"name":  topper_row["student_name"]   if topper_row else "",
			"reg":   topper_row["registration_id"] if topper_row else "",
			"marks": topper_row["total_marks"]     if topper_row else 0,
		},
		"grade_dist": [{"grade": g, "count": c} for g, c in sorted(grade_dist.items(), key=lambda x: -x[1])],
	}


# ── Email Results ─────────────────────────────────────────────────────────────

@frappe.whitelist()
def send_results_email(course, exam_plan):
	"""Queue result emails to all students in the course/exam_plan."""
	rows = frappe.db.sql(
		"""
		SELECT scm.student, scm.total_marks, scm.grade,
		       sm.first_name, sm.last_name, sm.official_email_id,
		       c.course_name, ep.exam_name
		FROM `tabStudent Course Marks` scm
		LEFT JOIN `tabStudent Master` sm ON sm.name = scm.student
		LEFT JOIN `tabCourse`         c  ON c.name  = scm.course
		LEFT JOIN `tabExam Plan`      ep ON ep.name = scm.exam_plan
		WHERE scm.course = %(course)s AND scm.exam_plan = %(exam_plan)s
		  AND sm.official_email_id IS NOT NULL AND sm.official_email_id != ''
		""",
		{"course": course, "exam_plan": exam_plan},
		as_dict=True,
	)

	sent = 0
	for r in rows:
		student_name = " ".join(filter(None, [r["first_name"], r["last_name"]]))
		total        = f"{frappe.utils.flt(r['total_marks']):.2f}" if r["total_marks"] is not None else "N/A"
		grade        = r["grade"] or "N/A"
		course_name  = r["course_name"] or course
		exam_name    = r["exam_name"]   or exam_plan

		subject = f"Your Result: {course_name} — {exam_name}"
		message = f"""
		<p>Dear {student_name},</p>
		<p>Your examination result for <strong>{course_name}</strong> under <strong>{exam_name}</strong> has been published.</p>
		<table style="border-collapse:collapse;margin:10px 0;">
		  <tr><td style="padding:6px 12px;border:1px solid #e2e8f0;font-weight:600;">Course</td>
		      <td style="padding:6px 12px;border:1px solid #e2e8f0;">{course_name}</td></tr>
		  <tr><td style="padding:6px 12px;border:1px solid #e2e8f0;font-weight:600;">Exam Plan</td>
		      <td style="padding:6px 12px;border:1px solid #e2e8f0;">{exam_name}</td></tr>
		  <tr><td style="padding:6px 12px;border:1px solid #e2e8f0;font-weight:600;">Total Marks</td>
		      <td style="padding:6px 12px;border:1px solid #e2e8f0;">{total}</td></tr>
		  <tr><td style="padding:6px 12px;border:1px solid #e2e8f0;font-weight:600;">Grade</td>
		      <td style="padding:6px 12px;border:1px solid #e2e8f0;font-size:18px;font-weight:700;color:#4f46e5;">{grade}</td></tr>
		</table>
		<p>Please contact your faculty for any queries.</p>
		"""
		frappe.sendmail(
			recipients=[r["official_email_id"]],
			subject=subject,
			message=message,
			now=False,
		)
		sent += 1

	return {"sent": sent}
