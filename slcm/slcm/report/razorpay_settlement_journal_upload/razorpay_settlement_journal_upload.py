# Copyright (c) 2026, Azim Premji Foundation and contributors
# For license information, please see license.txt
"""
Razorpay Settlement Journal Upload Report

Architecture mirrors FLE Razorpay Settlement Report exactly:
  1. Fetch settlement list from /v1/settlements
  2. Fetch per-payment recon from /v1/settlements/recon/combined
  3. Build local pay_xxx map from FLE Payment Log (transaction_id + gateway_response)
  4. Filter recon items: only items whose entity_id (pay_xxx) exists in FLE Payment Log
  5. Group filtered recon items by settlement_id to get accurate gross amounts
  6. Generate one Debit + one Credit journal row per settlement
  7. Export in Zoho Books Journal Import format (14 columns)

Amount accuracy:
  - Gross amount = sum of individual payment amounts from MATCHED recon items only
  - This matches the FLE Settlement Report gross total exactly
  - Falls back to settlement.amount (net) only when recon is unavailable

Journal amounts:
  Credit entry → gross (student paid)
  Debit entry  → gross (student paid)
  Both sides equal → always balanced
"""

import io
import json
from datetime import datetime, timezone

import frappe
import requests
from frappe import _
from frappe.utils import flt, formatdate, getdate, nowdate

# ── Constants ─────────────────────────────────────────────────────────────────
RAZORPAY_BASE  = "https://api.razorpay.com/v1"
PAGE_SIZE      = 100
RECON_PG_SIZE  = 1000

DEFAULT_CREDIT_ACCOUNT = "Foundation for a Legal Education Fee"
DEFAULT_DEBIT_ACCOUNT  = "UBI Bank PACE 520101045120011"
DEFAULT_CONTACT        = "Razorpay"
DEFAULT_PREFIX         = "JN-FP-"
DEFAULT_JOURNAL_TYPE   = "Both"
DEFAULT_CURRENCY       = "INR"
DEFAULT_DESCRIPTION    = "online"
DEFAULT_DEPARTMENT     = "PACE"
DEFAULT_COURSE         = "FLE"

# Zoho Books required column order — do NOT change
ZOHO_HEADERS = [
    "Journal Date", "Reference Number", "Journal Number Prefix",
    "Journal Number Suffix", "Notes", "Journal Type", "Currency",
    "Account", "Description", "Contact Name", "Debit", "Credit",
    "Department", "Course",
]
ZOHO_FIELD_MAP = {
    "Journal Date":          "journal_date",
    "Reference Number":      "reference_number",
    "Journal Number Prefix": "journal_number_prefix",
    "Journal Number Suffix": "journal_number_suffix",
    "Notes":                 "notes",
    "Journal Type":          "journal_type",
    "Currency":              "currency",
    "Account":               "account",
    "Description":           "description",
    "Contact Name":          "contact_name",
    "Debit":                 "debit",
    "Credit":                "credit",
    "Department":            "department",
    "Course":                "course",
}


# ── Entry point ───────────────────────────────────────────────────────────────

def execute(filters=None):
    filters = filters or {}

    try:
        settlements = _fetch_settlements(filters)
    except frappe.ValidationError:
        raise
    except Exception as exc:
        frappe.log_error(frappe.get_traceback(), "Razorpay Settlement Journal Upload")
        frappe.throw(
            _("Failed to fetch settlements from Razorpay: {0}").format(str(exc)),
            title=_("API Error"),
        )
        return _get_columns(), []

    if not settlements:
        frappe.msgprint(
            _("No settlements found for the selected date range."),
            indicator="orange", alert=True,
        )
        return _get_columns(), [], None, None, []

    config = _resolve_config(filters)

    # Step 1: build local pay_xxx → FLE match map (identical to FLE Settlement Report)
    local_pay_map = _build_local_pay_map()

    # Step 2: fetch recon items once — shared for matching AND amount calculation
    # Pass from_date/to_date as None so _resolve_year_months derives months from
    # the settlements themselves — avoids missing recon items that fall slightly
    # outside the user's selected date range due to timezone differences.
    api_key, api_secret = _get_credentials()
    auth = (api_key, api_secret)
    recon_items = _fetch_recon_items(
        auth=auth,
        settlements=settlements,
        from_date=None,
        to_date=None,
    )

    # Step 3: determine which settlement_ids have FLE-matched payments
    # Uses same logic as FLE Settlement Report: check entity_id in local_pay_map
    fle_matched_sids, fle_matched_utrs = _resolve_fle_matched_settlements(
        recon_items, settlements, local_pay_map
    )
    fle_map = {"by_settlement_id": fle_matched_sids, "by_utr": fle_matched_utrs}

    # Step 4: calculate gross amount per settlement from MATCHED recon items only
    # This ensures Journal Upload amounts match FLE Settlement Report exactly
    gross_by_sid = _calc_gross_from_matched_recon(
        recon_items, settlements, fle_matched_sids, filters
    )

    data, stats = _build_journal_rows(
        settlements, filters, config, fle_map, gross_by_sid
    )

    return (
        _get_columns(),
        data,
        None,
        _get_chart(stats["daily"]),
        _get_report_summary(stats),
    )


# ── Columns ───────────────────────────────────────────────────────────────────

def _get_columns():
    return [
        {"label": _("Journal Date"),           "fieldname": "journal_date",          "fieldtype": "Date",     "width": 110},
        {"label": _("Reference Number"),        "fieldname": "reference_number",      "fieldtype": "Data",     "width": 215},
        {"label": _("Journal Number Prefix"),   "fieldname": "journal_number_prefix", "fieldtype": "Data",     "width": 145},
        {"label": _("Journal Number Suffix"),   "fieldname": "journal_number_suffix", "fieldtype": "Int",      "width": 145},
        {"label": _("Notes"),                   "fieldname": "notes",                 "fieldtype": "Data",     "width": 380},
        {"label": _("Journal Type"),            "fieldname": "journal_type",          "fieldtype": "Data",     "width":  85},
        {"label": _("Currency"),                "fieldname": "currency",              "fieldtype": "Data",     "width":  75},
        {"label": _("Account"),                 "fieldname": "account",               "fieldtype": "Data",     "width": 280},
        {"label": _("Description"),             "fieldname": "description",           "fieldtype": "Data",     "width":  85},
        {"label": _("Contact Name"),            "fieldname": "contact_name",          "fieldtype": "Data",     "width": 115},
        {"label": _("Debit"),                   "fieldname": "debit",                 "fieldtype": "Currency", "width": 145},
        {"label": _("Credit"),                  "fieldname": "credit",                "fieldtype": "Currency", "width": 145},
        {"label": _("Department"),              "fieldname": "department",            "fieldtype": "Data",     "width":  90},
        {"label": _("Course"),                  "fieldname": "course",                "fieldtype": "Data",     "width":  70},
        # Display-only (never exported to Zoho)
        {"label": _("Row Type"),                "fieldname": "row_type",              "fieldtype": "Data",     "width":  85},
        {"label": _("Settlement Status"),       "fieldname": "settlement_status",     "fieldtype": "Data",     "width": 120},
        {"label": _("Gross Amount (₹)"),        "fieldname": "gross_amount",          "fieldtype": "Currency", "width": 145},
        {"label": _("Gateway Fees (₹)"),        "fieldname": "gateway_fees",          "fieldtype": "Currency", "width": 125},
        {"label": _("GST on Fees (₹)"),         "fieldname": "gst_on_fees",           "fieldtype": "Currency", "width": 115},
        {"label": _("Net Settled (₹)"),         "fieldname": "net_settled",           "fieldtype": "Currency", "width": 135},
        {"label": _("Payment Count"),           "fieldname": "payment_count",         "fieldtype": "Int",      "width":  90},
        {"label": _("UTR"),                     "fieldname": "utr",                   "fieldtype": "Data",     "width": 180},
        {"label": _("FLE Match"),               "fieldname": "fle_match",             "fieldtype": "Data",     "width":  90},
    ]


# ── Dynamic config ────────────────────────────────────────────────────────────

def _resolve_config(filters):
    rzp_merchant_name = ""
    try:
        settings = frappe.get_doc("Razorpay Settings")
        rzp_merchant_name = (settings.get("merchant_account_name") or "").strip()
    except Exception:
        pass

    return {
        "bank_account":   (filters.get("bank_account")   or "").strip() or rzp_merchant_name or DEFAULT_DEBIT_ACCOUNT,
        "credit_account": (filters.get("credit_account") or "").strip() or DEFAULT_CREDIT_ACCOUNT,
        "prefix":         (filters.get("journal_prefix") or "").strip() or DEFAULT_PREFIX,
        "department":     (filters.get("department")     or "").strip() or DEFAULT_DEPARTMENT,
        "course":         (filters.get("course")         or "").strip() or DEFAULT_COURSE,
        "contact":        DEFAULT_CONTACT,
    }


# ── FLE Payment Log local pay_xxx map ─────────────────────────────────────────

def _build_local_pay_map():
    """
    Identical to FLE Razorpay Settlement Report's _build_local_map().
    Returns dict: pay_xxx → {contact_name, student_id, payment_method, payment_notes}

    Resolution order (same as FLE Settlement Report):
      1. transaction_id field (if starts with pay_)
      2. gateway_response → razorpay_payment_id  (Integration Request format)
      3. gateway_response → payload.payment.entity.id  (webhook format)
    """
    if not frappe.db.table_exists("FLE Payment Log"):
        return {}

    rows = frappe.db.sql(
        """
        SELECT
            transaction_id,
            full_name                AS contact_name,
            reference_no             AS student_id,
            gateway_response,
            account_number_or_upi_id AS upi_or_account
        FROM `tabFLE Payment Log`
        """,
        as_dict=True,
    )

    pay_map = {}
    for row in rows:
        pid = (row.transaction_id or "").strip()

        if not pid or not pid.startswith("pay_"):
            pid = _extract_razorpay_pid(row.gateway_response)

        if not pid:
            continue
        if pid in pay_map:
            continue

        method, notes = _parse_gateway_response(row.gateway_response, row.upi_or_account)
        pay_map[pid] = {
            "contact_name":   row.contact_name or "",
            "student_id":     row.student_id   or "",
            "payment_method": method,
            "payment_notes":  notes,
        }
    return pay_map


def _extract_razorpay_pid(gateway_response):
    if not gateway_response:
        return ""
    try:
        gw  = json.loads(gateway_response)
        pid = (gw.get("razorpay_payment_id") or "").strip()
        if pid and pid.startswith("pay_"):
            return pid
        pid = (
            gw.get("payload", {})
               .get("payment", {})
               .get("entity", {})
               .get("id") or ""
        ).strip()
        if pid and pid.startswith("pay_"):
            return pid
    except Exception:
        pass
    return ""


def _parse_gateway_response(gateway_response, upi_or_account):
    method = ""
    notes  = upi_or_account or ""
    if not gateway_response:
        return method, notes
    try:
        resp   = json.loads(gateway_response)
        entity = resp.get("payload", {}).get("payment", {}).get("entity", resp)
        method = entity.get("method") or ""
        parts  = [
            str(entity[k])
            for k in ("bank", "wallet", "vpa", "description")
            if entity.get(k)
        ]
        if parts:
            notes = " | ".join(parts)
    except Exception:
        pass
    return method.title() if method else "", notes


# ── Recon fetcher ─────────────────────────────────────────────────────────────

def _fetch_recon_items(auth, settlements, from_date=None, to_date=None):
    """
    Fetch per-payment recon from /v1/settlements/recon/combined.
    Identical to FLE Settlement Report's _fetch_combined_recon():
      - Resolves settlement_id from UTR when missing
      - Deduplicates by entity_id
      - Returns [] gracefully on 404 (recon not enabled)
    """
    if not settlements:
        return []

    target_sids = {s["id"] for s in settlements if s.get("id")}
    utr_to_sid  = {
        s["utr"]: s["id"]
        for s in settlements
        if s.get("utr") and s.get("id")
    }
    year_months = _resolve_year_months(from_date, to_date, settlements)

    seen_eids = set()
    all_items = []

    for year, month in year_months:
        skip = 0
        while True:
            try:
                resp = requests.get(
                    f"{RAZORPAY_BASE}/settlements/recon/combined",
                    auth=auth,
                    params={"year": year, "month": month,
                            "count": RECON_PG_SIZE, "skip": skip},
                    timeout=60,
                )
            except Exception:
                break

            if resp.status_code == 404:
                return []
            if not resp.ok:
                break

            items = resp.json().get("items") or []
            for item in items:
                sid = item.get("settlement_id") or ""
                if not sid:
                    utr = item.get("settlement_utr") or ""
                    sid = utr_to_sid.get(utr) or ""
                if sid not in target_sids:
                    continue
                if not item.get("settlement_id") and sid:
                    item["settlement_id"] = sid

                eid = (item.get("entity_id") or item.get("payment_id") or "").strip()
                if eid:
                    if eid in seen_eids:
                        continue
                    seen_eids.add(eid)

                all_items.append(item)

            if len(items) < RECON_PG_SIZE:
                break
            skip += RECON_PG_SIZE

    return all_items


def _resolve_year_months(from_date, to_date, settlements):
    year_months = set()
    if from_date and to_date:
        try:
            start = datetime.strptime(str(from_date), "%Y-%m-%d")
            end   = datetime.strptime(str(to_date),   "%Y-%m-%d")
            if start <= end:
                y, m = start.year, start.month
                while (y, m) <= (end.year, end.month):
                    year_months.add((y, m))
                    m += 1
                    if m > 12:
                        m = 1
                        y += 1
        except ValueError:
            pass
    if not year_months and settlements:
        for s in settlements:
            ts = s.get("created_at")
            if ts:
                try:
                    dt = datetime.fromtimestamp(int(ts), tz=timezone.utc)
                    year_months.add((dt.year, dt.month))
                except Exception:
                    pass
    return sorted(year_months)


# ── FLE settlement matching ───────────────────────────────────────────────────

def _resolve_fle_matched_settlements(recon_items, settlements, local_pay_map):
    """
    Determine which settlement_ids have at least one FLE-matched payment.
    Mirrors FLE Settlement Report: checks entity_id (pay_xxx) in local_pay_map.
    Also checks settlement_id / settlement_utr direct DB fields as fast path.

    Returns: (matched_sids: set, matched_utrs: set)
    """
    matched_sids = set()
    matched_utrs = set()

    # Fast path: direct DB fields already populated by sync
    if frappe.db.table_exists("FLE Payment Log"):
        try:
            db_rows = frappe.db.sql(
                """
                SELECT settlement_id, settlement_utr
                FROM `tabFLE Payment Log`
                WHERE (settlement_id  IS NOT NULL AND settlement_id  != '')
                   OR (settlement_utr IS NOT NULL AND settlement_utr != '')
                """,
                as_dict=True,
            )
            for r in db_rows:
                if r.settlement_id:
                    matched_sids.add(r.settlement_id)
                if r.settlement_utr:
                    matched_utrs.add(r.settlement_utr)
        except Exception:
            pass

    # Primary path: match entity_id (pay_xxx) against local_pay_map
    # Identical to FLE Settlement Report line 204: `if show_fle_only and not local: continue`
    if local_pay_map and recon_items:
        sid_to_s = {s["id"]: s for s in settlements if s.get("id")}
        for item in recon_items:
            sid    = item.get("settlement_id") or ""
            pay_id = (item.get("entity_id") or item.get("payment_id") or "").strip()

            if pay_id and pay_id in local_pay_map and sid:
                matched_sids.add(sid)
                utr = (sid_to_s.get(sid) or {}).get("utr") or ""
                if utr:
                    matched_utrs.add(utr)

    return matched_sids, matched_utrs


# ── Gross amount calculation (FLE-filtered) ───────────────────────────────────

def _calc_gross_from_matched_recon(recon_items, settlements, fle_matched_sids, filters):
    """
    Calculate gross collected amount per settlement from MATCHED recon items only.

    KEY DIFFERENCE from previous implementation:
    Only sums recon items whose settlement_id is in fle_matched_sids.
    This ensures amounts match FLE Settlement Report exactly.

    Also tracks per-settlement: payment_count, total_fees, total_tax, net_settled.

    Returns dict: settlement_id → {
        gross, fees, tax, net_settled, payment_count
    }
    """
    fle_only = True  # always restrict to FLE-matched settlements
    data     = {}

    for item in recon_items:
        sid = item.get("settlement_id") or ""
        if not sid:
            continue

        # When fle_only: only include items from FLE-matched settlements
        if fle_only and sid not in fle_matched_sids:
            continue

        # Exclude non-payment entity types (refunds, adjustments)
        entity_type = (item.get("type") or item.get("entity_type") or "").lower()
        if entity_type and entity_type not in ("payment", ""):
            continue

        amt_paise  = int(item.get("amount") or 0)
        fee_paise  = int(item.get("fee")    or item.get("fees") or 0)
        tax_paise  = int(item.get("tax")    or 0)

        if sid not in data:
            data[sid] = {"gross_paise": 0, "fees_paise": 0, "tax_paise": 0, "count": 0}

        data[sid]["gross_paise"] += amt_paise
        data[sid]["fees_paise"]  += fee_paise
        data[sid]["tax_paise"]   += tax_paise
        data[sid]["count"]       += 1

    result = {}
    for sid, d in data.items():
        gross      = round(d["gross_paise"] / 100, 2)
        fees       = round(d["fees_paise"]  / 100, 2)
        tax        = round(d["tax_paise"]   / 100, 2)
        net        = round(gross - fees - tax, 2)
        result[sid] = {
            "gross":         gross,
            "gateway_fees":  fees,
            "gst_on_fees":   tax,
            "net_settled":   net,
            "payment_count": d["count"],
        }

    # Fallback: settlements with no recon items get settlement.amount (net settled)
    for s in settlements:
        sid = s.get("id") or ""
        if sid and sid not in result:
            net = round(flt(s.get("amount") or 0) / 100, 2)
            result[sid] = {
                "gross":         net,   # best estimate when recon unavailable
                "gateway_fees":  round(flt(s.get("fees") or 0) / 100, 2),
                "gst_on_fees":   round(flt(s.get("tax")  or 0) / 100, 2),
                "net_settled":   net,
                "payment_count": 0,
            }

    return result


# ── Journal row builder ───────────────────────────────────────────────────────

def _build_journal_rows(settlements, filters, config, fle_map, gross_by_sid):
    status_f   = (filters.get("settlement_status") or "").strip().lower()
    sid_f      = (filters.get("settlement_id")     or "").strip().lower()
    min_amt    = flt(filters.get("min_amount") or 0)
    max_amt    = flt(filters.get("max_amount") or 0)
    row_type_f = (filters.get("row_type") or "").strip()
    fle_only   = True  # always filter to FLE-matched settlements only

    rows         = []
    suffix       = _next_suffix(config["prefix"])
    daily_totals = {}
    filtered_out = 0

    # Sort ascending by settlement date
    settlements.sort(key=lambda s: s.get("settlement_time") or s.get("created_at") or 0)

    for s in settlements:
        settlement_id     = s.get("id") or ""
        settlement_status = (s.get("status") or "").lower()
        utr               = (s.get("utr") or "").strip()

        ts = s.get("settlement_time") or s.get("created_at")
        settlement_date = _unix_to_date(ts)

        if not settlement_date or not settlement_id:
            continue

        # Date filter (client-side guard)
        if filters.get("from_date") and settlement_date < getdate(filters["from_date"]):
            continue
        if filters.get("to_date") and settlement_date > getdate(filters["to_date"]):
            continue

        # Status filter
        if status_f and status_f != "all" and settlement_status != status_f:
            filtered_out += 1
            continue

        # Settlement ID partial filter
        if sid_f and sid_f not in settlement_id.lower():
            filtered_out += 1
            continue

        # FLE match check
        fle_matched = (
            settlement_id in fle_map.get("by_settlement_id", set())
            or (utr and utr in fle_map.get("by_utr", set()))
        )
        if fle_only and not fle_matched:
            filtered_out += 1
            continue

        # Get recon-based amounts for this settlement
        recon_data = gross_by_sid.get(settlement_id) or {}
        amount     = recon_data.get("gross") or 0

        if amount <= 0:
            continue

        # Amount range filter (on gross amount)
        if min_amt and amount < min_amt:
            filtered_out += 1
            continue
        if max_amt and amount > max_amt:
            filtered_out += 1
            continue

        date_str = formatdate(settlement_date, "dd-MM-yyyy")
        notes = (
            f"Online payment settlement on {date_str} "
            f"for bank account {config['bank_account']}"
            + (f" | UTR: {utr}" if utr else "")
        )

        daily_totals[str(settlement_date)] = (
            daily_totals.get(str(settlement_date), 0) + amount
        )

        shared = {
            "journal_date":          settlement_date,
            "reference_number":      settlement_id,
            "journal_number_prefix": config["prefix"],
            "journal_number_suffix": suffix,
            "notes":                 notes,
            "journal_type":          DEFAULT_JOURNAL_TYPE,
            "currency":              DEFAULT_CURRENCY,
            "description":           DEFAULT_DESCRIPTION,
            "contact_name":          config["contact"],
            "department":            config["department"],
            "course":                config["course"],
            "settlement_status":     settlement_status,
            "gross_amount":          amount,
            "gateway_fees":          recon_data.get("gateway_fees", 0),
            "gst_on_fees":           recon_data.get("gst_on_fees",  0),
            "net_settled":           recon_data.get("net_settled",  amount),
            "payment_count":         recon_data.get("payment_count", 0),
            "utr":                   utr,
            "fle_match":             "Yes" if fle_matched else "No",
        }

        credit_row = {**shared, "account": config["credit_account"], "debit": 0,      "credit": amount, "row_type": "Credit"}
        debit_row  = {**shared, "account": config["bank_account"],   "debit": amount, "credit": 0,      "row_type": "Debit"}

        if not row_type_f or row_type_f == "All":
            rows.append(credit_row)
            rows.append(debit_row)
        elif row_type_f == "Credit":
            rows.append(credit_row)
        elif row_type_f == "Debit":
            rows.append(debit_row)

        suffix += 1

    rows.sort(key=lambda r: (r["journal_date"], r["journal_number_suffix"], r["row_type"]))

    total_debit  = round(sum(r["debit"]  for r in rows), 2)
    total_credit = round(sum(r["credit"] for r in rows), 2)
    total_gross  = round(sum(r["gross_amount"] for r in rows if r["row_type"] == "Debit"), 2)
    total_fees   = round(sum(r["gateway_fees"] for r in rows if r["row_type"] == "Debit"), 2)
    total_gst    = round(sum(r["gst_on_fees"]  for r in rows if r["row_type"] == "Debit"), 2)
    total_net    = round(sum(r["net_settled"]   for r in rows if r["row_type"] == "Debit"), 2)
    balanced     = abs(total_debit - total_credit) < 0.01

    stats = {
        "total_settlements": len({r["reference_number"] for r in rows}),
        "total_rows":        len(rows),
        "filtered_out":      filtered_out,
        "total_gross":       total_gross,
        "total_fees":        total_fees,
        "total_gst":         total_gst,
        "total_net":         total_net,
        "total_debit":       total_debit,
        "total_credit":      total_credit,
        "balanced":          balanced,
        "daily":             daily_totals,
    }
    return rows, stats


def _next_suffix(prefix):
    series_name = f"{prefix}.####"
    try:
        result = frappe.db.sql(
            "SELECT current FROM `tabSeries` WHERE name = %s", (series_name,)
        )
        return int(result[0][0]) + 1 if result else 1
    except Exception:
        return 1


# ── Summary cards ─────────────────────────────────────────────────────────────

def _get_report_summary(stats):
    balanced = stats.get("balanced", False)
    return [
        {
            "value":     stats["total_settlements"],
            "label":     _("Settlements"),
            "datatype":  "Int",
            "indicator": "Blue",
        },
        {
            "value":     stats["total_gross"],
            "label":     _("Gross Amount (₹)"),
            "datatype":  "Currency",
            "currency":  "INR",
            "indicator": "Green",
        },
        {
            "value":     stats["total_fees"],
            "label":     _("Gateway Fees (₹)"),
            "datatype":  "Currency",
            "currency":  "INR",
            "indicator": "Orange",
        },
        {
            "value":     stats["total_gst"],
            "label":     _("GST on Fees (₹)"),
            "datatype":  "Currency",
            "currency":  "INR",
            "indicator": "Orange",
        },
        {
            "value":     stats["total_net"],
            "label":     _("Net Settled (₹)"),
            "datatype":  "Currency",
            "currency":  "INR",
            "indicator": "Blue",
        },
        {
            "value":     "Balanced ✓" if balanced else "UNBALANCED ✗",
            "label":     _("Debit = Credit"),
            "datatype":  "Data",
            "indicator": "Green" if balanced else "Red",
        },
    ]


# ── Trend chart ───────────────────────────────────────────────────────────────

def _get_chart(daily_totals):
    if not daily_totals:
        return None
    sorted_days = sorted(daily_totals)
    return {
        "data": {
            "labels":   [formatdate(d, "dd MMM") for d in sorted_days],
            "datasets": [{"name": _("Gross Amount (₹)"), "values": [daily_totals[d] for d in sorted_days]}],
        },
        "type":        "bar",
        "colors":      ["#5e64ff"],
        "axisOptions": {"xIsSeries": True},
    }


# ── Razorpay API ──────────────────────────────────────────────────────────────

def _fetch_settlements(filters):
    api_key, api_secret = _get_credentials()
    auth    = (api_key, api_secret)
    from_ts = _date_to_unix(filters.get("from_date"), end_of_day=False)
    to_ts   = _date_to_unix(filters.get("to_date"),   end_of_day=True)

    settlements = []
    skip = 0

    while True:
        params = {"count": PAGE_SIZE, "skip": skip}
        if from_ts:
            params["from"] = from_ts
        if to_ts:
            params["to"] = to_ts

        resp = requests.get(f"{RAZORPAY_BASE}/settlements", auth=auth, params=params, timeout=30)
        _raise_for_status(resp)

        items = resp.json().get("items") or []
        settlements.extend(items)

        if len(items) < PAGE_SIZE:
            break
        skip += PAGE_SIZE

    return settlements


def _get_credentials():
    try:
        settings = frappe.get_doc("Razorpay Settings")
        key    = settings.api_key
        secret = settings.get_password("api_secret")
        if key and secret:
            return key, secret
    except Exception:
        pass

    key = secret = None
    for k_attr in ("razorpay_api_key", "razorpay_key_id"):
        key = frappe.conf.get(k_attr)
        if key:
            break
    for s_attr in ("razorpay_api_secret", "razorpay_key_secret"):
        secret = frappe.conf.get(s_attr)
        if secret:
            break

    if key and secret:
        return key, secret

    frappe.throw(
        _(
            "Razorpay API credentials not found. "
            "Configure them in <b>Razorpay Settings</b> or in "
            "<code>site_config.json</code> as "
            "<code>razorpay_api_key</code> / <code>razorpay_api_secret</code>."
        ),
        title=_("Missing Configuration"),
    )


def _raise_for_status(resp):
    if resp.status_code == 401:
        frappe.throw(
            _("Razorpay authentication failed. Check your API Key and Secret in <b>Razorpay Settings</b>."),
            title=_("Authentication Error"),
        )
    if not resp.ok:
        try:
            err = resp.json().get("error", {}).get("description", resp.text[:300])
        except Exception:
            err = resp.text[:300]
        frappe.throw(
            _("Razorpay API error {0}: {1}").format(resp.status_code, err),
            title=_("API Error"),
        )


# ── Utilities ─────────────────────────────────────────────────────────────────

def _unix_to_date(ts):
    if not ts:
        return None
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).date()
    except Exception:
        return None


def _date_to_unix(date_str, end_of_day=False):
    if not date_str:
        return None
    try:
        from datetime import time as dtime
        d = getdate(date_str)
        t = dtime(23, 59, 59) if end_of_day else dtime(0, 0, 0)
        return int(datetime.combine(d, t).replace(tzinfo=timezone.utc).timestamp())
    except Exception:
        return None


def _format_zoho_date(val):
    try:
        return formatdate(val, "dd-MM-yyyy")
    except Exception:
        return str(val) if val else ""


def _cell(value, header):
    if value is None:
        return ""
    if header == "Journal Date":
        return _format_zoho_date(value)
    if header in ("Debit", "Credit"):
        v = flt(value)
        return v if v else ""
    return str(value) if value else ""


# ── Zoho Books export ─────────────────────────────────────────────────────────

@frappe.whitelist()
def download_zoho_upload_file(filters=None, file_format="csv"):
    """
    Generate Zoho Books–compatible journal upload file (14 columns only).
    Hard-blocks export if Debit ≠ Credit.
    """
    import base64

    if isinstance(filters, str):
        try:
            filters = json.loads(filters)
        except Exception:
            filters = {}
    filters = filters or {}

    settlements = _fetch_settlements(filters)
    if not settlements:
        frappe.throw(_("No settlements found for the selected date range."))

    config        = _resolve_config(filters)
    local_pay_map = _build_local_pay_map()
    api_key, api_secret = _get_credentials()
    recon_items   = _fetch_recon_items(
        auth=(api_key, api_secret),
        settlements=settlements,
        from_date=None,
        to_date=None,
    )
    fle_matched_sids, fle_matched_utrs = _resolve_fle_matched_settlements(
        recon_items, settlements, local_pay_map
    )
    fle_map      = {"by_settlement_id": fle_matched_sids, "by_utr": fle_matched_utrs}
    gross_by_sid = _calc_gross_from_matched_recon(
        recon_items, settlements, fle_matched_sids, filters
    )
    rows, stats = _build_journal_rows(settlements, filters, config, fle_map, gross_by_sid)

    if not rows:
        frappe.throw(_("No journal rows matched the applied filters."))

    if not stats["balanced"]:
        diff = round(stats["total_debit"] - stats["total_credit"], 2)
        frappe.throw(
            _(
                "Export blocked — journal is not balanced.<br>"
                "Total Debit: ₹{0} | Total Credit: ₹{1} | Difference: ₹{2}"
            ).format(
                f"{stats['total_debit']:,.2f}",
                f"{stats['total_credit']:,.2f}",
                f"{diff:,.2f}",
            ),
            title=_("Validation Failed — Unbalanced Journal"),
        )

    from_d     = filters.get("from_date", "")
    to_d       = filters.get("to_date", nowdate())
    date_label = f"{from_d}_{to_d}".strip("_") or nowdate()

    if file_format == "xlsx":
        content, filename, mime = _build_xlsx(rows, date_label, stats, config)
    else:
        content, filename, mime = _build_csv(rows, date_label)

    return {
        "filename":     filename,
        "content":      base64.b64encode(content).decode("utf-8"),
        "mime":         mime,
        "row_count":    len(rows),
        "balanced":     stats["balanced"],
        "total_debit":  stats["total_debit"],
        "total_credit": stats["total_credit"],
    }


@frappe.whitelist()
def get_dynamic_defaults():
    bank_account = DEFAULT_DEBIT_ACCOUNT
    try:
        settings = frappe.get_doc("Razorpay Settings")
        if settings.get("merchant_account_name"):
            bank_account = settings.merchant_account_name
    except Exception:
        pass
    return {
        "bank_account":   bank_account,
        "credit_account": DEFAULT_CREDIT_ACCOUNT,
        "journal_prefix": DEFAULT_PREFIX,
        "department":     DEFAULT_DEPARTMENT,
        "course":         DEFAULT_COURSE,
    }


# ── CSV builder ───────────────────────────────────────────────────────────────

def _build_csv(rows, date_label):
    import csv
    buf    = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(ZOHO_HEADERS)
    for r in rows:
        writer.writerow([_cell(r.get(ZOHO_FIELD_MAP[h]), h) for h in ZOHO_HEADERS])
    content  = buf.getvalue().encode("utf-8-sig")
    filename = f"zoho_journal_{date_label}.csv"
    return content, filename, "text/csv"


# ── XLSX builder ──────────────────────────────────────────────────────────────

def _build_xlsx(rows, date_label, stats, config):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        frappe.throw(_("openpyxl is not installed. Run: bench pip install openpyxl"))

    PURPLE    = "5E64FF"
    INDIGO    = "3949AB"
    WHITE     = "FFFFFF"
    GREEN_BG  = "E8F5E9"
    BLUE_BG   = "E3F2FD"
    STRIPE    = "F7F8FF"
    TOTAL_BG  = "E8EAF6"
    OK_CLR    = "2E7D32"
    ERR_CLR   = "C62828"
    BD        = "C5CAE9"

    thin_s  = Side(style="thin",   color=BD)
    thick_s = Side(style="medium", color=PURPLE)
    t_bdr   = Border(left=thin_s,  right=thin_s,  top=thin_s,  bottom=thin_s)
    h_bdr   = Border(left=thick_s, right=thick_s, top=thick_s, bottom=thick_s)

    n_cols    = len(ZOHO_HEADERS)
    debit_col  = ZOHO_HEADERS.index("Debit")  + 1
    credit_col = ZOHO_HEADERS.index("Credit") + 1
    balanced   = stats.get("balanced", False)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Journal Upload ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Journal Upload"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    # Title row
    ws.row_dimensions[1].height = 30
    ws.append([""] * n_cols)
    dept   = config.get("department") or DEFAULT_DEPARTMENT
    course = config.get("course")     or DEFAULT_COURSE
    bank   = config.get("bank_account") or DEFAULT_DEBIT_ACCOUNT
    tc = ws.cell(row=1, column=1,
        value=f"Razorpay Settlement Journal  |  Zoho Books Import  |  {date_label}  |  {dept}/{course}  |  Bank: {bank}")
    tc.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
    tc.fill      = PatternFill("solid", fgColor=PURPLE)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Header row
    ws.row_dimensions[2].height = 34
    ws.append(ZOHO_HEADERS)
    for ci, h in enumerate(ZOHO_HEADERS, 1):
        cell = ws.cell(row=2, column=ci)
        cell.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=INDIGO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = h_bdr

    # Data rows
    for ri, r in enumerate(rows, 3):
        rt = r.get("row_type", "")
        bg = PatternFill("solid", fgColor=GREEN_BG if rt == "Credit"
                          else BLUE_BG if rt == "Debit"
                          else (STRIPE if ri % 2 == 0 else WHITE))
        ws.row_dimensions[ri].height = 16

        for ci, h in enumerate(ZOHO_HEADERS, 1):
            val  = _cell(r.get(ZOHO_FIELD_MAP[h]), h)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = bg; cell.border = t_bdr

            if h in ("Debit", "Credit"):
                cell.font          = Font(bold=bool(val), size=9, name="Calibri",
                                          color="1565C0" if h == "Debit" else "2E7D32")
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            elif h == "Journal Date":
                cell.font      = Font(size=9, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif h in ("Journal Number Suffix", "Journal Type", "Currency",
                       "Department", "Course"):
                cell.font      = Font(size=9, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif h == "Account":
                clr = "6A1B9A" if rt == "Credit" else "1565C0"
                cell.font      = Font(size=9, name="Calibri", color=clr)
                cell.alignment = Alignment(vertical="center")
            else:
                cell.font      = Font(size=9, name="Calibri")
                cell.alignment = Alignment(vertical="center")

    # Total row
    total_ri = len(rows) + 3
    ws.row_dimensions[total_ri].height = 22
    for ci in range(1, n_cols + 1):
        cell = ws.cell(row=total_ri, column=ci)
        cell.fill      = PatternFill("solid", fgColor=TOTAL_BG)
        cell.font      = Font(bold=True, size=10, name="Calibri", color="1A237E")
        cell.border    = h_bdr
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if ci == 1:
            cell.value     = "TOTAL"
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        elif ci == debit_col:
            cell.value = stats["total_debit"]; cell.number_format = "#,##0.00"
        elif ci == credit_col:
            cell.value = stats["total_credit"]; cell.number_format = "#,##0.00"

    # Balance row
    bal_ri = total_ri + 1
    ws.row_dimensions[bal_ri].height = 20
    bc = ws.cell(row=bal_ri, column=1,
        value="✓  Debit = Credit — Balanced. Ready for Zoho Books import." if balanced
              else "✗  UNBALANCED — Do NOT import.")
    bc.font      = Font(bold=True, size=10, color=OK_CLR if balanced else ERR_CLR, name="Calibri")
    bc.fill      = PatternFill("solid", fgColor="E8F5E9" if balanced else "FFEBEE")
    bc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=bal_ri, start_column=1, end_row=bal_ri, end_column=n_cols)

    # Auto-size columns
    for ci, h in enumerate(ZOHO_HEADERS, 1):
        vals   = [str(_cell(r.get(ZOHO_FIELD_MAP[h]), h) or "") for r in rows]
        maxlen = max(len(str(h)), *(len(v) for v in vals)) if vals else len(str(h))
        ws.column_dimensions[get_column_letter(ci)].width = min(maxlen + 3, 50)

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 26

    s_hdr = Font(bold=True, color=WHITE, size=10, name="Calibri")
    s_hfill = PatternFill("solid", fgColor=PURPLE)
    s_lbl = Font(bold=True, size=10, name="Calibri", color="424242")
    s_val = Font(size=10, name="Calibri")
    s_ok  = Font(bold=True, size=10, name="Calibri", color=OK_CLR)
    s_err = Font(bold=True, size=10, name="Calibri", color=ERR_CLR)
    s_bdr = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
    )

    summary_rows = [
        ("Report Details", ""),
        ("Report Name",           "Razorpay Settlement Journal Upload"),
        ("Date Range",            date_label),
        ("Generated On",          nowdate()),
        ("", ""),
        ("Journal Configuration", ""),
        ("Credit Account",        config.get("credit_account") or DEFAULT_CREDIT_ACCOUNT),
        ("Bank Account (Debit)",  config.get("bank_account")   or DEFAULT_DEBIT_ACCOUNT),
        ("Journal Prefix",        config.get("prefix")         or DEFAULT_PREFIX),
        ("Department",            config.get("department")     or DEFAULT_DEPARTMENT),
        ("Course",                config.get("course")         or DEFAULT_COURSE),
        ("", ""),
        ("Settlement Statistics", ""),
        ("Total Settlements",     stats["total_settlements"]),
        ("Total Journal Rows",    stats["total_rows"]),
        ("", ""),
        ("Amounts (INR)", ""),
        ("Gross Amount (Student Payments)", stats["total_gross"]),
        ("Gateway Fees",                   stats["total_fees"]),
        ("GST on Fees",                    stats["total_gst"]),
        ("Net Settled (Bank Credit)",      stats["total_net"]),
        ("", ""),
        ("Journal Totals", ""),
        ("Total Debit",                    stats["total_debit"]),
        ("Total Credit",                   stats["total_credit"]),
        ("Difference (Debit − Credit)",    round(stats["total_debit"] - stats["total_credit"], 2)),
        ("", ""),
        ("Validation", ""),
        ("Debit = Credit",   "YES — Balanced ✓" if balanced else "NO — UNBALANCED ✗"),
        ("Ready for Import", "YES" if balanced else "NO — Fix before importing"),
    ]

    ws2.row_dimensions[1].height = 28
    t2 = ws2.cell(row=1, column=1, value="Zoho Books Journal Upload — Summary")
    t2.font = Font(bold=True, size=13, color=PURPLE, name="Calibri")
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells("A1:B1")

    for ri, (label, value) in enumerate(summary_rows, 2):
        ws2.row_dimensions[ri].height = 18
        cl = ws2.cell(row=ri, column=1, value=label)
        cv = ws2.cell(row=ri, column=2, value=value)
        if not label:
            continue
        if value == "":
            cl.font = s_hdr; cl.fill = s_hfill
            cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws2.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=2)
            continue
        cl.font = s_lbl; cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cl.border = s_bdr; cv.border = s_bdr
        if label in ("Gross Amount (Student Payments)", "Gateway Fees", "GST on Fees",
                     "Net Settled (Bank Credit)", "Total Debit", "Total Credit",
                     "Difference (Debit − Credit)"):
            cv.number_format = "#,##0.00"; cv.font = s_val
        elif label in ("Debit = Credit", "Ready for Import"):
            cv.font = s_ok if balanced else s_err
        else:
            cv.font = s_val
        cv.alignment = Alignment(
            horizontal="right" if isinstance(value, (int, float)) else "left",
            vertical="center",
        )

    # ── Sheet 3: Instructions ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Import Instructions")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 90

    steps = [
        ("Zoho Books Journal Import — Step-by-step Guide", True),
        ("", False),
        ("Step 1 — Verify data in the 'Journal Upload' sheet.", False),
        ("Step 2 — Check 'Summary' sheet: Debit = Credit must show YES.", False),
        ("Step 3 — In Zoho Books: Accountant → Journal → ⋮ → Import Journals.", False),
        ("Step 4 — Upload this XLSX file (or the CSV version).", False),
        ("Step 5 — Map columns if prompted, preview and confirm.", False),
        ("", False),
        ("Amount Notes", True),
        ("Gross Amount = sum of individual student payments (what students paid)", False),
        ("Gateway Fees = Razorpay transaction charges", False),
        ("GST on Fees  = GST on Razorpay charges", False),
        ("Net Settled  = Gross − Fees − GST (actual bank credit)", False),
        ("Journal Debit/Credit use Gross Amount — both sides are equal", False),
        ("", False),
        ("Column Reference", True),
        ("Journal Date          — dd-MM-yyyy  (e.g. 21-02-2026)", False),
        ("Reference Number      — Razorpay Settlement ID  (setl_xxx)", False),
        ("Journal Number Prefix — Configured prefix  (default: JN-FP-)", False),
        ("Journal Number Suffix — Auto-incremented integer", False),
        ("Notes                 — Includes bank account name and UTR", False),
        ("Account               — Must exactly match Zoho Books chart of accounts", False),
        ("Debit / Credit        — Only one value per row; other is blank", False),
    ]

    for ri, (text, heading) in enumerate(steps, 1):
        ws3.row_dimensions[ri].height = 18
        cell = ws3.cell(row=ri, column=1, value=text)
        cell.font = (
            Font(bold=True, size=11, color=PURPLE, name="Calibri") if heading
            else Font(size=10, name="Calibri", color="424242")
        )
        cell.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        f"zoho_journal_{date_label}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
